from datetime import date as dt_date, datetime

import json
import base64
import hashlib
import os
import re
import unicodedata
from io import BytesIO
from uuid import uuid4

from flask import current_app, jsonify, render_template, request, send_file, url_for
from flask import g
from flask_login import login_required
from sqlalchemy import func, or_
from sqlalchemy.orm import joinedload
from typing import Optional
from werkzeug.utils import secure_filename

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app import db
from app.models import Category, Expense, FileAsset, InventoryLot, InventoryMovement, Product, Sale, Supplier
from app.files.storage import upload_to_r2_and_create_asset
from app.permissions import module_required, module_required_any
from app.tenancy import ensure_request_context
from app.inventory import bp


_CODIGO_INTERNO_MAX_LEN = 8
_CODIGO_INTERNO_AUTO_LEN = 8
_CODIGO_INTERNO_PATTERN = re.compile(r'^[A-Za-z0-9]{8}$')


def _company_id() -> str:
    try:
        return str(getattr(g, 'company_id', '') or '').strip()
    except Exception:
        return ''


@bp.route('/')
@bp.route('/index')
@login_required
@module_required('inventory')
def index():
    """Inventario avanzado."""
    return render_template('inventory/index.html', title='Inventario')


def _parse_date_iso(raw, fallback=None):
    try:
        return dt_date.fromisoformat(str(raw).strip())
    except Exception:
        return fallback


def _parse_date_flexible(raw, fallback=None):
    s = str(raw or '').strip()
    if not s:
        return fallback
    # Accept ISO
    try:
        return dt_date.fromisoformat(s)
    except Exception:
        pass
    # Accept dd/mm/yyyy
    try:
        parts = s.split('/')
        if len(parts) == 3:
            dd = int(parts[0])
            mm = int(parts[1])
            yy = int(parts[2])
            return dt_date(yy, mm, dd)
    except Exception:
        pass
    return fallback


def _num(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _serialize_category(cat: Optional[Category]):
    if not cat:
        return None
    return {
        'id': cat.id,
        'name': cat.name,
        'parent_id': cat.parent_id,
        'active': bool(getattr(cat, 'active', True)),
    }


def _image_url(p: Product):
    file_id = str(getattr(p, 'image_file_id', '') or '').strip()
    if file_id:
        try:
            return url_for('files.download_file_api', file_id=file_id)
        except Exception:
            return ''
    filename = str(getattr(p, 'image_filename', '') or '').strip()
    if not filename:
        return ''
    try:
        return url_for('static', filename=f'uploads/{filename}')
    except Exception:
        return ''


def _serialize_product(p: Product):
    cat_obj = _serialize_category(p.category) if p.category else None
    cat_name = ''
    try:
        cat_name = str((cat_obj or {}).get('name') or '').strip()
    except Exception:
        cat_name = ''
    return {
        'id': p.id,
        'name': p.name,
        'description': p.description or '',
        'category_id': p.category_id,
        'category': cat_name,
        'category_obj': cat_obj,
        'sale_price': p.sale_price,
        'codigo_interno': p.internal_code or '',
        'internal_code': p.internal_code or '',
        'barcode': p.barcode or '',
        'image_url': _image_url(p),
        'unit_name': (p.unit_name or 'Unidad'),
        'uses_lots': bool(p.uses_lots),
        'method': p.method or 'FIFO',
        'min_stock': p.min_stock,
        'reorder_point': getattr(p, 'reorder_point', 0.0) or 0.0,
        'primary_supplier_id': str(getattr(p, 'primary_supplier_id', '') or ''),
        'primary_supplier_name': getattr(p, 'primary_supplier_name', '') or '',
        'active': bool(p.active),
    }


def _normalize_name(name: str) -> str:
    raw = str(name or '')
    try:
        raw = unicodedata.normalize('NFKD', raw)
        raw = ''.join([c for c in raw if not unicodedata.combining(c)])
    except Exception:
        raw = str(name or '')
    raw = raw.strip().lower()
    raw = ' '.join(raw.split())
    return raw


def _normalize_header(name: str) -> str:
    key = _normalize_name(name)
    key = key.replace('-', ' ').replace('/', ' ').replace('.', ' ')
    key = '_'.join([p for p in key.split(' ') if p])
    key = key.replace('__', '_')
    return key


def _find_category_by_norm(company_id: str, norm: str):
    cid = str(company_id or '').strip()
    if not cid:
        return None
    target = _normalize_name(norm)
    if not target:
        return None
    rows = (
        db.session.query(Category)
        .filter(Category.company_id == cid, Category.parent_id == None)  # noqa: E711
        .order_by(Category.name.asc())
        .limit(10000)
        .all()
    )
    for r in rows:
        if _normalize_name(getattr(r, 'name', '') or '') == target:
            return r
    return None


def _get_or_create_category(company_id: str, name: str):
    cid = str(company_id or '').strip()
    raw = str(name or '').strip()
    if not cid or not raw:
        return None, False
    existing = _find_category_by_norm(cid, raw)
    if existing:
        return existing, False
    row = Category(company_id=cid, name=raw, active=True)
    db.session.add(row)
    db.session.flush()
    return row, True


def _find_supplier_by_norm(company_id: str, norm: str):
    cid = str(company_id or '').strip()
    if not cid:
        return None
    target = _normalize_name(norm)
    if not target:
        return None
    rows = (
        db.session.query(Supplier)
        .filter(Supplier.company_id == cid)
        .order_by(Supplier.updated_at.desc(), Supplier.created_at.desc())
        .limit(10000)
        .all()
    )
    for r in rows:
        if _normalize_name(getattr(r, 'name', '') or '') == target:
            return r
    return None


def _get_or_create_supplier(company_id: str, name: str):
    cid = str(company_id or '').strip()
    raw = str(name or '').strip()
    if not cid or not raw:
        return None, False
    existing = _find_supplier_by_norm(cid, raw)
    if existing:
        return existing, False
    sid = uuid4().hex
    row = Supplier(
        id=sid,
        company_id=cid,
        name=raw,
        supplier_type='Inventory',
        status='Active',
        categories_json=json.dumps(['Inventario'], ensure_ascii=False),
    )
    db.session.add(row)
    db.session.flush()
    return row, True


def _product_method_locked(product_id: int) -> bool:
    has_lots = db.session.query(InventoryLot.id).filter(InventoryLot.product_id == int(product_id)).first() is not None
    if has_lots:
        return True
    has_movements = db.session.query(InventoryMovement.id).filter(InventoryMovement.product_id == int(product_id)).first() is not None
    return bool(has_movements)


def _generate_unique_internal_code() -> str:
    # Legacy helper (kept for backward compatibility): generate an 8-char code.
    for _ in range(30):
        code = uuid4().hex[:_CODIGO_INTERNO_AUTO_LEN].upper()
        exists = db.session.query(Product.id).filter(Product.internal_code == code).first() is not None
        if not exists:
            return code
    return uuid4().hex[:_CODIGO_INTERNO_AUTO_LEN].upper()


def _normalize_codigo_interno(raw: str) -> str:
    return str(raw or '').strip().upper()


def _validate_codigo_interno_or_raise(code: str) -> str:
    c = _normalize_codigo_interno(code)
    if not c:
        return ''
    if len(c) != _CODIGO_INTERNO_MAX_LEN:
        raise ValueError('codigo_interno_length')
    if _CODIGO_INTERNO_PATTERN.match(c) is None:
        raise ValueError('codigo_interno_invalid')
    return c


def _is_valid_codigo_interno(code: str) -> bool:
    try:
        c = _normalize_codigo_interno(code)
        if not c:
            return False
        if len(c) != _CODIGO_INTERNO_MAX_LEN:
            return False
        return _CODIGO_INTERNO_PATTERN.match(c) is not None
    except Exception:
        return False


def _strip_accents(s: str) -> str:
    try:
        return ''.join(ch for ch in unicodedata.normalize('NFKD', s) if not unicodedata.combining(ch))
    except Exception:
        return s


def _alnum_upper(s: str) -> str:
    raw = _strip_accents(str(s or '').upper())
    return re.sub(r'[^A-Z0-9]', '', raw)


def _code3(s: str, fallback: str = 'XXX') -> str:
    base = _alnum_upper(s)
    if not base:
        base = _alnum_upper(fallback)
    base = (base + 'XXX')[:3]
    return base


def _codigo_prefix_from(name: str, category_name: str) -> str:
    nnn = _code3(name, 'XXX')
    ccc = _code3(category_name, 'GEN')
    if not ccc:
        ccc = 'GEN'
    return (nnn + ccc)[:6]


def _generate_codigo_interno(company_id: str, name: str, category_name: str = '', used: Optional[set] = None) -> str:
    prefix = _codigo_prefix_from(name, category_name or 'GEN')
    used_set = used if isinstance(used, set) else None

    taken = set()
    if used_set is not None:
        # Batch generation path (migration): rely on in-memory set for uniqueness.
        taken = used_set
    else:
        try:
            like = f"{prefix}%"
            rows = (
                db.session.query(Product.internal_code)
                .filter(Product.company_id == str(company_id or '').strip())
                .filter(Product.internal_code.ilike(like))
                .all()
            )
            for (code,) in rows:
                c = str(code or '').strip().upper()
                if len(c) == 8 and c.startswith(prefix):
                    taken.add(c)
        except Exception:
            taken = set()

    for n in range(1, 100):
        suf = str(n).zfill(2)
        candidate = prefix + suf
        if candidate in taken:
            continue
        try:
            exists_global = db.session.query(Product.id).filter(Product.internal_code == candidate).first() is not None
        except Exception:
            exists_global = False
        if exists_global:
            continue
        if used_set is not None and taken is used_set:
            used_set.add(candidate)
        return candidate

    # Fallback (extremely rare): prefix space exhausted (01-99). Ensure DB uniqueness.
    try:
        current_app.logger.warning('codigo_interno prefix exhausted for prefix=%s (company_id=%s)', prefix, str(company_id or '').strip())
    except Exception:
        pass
    return _generate_unique_internal_code()


def _ensure_codigo_interno(product: Product) -> None:
    raw_existing = str(getattr(product, 'internal_code', '') or '').strip()
    if raw_existing:
        normalized = _normalize_codigo_interno(raw_existing)
        # Keep as-is only if valid under current rules.
        if _is_valid_codigo_interno(normalized):
            if normalized != raw_existing:
                product.internal_code = normalized
            return
    try:
        cid = str(getattr(product, 'company_id', '') or '').strip() or _company_id()
    except Exception:
        cid = _company_id()
    try:
        cat_name = ''
        if getattr(product, 'category', None) is not None:
            cat_name = str(getattr(getattr(product, 'category', None), 'name', '') or '').strip()
    except Exception:
        cat_name = ''
    if not cat_name:
        cat_name = 'GEN'
    product.internal_code = _generate_codigo_interno(cid, getattr(product, 'name', '') or '', cat_name)


@bp.post('/api/products/migrate-codigo-interno')
@login_required
@module_required('inventory')
def migrate_codigo_interno_all_products():
    """Re-generate codigo_interno for all products of the company using NNNCCC## format."""
    try:
        ensure_request_context()
    except Exception:
        pass
    payload = request.get_json(silent=True) or {}
    dry_run = bool(payload.get('dry_run')) or (str(request.args.get('dry_run') or '').strip() in ('1', 'true', 'True'))
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    rows = (
        db.session.query(Product)
        .options(joinedload(Product.category))
        .filter(Product.company_id == cid)
        .order_by(Product.id.asc())
        .all()
    )
    used = set()
    updated = 0
    errors = []
    changes = []

    for p in (rows or []):
        try:
            cat_name = ''
            try:
                cat_name = str(getattr(getattr(p, 'category', None), 'name', '') or '').strip()
            except Exception:
                cat_name = ''
            if not cat_name:
                cat_name = 'GEN'
            next_code = _generate_codigo_interno(cid, getattr(p, 'name', '') or '', cat_name, used=used)
            if not next_code or len(next_code) != 8:
                raise ValueError('codigo_interno_generation_failed')
            before = str(getattr(p, 'internal_code', '') or '').strip()
            if before != next_code:
                p.internal_code = next_code
                updated += 1
                changes.append({'id': int(getattr(p, 'id', 0) or 0), 'name': str(getattr(p, 'name', '') or ''), 'before': before, 'after': next_code})
        except Exception:
            errors.append({'id': getattr(p, 'id', None), 'error': 'failed'})

    try:
        if dry_run:
            db.session.rollback()
        else:
            db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400

    return jsonify({'ok': True, 'dry_run': dry_run, 'total': len(rows or []), 'updated': updated, 'errors': errors, 'changes': changes})


def _ensure_codigo_interno_on_products(products) -> bool:
    changed = False
    for p in (products or []):
        if not p:
            continue
        before = str(getattr(p, 'internal_code', '') or '').strip()
        _ensure_codigo_interno(p)
        after = str(getattr(p, 'internal_code', '') or '').strip()
        if before != after:
            changed = True
    return changed


def _encode_base36(n: int) -> str:
    chars = '0123456789ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    if n <= 0:
        return '0'
    out = ''
    x = int(n)
    while x > 0:
        x, r = divmod(x, 36)
        out = chars[r] + out
    return out


def _generate_lot_code() -> str:
    ts = int(datetime.utcnow().timestamp() * 1000)
    base = _encode_base36(ts).rjust(10, '0')[-10:]
    for _ in range(20):
        code = base
        exists = db.session.query(InventoryLot.id).filter(InventoryLot.lot_code == code).first() is not None
        if not exists:
            return code
        ts2 = int(datetime.utcnow().timestamp() * 1000) + int(uuid4().hex[:4], 16)
        base = _encode_base36(ts2).rjust(10, '0')[-10:]
    return uuid4().hex[:10].upper()


def _allowed_image(filename: str) -> bool:
    ext = os.path.splitext(filename or '')[1].lower()
    return ext in ('.png', '.jpg', '.jpeg')


@bp.get('/api/categories')
@login_required
@module_required('inventory')
def list_categories_api():
    limit = int(request.args.get('limit') or 5000)
    if limit <= 0 or limit > 10000:
        limit = 5000
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    q = db.session.query(Category).filter(Category.company_id == cid, Category.parent_id == None)  # noqa: E711
    active = (request.args.get('active') or '').strip()
    if active in ('1', 'true', 'True'):
        q = q.filter(Category.active == True)  # noqa: E712
    q = q.order_by(Category.name.asc()).limit(limit)
    rows = q.all()
    return jsonify({'ok': True, 'items': [_serialize_category(r) for r in rows]})


@bp.post('/api/categories')
@login_required
@module_required('inventory')
def create_category_api():
    payload = request.get_json(silent=True) or {}
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    name = str(payload.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'name_required'}), 400
    existing = _find_category_by_norm(cid, name)
    if existing:
        return jsonify({'ok': False, 'error': 'already_exists', 'item': _serialize_category(existing)}), 400
    row = Category(company_id=cid, name=name, active=bool(payload.get('active', True)))
    db.session.add(row)
    db.session.commit()
    return jsonify({'ok': True, 'item': _serialize_category(row)})


@bp.get('/api/import-excel/template')
@login_required
@module_required('inventory')
def inventory_import_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventario'
    headers = [
        'nombre',
        'categoria',
        'codigo_interno',
        'precio_lista',
        'descripcion',
        'costo_unitario',
        'cantidad',
        'proveedor',
        'vencimiento',
        'nota_lote',
        'stock_minimo',
        'punto_pedido',
    ]
    ws.append(headers)

    try:
        ws.freeze_panes = 'A2'
    except Exception:
        pass

    try:
        bold = Font(bold=True)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = bold
    except Exception:
        pass

    try:
        widths = {
            'A': 16,  # nombre
            'B': 14,  # categoria
            'C': 16,  # codigo_interno
            'D': 12,  # precio_lista
            'E': 18,  # descripcion
            'F': 14,  # costo_unitario
            'G': 10,  # cantidad
            'H': 14,  # proveedor
            'I': 12,  # vencimiento
            'J': 14,  # nota_lote
            'K': 12,  # stock_minimo
            'L': 12,  # punto_pedido
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
    except Exception:
        pass
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name='plantilla_inventario.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@bp.post('/api/import-excel')
@login_required
@module_required('inventory')
def inventory_import_excel():
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    f = None
    try:
        f = request.files.get('file')
    except Exception:
        f = None
    if not f:
        return jsonify({'ok': False, 'error': 'file_required'}), 400

    filename = str(getattr(f, 'filename', '') or '').strip().lower()
    if not filename.endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'invalid_file_type'}), 400

    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        return jsonify({'ok': False, 'error': 'invalid_excel'}), 400

    try:
        ws = wb.worksheets[0]
    except Exception:
        return jsonify({'ok': False, 'error': 'invalid_excel'}), 400

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        header_row = None
    if not header_row:
        return jsonify({'ok': False, 'error': 'missing_header'}), 400

    header_map = {}
    for idx, val in enumerate(header_row):
        key = _normalize_header(val)
        if not key:
            continue
        if key not in header_map:
            header_map[key] = idx

    required = ['nombre', 'categoria', 'precio_lista']
    missing = [k for k in required if k not in header_map]
    if missing:
        return jsonify({'ok': False, 'error': 'missing_required_columns', 'missing': missing}), 400

    def get_cell(row_vals, key: str):
        try:
            pos = header_map.get(key)
            if pos is None:
                return None
            if pos >= len(row_vals):
                return None
            return row_vals[pos]
        except Exception:
            return None

    errors = []
    created_products = 0
    created_lots = 0
    created_categories = 0
    created_suppliers = 0
    processed_rows = 0

    today = dt_date.today()
    try:
        received_at = datetime.combine(today, datetime.min.time())
    except Exception:
        received_at = datetime.utcnow()

    for i, row_vals in enumerate(rows_iter, start=2):
        vals = row_vals or ()
        is_blank = True
        for v in vals:
            if v is None:
                continue
            if str(v).strip() != '':
                is_blank = False
                break
        if is_blank:
            continue

        processed_rows += 1

        nested = db.session.begin_nested()
        try:
            nombre = str(get_cell(vals, 'nombre') or '').strip()
            categoria = str(get_cell(vals, 'categoria') or '').strip()
            precio_lista = _num(get_cell(vals, 'precio_lista'))

            if not nombre:
                raise ValueError('nombre_required')
            if not categoria:
                raise ValueError('categoria_required')
            if precio_lista < 0:
                raise ValueError('precio_lista_invalid')

            costo_unitario_raw = get_cell(vals, 'costo_unitario')
            costo_unitario = _num(costo_unitario_raw) if costo_unitario_raw is not None and str(costo_unitario_raw).strip() != '' else 0.0
            if costo_unitario < 0:
                raise ValueError('costo_unitario_invalid')

            cantidad_raw = get_cell(vals, 'cantidad')
            cantidad = _num(cantidad_raw) if cantidad_raw is not None and str(cantidad_raw).strip() != '' else 0.0
            if cantidad < 0:
                raise ValueError('cantidad_invalid')

            descripcion = str(get_cell(vals, 'descripcion') or '').strip() or None
            codigo_interno_raw = str(get_cell(vals, 'codigo_interno') or '').strip() or ''
            nota_lote = str(get_cell(vals, 'nota_lote') or '').strip() or None

            stock_minimo_raw = get_cell(vals, 'stock_minimo')
            stock_minimo = _num(stock_minimo_raw) if stock_minimo_raw is not None and str(stock_minimo_raw).strip() != '' else 0.0
            if stock_minimo < 0:
                raise ValueError('stock_minimo_invalid')

            punto_pedido_raw = get_cell(vals, 'punto_pedido')
            punto_pedido = _num(punto_pedido_raw) if punto_pedido_raw is not None and str(punto_pedido_raw).strip() != '' else 0.0
            if punto_pedido < 0:
                raise ValueError('punto_pedido_invalid')

            proveedor = str(get_cell(vals, 'proveedor') or '').strip() or ''

            vencimiento_raw = str(get_cell(vals, 'vencimiento') or '').strip()
            exp_dt = _parse_date_flexible(vencimiento_raw, None) if vencimiento_raw else None
            if vencimiento_raw and not exp_dt:
                raise ValueError('vencimiento_invalid')

            cat_row, cat_created = _get_or_create_category(cid, categoria)
            if not cat_row:
                raise ValueError('categoria_required')
            if cat_created:
                created_categories += 1

            supplier_id = None
            supplier_name = None
            if proveedor:
                srow, screated = _get_or_create_supplier(cid, proveedor)
                if screated:
                    created_suppliers += 1
                supplier_id = str(getattr(srow, 'id', '') or '').strip() or None
                supplier_name = str(getattr(srow, 'name', '') or '').strip() or None

            internal_code = ''
            if codigo_interno_raw:
                internal_code = _validate_codigo_interno_or_raise(codigo_interno_raw)
                exists = db.session.query(Product.id).filter(Product.internal_code == internal_code).first() is not None
                if exists:
                    raise ValueError('codigo_interno_already_exists')
            if not internal_code:
                internal_code = _generate_codigo_interno(cid, nombre, categoria or 'GEN')
            prod = Product(
                company_id=cid,
                name=nombre,
                description=descripcion,
                category_id=cat_row.id,
                sale_price=precio_lista,
                internal_code=internal_code,
                unit_name='Unidad',
                uses_lots=True,
                method='FIFO',
                min_stock=stock_minimo,
                reorder_point=punto_pedido,
                primary_supplier_id=supplier_id,
                primary_supplier_name=supplier_name,
                active=True,
            )
            db.session.add(prod)
            db.session.flush()
            created_products += 1

            if cantidad > 0:
                lot = InventoryLot(
                    company_id=cid,
                    product_id=prod.id,
                    qty_initial=cantidad,
                    qty_available=cantidad,
                    unit_cost=costo_unitario,
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    expiration_date=exp_dt,
                    lot_code=_generate_lot_code(),
                    note=nota_lote or 'Ingreso por inventario (importación Excel)',
                )
                if not supplier_id and not supplier_name:
                    lot.supplier_id = None
                    lot.supplier_name = 'Ajuste interno'
                lot.received_at = received_at
                db.session.add(lot)
                db.session.flush()

                mv = InventoryMovement(
                    company_id=cid,
                    movement_date=today,
                    type='purchase',
                    sale_ticket=None,
                    product_id=prod.id,
                    lot_id=lot.id,
                    qty_delta=cantidad,
                    unit_cost=costo_unitario,
                    total_cost=cantidad * costo_unitario,
                )
                db.session.add(mv)

                exp_amount = round(float(cantidad * costo_unitario), 2)
                exp_supplier_name = supplier_name or (lot.supplier_name or '')
                exp = Expense(
                    id=uuid4().hex,
                    company_id=cid,
                    expense_date=today,
                    payment_method='Ajuste interno',
                    amount=exp_amount,
                    category='Inventario',
                    supplier_id=supplier_id,
                    supplier_name=exp_supplier_name,
                    note=('Ingreso por inventario (importación Excel) - ' + str(nombre) + ' (' + str(lot.lot_code or '') + ')'),
                    description=None,
                    expense_type='Variable',
                    frequency='Único',
                    origin='inventory',
                )
                try:
                    exp.meta_json = json.dumps({
                        'origin_ref': {
                            'kind': 'inventory_import_excel',
                            'product_id': int(prod.id),
                            'lot_id': int(lot.id),
                            'lot_code': str(lot.lot_code or ''),
                        }
                    }, ensure_ascii=False)
                except Exception:
                    exp.meta_json = None
                db.session.add(exp)
                created_lots += 1

            nested.commit()
        except Exception as e:
            try:
                nested.rollback()
            except Exception:
                pass
            code = str(e) if str(e) else 'row_error'
            errors.append({'row': i, 'error': code})
            continue

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400

    return jsonify({
        'ok': True,
        'summary': {
            'rows_processed': processed_rows,
            'items_created': created_products,
            'lots_created': created_lots,
            'categories_created': created_categories,
            'suppliers_created': created_suppliers,
            'errors': len(errors),
        },
        'errors': errors,
    })


@bp.post('/api/import-excel/preview')
@login_required
@module_required('inventory')
def inventory_import_excel_preview():
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    f = None
    try:
        f = request.files.get('file')
    except Exception:
        f = None
    if not f:
        return jsonify({'ok': False, 'error': 'file_required'}), 400

    filename = str(getattr(f, 'filename', '') or '').strip().lower()
    if not filename.endswith('.xlsx'):
        return jsonify({'ok': False, 'error': 'invalid_file_type'}), 400

    try:
        wb = load_workbook(f, data_only=True)
    except Exception:
        return jsonify({'ok': False, 'error': 'invalid_excel'}), 400

    try:
        ws = wb.worksheets[0]
    except Exception:
        return jsonify({'ok': False, 'error': 'invalid_excel'}), 400

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        header_row = None
    if not header_row:
        return jsonify({'ok': False, 'error': 'missing_header'}), 400

    header_map = {}
    for idx, val in enumerate(header_row):
        key = _normalize_header(val)
        if not key:
            continue
        if key not in header_map:
            header_map[key] = idx

    required = ['nombre', 'categoria', 'precio_lista']
    missing = [k for k in required if k not in header_map]
    if missing:
        return jsonify({'ok': False, 'error': 'missing_required_columns', 'missing': missing}), 400

    def get_cell(row_vals, key: str):
        try:
            pos = header_map.get(key)
            if pos is None:
                return None
            if pos >= len(row_vals):
                return None
            return row_vals[pos]
        except Exception:
            return None

    # Existing names (normalized)
    existing_cat = set()
    existing_sup = set()
    existing_prod = set()
    try:
        cats = (
            db.session.query(Category)
            .filter(Category.company_id == cid, Category.parent_id == None)  # noqa: E711
            .limit(10000)
            .all()
        )
        for c in (cats or []):
            existing_cat.add(_normalize_name(getattr(c, 'name', '') or ''))
    except Exception:
        existing_cat = set()
    try:
        sups = db.session.query(Supplier).filter(Supplier.company_id == cid).limit(10000).all()
        for s in (sups or []):
            existing_sup.add(_normalize_name(getattr(s, 'name', '') or ''))
    except Exception:
        existing_sup = set()
    try:
        prods = db.session.query(Product).filter(Product.company_id == cid).limit(100000).all()
        for p in (prods or []):
            existing_prod.add(_normalize_name(getattr(p, 'name', '') or ''))
    except Exception:
        existing_prod = set()

    out_rows = []
    new_cats = set()
    new_sups = set()

    for i, row_vals in enumerate(rows_iter, start=2):
        vals = row_vals or ()
        is_blank = True
        for v in vals:
            if v is None:
                continue
            if str(v).strip() != '':
                is_blank = False
                break
        if is_blank:
            continue

        nombre = str(get_cell(vals, 'nombre') or '').strip()
        categoria = str(get_cell(vals, 'categoria') or '').strip()
        codigo_interno = str(get_cell(vals, 'codigo_interno') or '').strip()
        precio_lista = str(get_cell(vals, 'precio_lista') or '').strip()
        descripcion = str(get_cell(vals, 'descripcion') or '').strip()
        cantidad = str(get_cell(vals, 'cantidad') or '').strip()
        costo_unitario = str(get_cell(vals, 'costo_unitario') or '').strip()
        proveedor = str(get_cell(vals, 'proveedor') or '').strip()
        vencimiento = str(get_cell(vals, 'vencimiento') or '').strip()
        nota_lote = str(get_cell(vals, 'nota_lote') or '').strip()
        stock_minimo = str(get_cell(vals, 'stock_minimo') or '').strip()
        punto_pedido = str(get_cell(vals, 'punto_pedido') or '').strip()

        cat_norm = _normalize_name(categoria)
        sup_norm = _normalize_name(proveedor)
        prod_norm = _normalize_name(nombre)

        will_create_category = bool(cat_norm) and (cat_norm not in existing_cat)
        will_create_supplier = bool(sup_norm) and (sup_norm not in existing_sup)
        product_exists = bool(prod_norm) and (prod_norm in existing_prod)

        if will_create_category:
            new_cats.add(cat_norm)
        if will_create_supplier:
            new_sups.add(sup_norm)

        out_rows.append({
            'row': i,
            'nombre': nombre,
            'categoria': categoria,
            'codigo_interno': codigo_interno,
            'precio_lista': precio_lista,
            'descripcion': descripcion,
            'cantidad': cantidad,
            'costo_unitario': costo_unitario,
            'proveedor': proveedor,
            'vencimiento': vencimiento,
            'nota_lote': nota_lote,
            'stock_minimo': stock_minimo,
            'punto_pedido': punto_pedido,
            'will_create_category': will_create_category,
            'will_create_supplier': will_create_supplier,
            'product_exists': product_exists,
        })

    return jsonify({
        'ok': True,
        'rows': out_rows,
        'summary': {
            'rows_detected': len(out_rows),
            'new_categories': len(new_cats),
            'new_suppliers': len(new_sups),
        }
    })


@bp.post('/api/import-excel/commit')
@login_required
@module_required('inventory')
def inventory_import_excel_commit():
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows')
    if not isinstance(rows, list) or not rows:
        return jsonify({'ok': False, 'error': 'rows_required'}), 400

    created_products = 0
    created_lots = 0
    created_categories = 0
    created_suppliers = 0
    errors = []

    today = dt_date.today()
    try:
        received_at = datetime.combine(today, datetime.min.time())
    except Exception:
        received_at = datetime.utcnow()

    for idx, r in enumerate(rows, start=1):
        d = r if isinstance(r, dict) else {}
        row_num = d.get('row') or idx
        nested = db.session.begin_nested()
        try:
            nombre = str(d.get('nombre') or '').strip()
            categoria = str(d.get('categoria') or '').strip()
            codigo_interno_raw = str(d.get('codigo_interno') or '').strip() or ''
            descripcion = str(d.get('descripcion') or '').strip() or None

            precio_lista = _num(d.get('precio_lista'))
            if not nombre:
                raise ValueError('nombre_required')
            if not categoria:
                raise ValueError('categoria_required')
            if precio_lista < 0:
                raise ValueError('precio_lista_invalid')

            cantidad_raw = str(d.get('cantidad') or '').strip()
            cantidad = _num(cantidad_raw) if cantidad_raw != '' else 0.0
            if cantidad < 0:
                raise ValueError('cantidad_invalid')

            costo_raw = str(d.get('costo_unitario') or '').strip()
            costo_unitario = _num(costo_raw) if costo_raw != '' else 0.0
            if costo_unitario < 0:
                raise ValueError('costo_unitario_invalid')

            proveedor = str(d.get('proveedor') or '').strip() or ''
            vencimiento_raw = str(d.get('vencimiento') or '').strip() or ''
            exp_dt = _parse_date_flexible(vencimiento_raw, None) if vencimiento_raw else None
            if vencimiento_raw and not exp_dt:
                raise ValueError('vencimiento_invalid')

            nota_lote = str(d.get('nota_lote') or '').strip() or None

            stock_min_raw = str(d.get('stock_minimo') or '').strip()
            stock_minimo = _num(stock_min_raw) if stock_min_raw != '' else 0.0
            if stock_minimo < 0:
                raise ValueError('stock_minimo_invalid')

            punto_raw = str(d.get('punto_pedido') or '').strip()
            punto_pedido = _num(punto_raw) if punto_raw != '' else 0.0
            if punto_pedido < 0:
                raise ValueError('punto_pedido_invalid')

            cat_row, cat_created = _get_or_create_category(cid, categoria)
            if not cat_row:
                raise ValueError('categoria_required')
            if cat_created:
                created_categories += 1

            supplier_id = None
            supplier_name = None
            if proveedor:
                srow, screated = _get_or_create_supplier(cid, proveedor)
                if screated:
                    created_suppliers += 1
                supplier_id = str(getattr(srow, 'id', '') or '').strip() or None
                supplier_name = str(getattr(srow, 'name', '') or '').strip() or None

            internal_code = ''
            if codigo_interno_raw:
                internal_code = _validate_codigo_interno_or_raise(codigo_interno_raw)
                exists = db.session.query(Product.id).filter(Product.internal_code == internal_code).first() is not None
                if exists:
                    raise ValueError('codigo_interno_already_exists')
            if not internal_code:
                internal_code = _generate_codigo_interno(cid, nombre, categoria or 'GEN')
            prod = Product(
                company_id=cid,
                name=nombre,
                description=descripcion,
                category_id=cat_row.id,
                sale_price=precio_lista,
                internal_code=internal_code,
                unit_name='Unidad',
                uses_lots=True,
                method='FIFO',
                min_stock=stock_minimo,
                reorder_point=punto_pedido,
                primary_supplier_id=supplier_id,
                primary_supplier_name=supplier_name,
                active=True,
            )
            db.session.add(prod)
            db.session.flush()
            created_products += 1

            if cantidad > 0:
                lot = InventoryLot(
                    company_id=cid,
                    product_id=prod.id,
                    qty_initial=cantidad,
                    qty_available=cantidad,
                    unit_cost=costo_unitario,
                    supplier_id=supplier_id,
                    supplier_name=supplier_name,
                    expiration_date=exp_dt,
                    lot_code=_generate_lot_code(),
                    note=nota_lote or 'Ingreso por inventario (importación Excel)',
                )
                if not supplier_id and not supplier_name:
                    lot.supplier_id = None
                    lot.supplier_name = 'Ajuste interno'
                lot.received_at = received_at
                db.session.add(lot)
                db.session.flush()

                mv = InventoryMovement(
                    company_id=cid,
                    movement_date=today,
                    type='purchase',
                    sale_ticket=None,
                    product_id=prod.id,
                    lot_id=lot.id,
                    qty_delta=cantidad,
                    unit_cost=costo_unitario,
                    total_cost=cantidad * costo_unitario,
                )
                db.session.add(mv)

                exp_amount = round(float(cantidad * costo_unitario), 2)
                exp_supplier_name = supplier_name or (lot.supplier_name or '')
                exp = Expense(
                    id=uuid4().hex,
                    company_id=cid,
                    expense_date=today,
                    payment_method='Ajuste interno',
                    amount=exp_amount,
                    category='Inventario',
                    supplier_id=supplier_id,
                    supplier_name=exp_supplier_name,
                    note=('Ingreso por inventario (importación Excel) - ' + str(nombre) + ' (' + str(lot.lot_code or '') + ')'),
                    description=None,
                    expense_type='Variable',
                    frequency='Único',
                    origin='inventory',
                )
                try:
                    exp.meta_json = json.dumps({
                        'origin_ref': {
                            'kind': 'inventory_import_excel',
                            'product_id': int(prod.id),
                            'lot_id': int(lot.id),
                            'lot_code': str(lot.lot_code or ''),
                        }
                    }, ensure_ascii=False)
                except Exception:
                    exp.meta_json = None
                db.session.add(exp)
                created_lots += 1

            nested.commit()
        except Exception as e:
            try:
                nested.rollback()
            except Exception:
                pass
            code = str(e) if str(e) else 'row_error'
            errors.append({'row': row_num, 'error': code})
            continue

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400

    return jsonify({
        'ok': True,
        'summary': {
            'items_created': created_products,
            'lots_created': created_lots,
            'categories_created': created_categories,
            'suppliers_created': created_suppliers,
            'errors': len(errors),
        },
        'errors': errors,
    })


@bp.delete('/api/categories/<int:category_id>')
@login_required
@module_required('inventory')
def delete_category_api(category_id: int):
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    try:
        cat_id = int(category_id)
    except Exception:
        return jsonify({'ok': False, 'error': 'invalid_id'}), 400

    row = db.session.query(Category).filter(Category.company_id == cid, Category.id == cat_id).first()
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    child = db.session.query(Category.id).filter(Category.company_id == cid, Category.parent_id == row.id).first()
    if child:
        return jsonify({'ok': False, 'error': 'has_children'}), 400

    used = db.session.query(Product.id).filter(Product.company_id == cid, Product.category_id == row.id).first()
    if used:
        return jsonify({'ok': False, 'error': 'in_use'}), 400

    db.session.delete(row)
    db.session.commit()
    return jsonify({'ok': True})


@bp.get('/api/products/<int:product_id>/method_lock')
@login_required
@module_required('inventory')
def get_product_method_lock(product_id: int):
    row = db.session.get(Product, int(product_id))
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    locked = _product_method_locked(row.id)
    return jsonify({'ok': True, 'locked': bool(locked)})


def _serialize_lot(l: InventoryLot):
    return {
        'id': l.id,
        'product_id': l.product_id,
        'product_name': (l.product.name if l.product else ''),
        'codigo_interno': ((l.product.internal_code or '') if l.product else ''),
        'qty_initial': l.qty_initial,
        'qty_available': l.qty_available,
        'unit_cost': l.unit_cost,
        'received_at': l.received_at.isoformat() if l.received_at else None,
        'supplier_id': str(getattr(l, 'supplier_id', '') or ''),
        'supplier_name': l.supplier_name or '',
        'expiration_date': l.expiration_date.isoformat() if l.expiration_date else None,
        'lot_code': l.lot_code or '',
        'note': l.note or '',
        'origin_sale_ticket': l.origin_sale_ticket or '',
    }


def _serialize_movement(m: InventoryMovement):
    return {
        'id': m.id,
        'date': m.movement_date.isoformat() if m.movement_date else '',
        'type': m.type,
        'sale_ticket': m.sale_ticket or '',
        'product_id': m.product_id,
        'product_name': (m.product.name if m.product else ''),
        'lot_id': m.lot_id,
        'qty_delta': m.qty_delta,
        'unit_cost': m.unit_cost,
        'total_cost': m.total_cost,
        'created_at': m.created_at.isoformat() if m.created_at else None,
    }


@bp.get('/api/products')
@login_required
@module_required('inventory')
def list_products():
    limit = int(request.args.get('limit') or 500)
    if limit <= 0 or limit > 5000:
        limit = 500
    active = (request.args.get('active') or '').strip()
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    q = db.session.query(Product).filter(Product.company_id == cid)
    if active in ('1', 'true', 'True'):
        q = q.filter(Product.active == True)  # noqa: E712
    q = q.order_by(Product.name.asc()).limit(limit)
    rows = q.all()
    try:
        changed = _ensure_codigo_interno_on_products(rows)
        if changed:
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return jsonify({'ok': True, 'items': [_serialize_product(r) for r in rows]})


@bp.get('/api/products/search')
@login_required
@module_required('inventory')
def search_products():
    qraw = (request.args.get('q') or '').strip()
    limit = int(request.args.get('limit') or 50)
    if limit <= 0 or limit > 200:
        limit = 50

    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    q = db.session.query(Product).filter(Product.company_id == cid, Product.active == True)  # noqa: E712
    if qraw:
        term = f"%{qraw}%"
        q = q.filter(
            or_(
                Product.name.ilike(term),
                Product.internal_code.ilike(term),
                Product.barcode.ilike(term),
            )
        )
    q = q.order_by(Product.name.asc()).limit(limit)
    rows = q.all()
    try:
        changed = _ensure_codigo_interno_on_products(rows)
        if changed:
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return jsonify({'ok': True, 'items': [_serialize_product(r) for r in rows]})


@bp.post('/api/products')
@login_required
@module_required('inventory')
def create_product():
    payload = request.get_json(silent=True) or {}
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    name = str(payload.get('name') or '').strip()
    if not name:
        return jsonify({'ok': False, 'error': 'name_required'}), 400

    cat_id = payload.get('category_id')
    category = None
    if cat_id is not None and cat_id != '':
        try:
            category = db.session.query(Category).filter(Category.company_id == cid, Category.id == int(cat_id)).first()
        except Exception:
            category = None
    if not category:
        return jsonify({'ok': False, 'error': 'category_required'}), 400
    if hasattr(category, 'active') and not bool(category.active):
        return jsonify({'ok': False, 'error': 'category_inactive'}), 400

    codigo_interno_raw = None
    if 'codigo_interno' in payload:
        codigo_interno_raw = str(payload.get('codigo_interno') or '').strip()
    elif 'internal_code' in payload:
        codigo_interno_raw = str(payload.get('internal_code') or '').strip()
    internal_code = None
    if codigo_interno_raw:
        internal_code = _validate_codigo_interno_or_raise(codigo_interno_raw)
        exists = db.session.query(Product.id).filter(Product.internal_code == internal_code).first() is not None
        if exists:
            return jsonify({'ok': False, 'error': 'codigo_interno_already_exists'}), 400
    if not internal_code:
        internal_code = _generate_codigo_interno(cid, name, str(getattr(category, 'name', '') or '').strip() or 'GEN')

    supplier_id = str(payload.get('primary_supplier_id') or '').strip() or None
    supplier_name = str(payload.get('primary_supplier_name') or '').strip() or None
    supplier_row = None
    if supplier_id:
        supplier_row = db.session.query(Supplier).filter(Supplier.company_id == cid, Supplier.id == supplier_id).first()
        if not supplier_row:
            return jsonify({'ok': False, 'error': 'supplier_not_found'}), 400
        supplier_name = str(supplier_row.name or '').strip() or supplier_name

    row = Product(
        company_id=cid,
        name=name,
        description=str(payload.get('description') or '').strip() or None,
        category_id=(category.id if category else None),
        sale_price=_num(payload.get('sale_price')),
        internal_code=internal_code,
        barcode=str(payload.get('barcode') or '').strip() or None,
        image_filename=str(payload.get('image_filename') or '').strip() or None,
        unit_name='Unidad',
        uses_lots=True,
        method='FIFO',
        min_stock=_num(payload.get('min_stock')),
        reorder_point=_num(payload.get('reorder_point')),
        primary_supplier_id=supplier_id,
        primary_supplier_name=supplier_name,
        active=True,
    )
    _ensure_codigo_interno(row)
    db.session.add(row)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400
    return jsonify({'ok': True, 'item': _serialize_product(row)})


@bp.put('/api/products/<int:product_id>')
@login_required
@module_required('inventory')
def update_product(product_id: int):
    row = db.session.get(Product, int(product_id))
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    payload = request.get_json(silent=True) or {}
    name = str(payload.get('name') or row.name).strip()
    if not name:
        return jsonify({'ok': False, 'error': 'name_required'}), 400

    cat_id = payload.get('category_id')
    category = None
    if cat_id is not None and cat_id != '':
        try:
            category = db.session.get(Category, int(cat_id))
        except Exception:
            category = None
    if category and hasattr(category, 'active') and not bool(category.active):
        return jsonify({'ok': False, 'error': 'category_inactive'}), 400

    row.name = name
    row.description = str(payload.get('description') or '').strip() or None
    if category:
        row.category_id = category.id
    row.sale_price = _num(payload.get('sale_price') if payload.get('sale_price') is not None else row.sale_price)
    if payload.get('codigo_interno') is not None or payload.get('internal_code') is not None:
        raw_ci = payload.get('codigo_interno') if payload.get('codigo_interno') is not None else payload.get('internal_code')
        raw_ci = str(raw_ci or '').strip()
        if raw_ci:
            next_ci = _validate_codigo_interno_or_raise(raw_ci)
            exists = db.session.query(Product.id).filter(Product.internal_code == next_ci, Product.id != row.id).first() is not None
            if exists:
                return jsonify({'ok': False, 'error': 'codigo_interno_already_exists'}), 400
            row.internal_code = next_ci
        else:
            row.internal_code = None
    row.barcode = str(payload.get('barcode') or '').strip() or None
    if payload.get('primary_supplier_id') is not None:
        next_sid = str(payload.get('primary_supplier_id') or '').strip() or None
        if next_sid:
            srow = db.session.get(Supplier, next_sid)
            if not srow:
                return jsonify({'ok': False, 'error': 'supplier_not_found'}), 400
            row.primary_supplier_id = next_sid
            row.primary_supplier_name = str(srow.name or '').strip() or row.primary_supplier_name
        else:
            row.primary_supplier_id = None

    if payload.get('primary_supplier_name') is not None:
        next_name = str(payload.get('primary_supplier_name') or '').strip() or None
        row.primary_supplier_name = next_name
        if next_name is None:
            row.primary_supplier_id = None
    if payload.get('reorder_point') is not None:
        row.reorder_point = _num(payload.get('reorder_point'))
    row.unit_name = 'Unidad'
    if payload.get('uses_lots') is not None:
        row.uses_lots = bool(payload.get('uses_lots'))
    if payload.get('method') is not None:
        next_method = str(payload.get('method') or '').strip() or 'FIFO'
        if next_method != (row.method or 'FIFO'):
            if _product_method_locked(row.id):
                return jsonify({'ok': False, 'error': 'method_locked'}), 400
            row.method = next_method
    if payload.get('min_stock') is not None:
        row.min_stock = _num(payload.get('min_stock'))
    if payload.get('active') is not None:
        row.active = bool(payload.get('active'))

    _ensure_codigo_interno(row)
    db.session.commit()
    return jsonify({'ok': True, 'item': _serialize_product(row)})


@bp.post('/api/products/<int:product_id>/image')
@login_required
@module_required('inventory')
def upload_product_image(product_id: int):
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    row = db.session.query(Product).filter(Product.company_id == cid, Product.id == int(product_id)).first()
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404
    f = request.files.get('image')
    if not f or not getattr(f, 'filename', ''):
        return jsonify({'ok': False, 'error': 'file_required'}), 400
    if not _allowed_image(f.filename):
        return jsonify({'ok': False, 'error': 'invalid_file_type'}), 400

    try:
        asset = upload_to_r2_and_create_asset(
            company_id=cid,
            file_storage=f,
            entity_type='product',
            entity_id=str(row.id),
            key_prefix='products/images',
        )
        row.image_file_id = asset.id
        row.image_filename = None
        db.session.commit()
        return jsonify({'ok': True, 'item': _serialize_product(row)})
    except Exception:
        current_app.logger.exception('Failed to upload product image to R2')
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': 'upload_failed'}), 400


@bp.delete('/api/products/<int:product_id>')
@login_required
@module_required('inventory')
def delete_product(product_id: int):
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    row = db.session.query(Product).filter(Product.company_id == cid, Product.id == int(product_id)).first()
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    has_lots = db.session.query(InventoryLot.id).filter(InventoryLot.company_id == cid, InventoryLot.product_id == row.id).first() is not None
    has_movements = db.session.query(InventoryMovement.id).filter(InventoryMovement.company_id == cid, InventoryMovement.product_id == row.id).first() is not None

    # Para mantener integridad y UX consistente, se desactiva (no se elimina) desde la UI.
    row.active = False
    db.session.commit()
    return jsonify({'ok': True, 'soft_deleted': True, 'had_history': bool(has_lots or has_movements)})


@bp.delete('/api/products/<int:product_id>/hard')
@login_required
@module_required('inventory')
def hard_delete_product(product_id: int):
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = str(getattr(g, 'company_id', '') or '').strip()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    row = db.session.query(Product).filter(Product.company_id == cid, Product.id == int(product_id)).first()
    if not row:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    try:
        img_asset_id = str(getattr(row, 'image_file_id', '') or '').strip()

        db.session.query(InventoryMovement).filter(
            InventoryMovement.company_id == cid,
            InventoryMovement.product_id == row.id,
        ).delete(synchronize_session=False)

        db.session.query(InventoryLot).filter(
            InventoryLot.company_id == cid,
            InventoryLot.product_id == row.id,
        ).delete(synchronize_session=False)

        if img_asset_id:
            db.session.query(FileAsset).filter(FileAsset.company_id == cid, FileAsset.id == img_asset_id).delete(synchronize_session=False)
        db.session.query(FileAsset).filter(
            FileAsset.company_id == cid,
            FileAsset.entity_type == 'product',
            FileAsset.entity_id == str(row.id),
        ).delete(synchronize_session=False)

        db.session.delete(row)
        db.session.commit()
        return jsonify({'ok': True})
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400


@bp.post('/api/products/prices')
@login_required
@module_required('inventory')
def bulk_update_product_prices():
    payload = request.get_json(silent=True) or {}
    items = payload.get('items')
    if not isinstance(items, list) or not items:
        return jsonify({'ok': False, 'error': 'items_required'}), 400

    updated = []
    errors = []

    for it in items:
        d = it if isinstance(it, dict) else {}
        pid_raw = d.get('id')
        try:
            pid = int(pid_raw)
        except Exception:
            errors.append({'id': pid_raw, 'error': 'invalid_id'})
            continue

        row = db.session.get(Product, pid)
        if not row:
            errors.append({'id': pid, 'error': 'not_found'})
            continue

        if 'sale_price' not in d:
            errors.append({'id': pid, 'error': 'sale_price_required'})
            continue

        row.sale_price = _num(d.get('sale_price'))
        updated.append(row)

    if not updated and errors:
        return jsonify({'ok': False, 'error': 'no_updates', 'errors': errors}), 400

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400

    return jsonify({'ok': True, 'items': [_serialize_product(r) for r in updated], 'errors': errors})


@bp.post('/api/empty')
@login_required
@module_required('inventory')
def empty_inventory():
    """Dangerous operation: clears current inventory (stock) for the current company.

    - Deletes ALL InventoryLot rows for the company (current stock)
    - Keeps InventoryMovement rows (history), but detaches them from lots by setting lot_id=NULL
    - Products are intentionally NOT deleted.
    """
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    payload = request.get_json(silent=True) or {}
    confirm_raw = str(payload.get('confirm') or '').strip().lower()
    if confirm_raw != 'vaciar inventario':
        return jsonify({'ok': False, 'error': 'confirm_required'}), 400

    try:
        detached = (
            db.session.query(InventoryMovement)
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.lot_id.isnot(None))
            .update({InventoryMovement.lot_id: None}, synchronize_session=False)
        )
        deleted_lots = (
            db.session.query(InventoryLot)
            .filter(InventoryLot.company_id == cid)
            .delete(synchronize_session=False)
        )
        db.session.commit()
        return jsonify({'ok': True, 'detached_movements': int(detached or 0), 'deleted_lots': int(deleted_lots or 0)})
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': 'db_error'}), 400


@bp.post('/api/clear')
@login_required
@module_required('inventory')
def clear_inventory():
    """Dangerous operation: clears inventory and removes products from Inventory module.

    - Deletes ALL InventoryLot rows for the company (current stock)
    - Keeps InventoryMovement rows (history), but detaches them from lots by setting lot_id=NULL
    - Products are NOT physically deleted to preserve history integrity (InventoryMovement.product_id FK),
      but are deactivated (active=False) so they disappear from normal inventory listings.
    """
    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    payload = request.get_json(silent=True) or {}
    confirm_raw = str(payload.get('confirm') or '').strip().lower()
    if confirm_raw != 'borrar inventario':
        return jsonify({'ok': False, 'error': 'confirm_required'}), 400

    try:
        detached = (
            db.session.query(InventoryMovement)
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.lot_id.isnot(None))
            .update({InventoryMovement.lot_id: None}, synchronize_session=False)
        )
        deleted_lots = (
            db.session.query(InventoryLot)
            .filter(InventoryLot.company_id == cid)
            .delete(synchronize_session=False)
        )
        deactivated_products = (
            db.session.query(Product)
            .filter(Product.company_id == cid)
            .update({Product.active: False}, synchronize_session=False)
        )
        db.session.commit()
        return jsonify({
            'ok': True,
            'detached_movements': int(detached or 0),
            'deleted_lots': int(deleted_lots or 0),
            'deactivated_products': int(deactivated_products or 0),
        })
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        return jsonify({'ok': False, 'error': 'db_error'}), 400


@bp.get('/api/lots')
@login_required
@module_required('inventory')
def list_lots():
    limit = int(request.args.get('limit') or 5000)
    if limit <= 0 or limit > 20000:
        limit = 5000
    product_id = (request.args.get('product_id') or '').strip()
    try:
        ensure_request_context()
    except Exception:
        pass
    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    q = (
        db.session.query(InventoryLot)
        .join(Product)
        .filter(InventoryLot.company_id == cid)
        .filter(Product.company_id == cid)
    )
    if product_id:
        try:
            q = q.filter(InventoryLot.product_id == int(product_id))
        except Exception:
            return jsonify({'ok': True, 'items': []})
    q = q.order_by(InventoryLot.received_at.desc(), InventoryLot.id.desc()).limit(limit)
    rows = q.all()
    try:
        prods = []
        seen = set()
        for l in (rows or []):
            p = getattr(l, 'product', None)
            if not p:
                continue
            pid = getattr(p, 'id', None)
            if pid is None:
                continue
            if pid in seen:
                continue
            seen.add(pid)
            prods.append(p)
        changed = _ensure_codigo_interno_on_products(prods)
        if changed:
            db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
    return jsonify({'ok': True, 'items': [_serialize_lot(r) for r in rows]})


@bp.post('/api/lots')
@login_required
@module_required('inventory')
def create_lot():
    payload = request.get_json(silent=True) or {}
    try:
        product_id = int(payload.get('product_id'))
    except Exception:
        return jsonify({'ok': False, 'error': 'product_id_required'}), 400

    prod = db.session.get(Product, product_id)
    if not prod:
        return jsonify({'ok': False, 'error': 'product_not_found'}), 404

    qty = _num(payload.get('qty'))
    if qty <= 0:
        return jsonify({'ok': False, 'error': 'qty_required'}), 400

    unit_cost = _num(payload.get('unit_cost'))
    supplier_id = str(payload.get('supplier_id') or '').strip() or None
    supplier_name = str(payload.get('supplier_name') or '').strip() or None
    if supplier_id:
        srow = db.session.get(Supplier, supplier_id)
        if not srow:
            return jsonify({'ok': False, 'error': 'supplier_not_found'}), 400
        supplier_name = str(srow.name or '').strip() or supplier_name
    lot_code = _generate_lot_code()
    note = str(payload.get('note') or '').strip() or None
    exp_raw = str(payload.get('expiration_date') or '').strip()
    exp_dt = _parse_date_iso(exp_raw, None) if exp_raw else None
    row = InventoryLot(
        product_id=product_id,
        qty_initial=qty,
        qty_available=qty,
        unit_cost=unit_cost,
        supplier_id=supplier_id,
        supplier_name=supplier_name,
        expiration_date=exp_dt,
        lot_code=lot_code,
        note=note,
    )
    db.session.add(row)
    db.session.flush()

    movement_date = dt_date.today()
    try:
        received_at = datetime.combine(movement_date, datetime.min.time())
    except Exception:
        received_at = datetime.utcnow()

    mv = InventoryMovement(
        movement_date=movement_date,
        type='purchase',
        sale_ticket=None,
        product_id=product_id,
        lot_id=row.id,
        qty_delta=qty,
        unit_cost=unit_cost,
        total_cost=qty * unit_cost,
    )
    row.received_at = received_at
    db.session.add(mv)
    db.session.commit()
    return jsonify({'ok': True, 'item': _serialize_lot(row)})


@bp.put('/api/lots/<int:lot_id>')
@login_required
@module_required('inventory')
def update_lot(lot_id: int):
    lot = db.session.get(InventoryLot, int(lot_id))
    if not lot:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    payload = request.get_json(silent=True) or {}

    qty_initial_raw = payload.get('qty_initial')
    unit_cost_raw = payload.get('unit_cost')
    wants_qty_cost_change = (qty_initial_raw is not None) or (unit_cost_raw is not None)

    date_raw = payload.get('date')
    wants_date_change = date_raw is not None

    # Proteger consistencia: si el lote ya tuvo consumos, no permitimos cambiar cantidad/costo.
    # FIFO/PEPS depende de received_at: si ya hay consumos u otros movimientos aplicados al lote,
    # no se deben permitir cambios que alteren el orden o la magnitud del ingreso.
    if wants_qty_cost_change or wants_date_change:
        blocked = (
            db.session.query(InventoryMovement.id)
            .filter(InventoryMovement.lot_id == lot.id)
            .filter(InventoryMovement.type.notin_(['purchase', 'return']))
            .first()
        )
        if blocked:
            return jsonify({'ok': False, 'error': 'lot_has_sales'}), 400

        used_sale = (
            db.session.query(InventoryMovement.id)
            .filter(InventoryMovement.lot_id == lot.id)
            .filter(InventoryMovement.type == 'sale')
            .first()
        )
        if used_sale:
            return jsonify({'ok': False, 'error': 'lot_has_sales'}), 400

    qty_initial = None
    unit_cost = None
    if qty_initial_raw is not None:
        qty_initial = _num(qty_initial_raw)
        if qty_initial <= 0:
            return jsonify({'ok': False, 'error': 'qty_required'}), 400
    if unit_cost_raw is not None:
        unit_cost = _num(unit_cost_raw)
        if unit_cost < 0:
            return jsonify({'ok': False, 'error': 'unit_cost_invalid'}), 400

    supplier_id = str(payload.get('supplier_id') or '').strip() or None
    supplier_name = str(payload.get('supplier_name') or '').strip() or None
    if supplier_id:
        srow = db.session.get(Supplier, supplier_id)
        if not srow:
            return jsonify({'ok': False, 'error': 'supplier_not_found'}), 400
        supplier_name = str(srow.name or '').strip() or supplier_name

    lot_code = str(getattr(lot, 'lot_code', '') or '').strip() or None
    note = str(payload.get('note') or '').strip() or None

    exp_raw = str(payload.get('expiration_date') or '').strip()
    exp_dt = _parse_date_iso(exp_raw, None) if exp_raw else None

    movement_date = _parse_date_iso(payload.get('date'), None)
    if not movement_date:
        return jsonify({'ok': False, 'error': 'date_required'}), 400

    try:
        received_at = datetime.combine(movement_date, datetime.min.time())
    except Exception:
        received_at = datetime.utcnow()

    lot.supplier_id = supplier_id
    lot.supplier_name = supplier_name
    lot.lot_code = lot_code
    lot.note = note
    lot.expiration_date = exp_dt
    lot.received_at = received_at

    if qty_initial is not None:
        lot.qty_initial = qty_initial
        lot.qty_available = qty_initial
    if unit_cost is not None:
        lot.unit_cost = unit_cost

    # Mantener consistencia: el movimiento de ingreso del lote define la fecha del ingreso.
    try:
        mv_update = {InventoryMovement.movement_date: movement_date}
        if qty_initial is not None:
            mv_update[InventoryMovement.qty_delta] = qty_initial
        if unit_cost is not None:
            mv_update[InventoryMovement.unit_cost] = unit_cost
        if qty_initial is not None or unit_cost is not None:
            q = qty_initial if qty_initial is not None else float(lot.qty_initial or 0)
            c = unit_cost if unit_cost is not None else float(lot.unit_cost or 0)
            mv_update[InventoryMovement.total_cost] = q * c

        # Un lote puede venir de compra (purchase) o de devolución (return).
        # Actualizamos el movimiento de ingreso correspondiente.
        inbound = (
            db.session.query(InventoryMovement)
            .filter(InventoryMovement.lot_id == lot.id)
            .filter(InventoryMovement.type.in_(['purchase', 'return']))
        )
        inbound.update(mv_update, synchronize_session=False)
    except Exception:
        current_app.logger.exception('Failed to update inbound movement for inventory lot')

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'ok': False, 'error': 'db_error'}), 400

    return jsonify({'ok': True, 'item': _serialize_lot(lot)})


@bp.post('/api/lots/<int:lot_id>/adjust')
@login_required
@module_required('inventory')
def adjust_lot(lot_id: int):
    from flask_login import current_user

    try:
        ensure_request_context()
    except Exception:
        pass

    cid = _company_id()
    if not cid:
        return jsonify({'ok': False, 'error': 'no_company'}), 400

    lot = db.session.get(InventoryLot, int(lot_id))
    if not lot:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    if str(getattr(lot, 'company_id', '') or '') != cid:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    payload = request.get_json(silent=True) or {}
    date_adj = _parse_date_iso(payload.get('date') or payload.get('fecha'), dt_date.today())
    note_user = str(payload.get('note') or payload.get('nota') or '').strip()

    qty_initial_raw = payload.get('qty_initial')
    unit_cost_raw = payload.get('unit_cost')
    if qty_initial_raw is None and unit_cost_raw is None:
        return jsonify({'ok': False, 'error': 'qty_or_cost_required'}), 400

    old_qty_initial = float(getattr(lot, 'qty_initial', 0.0) or 0.0)
    old_unit_cost = float(getattr(lot, 'unit_cost', 0.0) or 0.0)

    new_qty_initial = old_qty_initial
    if qty_initial_raw is not None:
        new_qty_initial = _num(qty_initial_raw)
        if new_qty_initial <= 0:
            return jsonify({'ok': False, 'error': 'qty_required'}), 400

    new_unit_cost = old_unit_cost
    if unit_cost_raw is not None:
        new_unit_cost = _num(unit_cost_raw)
        if new_unit_cost < 0:
            return jsonify({'ok': False, 'error': 'unit_cost_invalid'}), 400

    sold_rows = (
        db.session.query(
            InventoryMovement.sale_ticket,
            InventoryMovement.movement_date,
            func.coalesce(func.sum(-InventoryMovement.qty_delta), 0.0),
        )
        .filter(InventoryMovement.company_id == cid)
        .filter(InventoryMovement.lot_id == lot.id)
        .filter(InventoryMovement.type == 'sale')
        .group_by(InventoryMovement.sale_ticket, InventoryMovement.movement_date)
        .all()
    )
    qty_sold_total = sum(float(qty or 0.0) for (_, __, qty) in (sold_rows or []))

    if float(new_qty_initial or 0.0) + 1e-9 < float(qty_sold_total or 0.0):
        return jsonify({'ok': False, 'error': 'qty_below_consumed', 'consumed_qty': round(qty_sold_total, 6)}), 400

    adjustment_id = uuid4().hex

    product_name = str(getattr(getattr(lot, 'product', None), 'name', '') or '').strip() or 'Producto'
    lot_label = str(getattr(lot, 'lot_code', '') or '').strip() or str(lot.id)

    delta_qty = float(new_qty_initial or 0.0) - float(old_qty_initial or 0.0)
    delta_cost = float(new_unit_cost or 0.0) - float(old_unit_cost or 0.0)
    stock_base = float(old_qty_initial or 0.0)
    impact_cost = float(delta_cost or 0.0) * float(stock_base or 0.0)
    impact_qty = float(delta_qty or 0.0) * float(new_unit_cost or 0.0)

    ingreso = 0.0
    gasto = 0.0
    if impact_cost > 1e-9:
        gasto += abs(impact_cost)
    elif impact_cost < -1e-9:
        ingreso += abs(impact_cost)
    if impact_qty > 1e-9:
        ingreso += abs(impact_qty)
    elif impact_qty < -1e-9:
        gasto += abs(impact_qty)

    neto = float(ingreso or 0.0) - float(gasto or 0.0)
    amount_abs = abs(float(neto or 0.0))

    supplier_id = str(getattr(lot, 'supplier_id', '') or '').strip() or None
    supplier_name = str(getattr(lot, 'supplier_name', '') or '').strip() or None
    if not supplier_id and not supplier_name:
        supplier_id = 'system_adjustment'
        supplier_name = 'Ajuste interno / Sin proveedor'

    summary_sign = '+' if neto >= 0 else '-'
    summary_reason = 'Ajuste interno de inventario'
    if abs(impact_cost) > 1e-9 and abs(impact_qty) > 1e-9:
        summary_reason = 'Ajuste de inventario (costo + cantidad)'
    elif abs(impact_cost) > 1e-9:
        summary_reason = 'Revalorización de inventario'
    elif abs(impact_qty) > 1e-9:
        summary_reason = 'Corrección de stock'

    summary_line = (
        f"Ajuste de inventario – Lote {lot_label} · {supplier_name or '-'} · "
        f"Total {summary_sign}${round(amount_abs, 2)}"
    )

    base_note_lines = [
        summary_line,
        '',
        f"Producto: {product_name}",
        f"Proveedor: {supplier_name or '-'}",
        f"Stock: {round(old_qty_initial, 6)} → {round(new_qty_initial, 6)}",
        f"Costo: ${round(old_unit_cost, 6)} → ${round(new_unit_cost, 6)}",
        f"Tipo: {summary_reason}",
    ]
    if abs(impact_cost) > 1e-9:
        base_note_lines.append(f"Impacto por costo: {'+' if impact_cost >= 0 else '-'}${round(abs(impact_cost), 2)}")
    if abs(impact_qty) > 1e-9:
        base_note_lines.append(f"Impacto por cantidad: {'+' if impact_qty >= 0 else '-'}${round(abs(impact_qty), 2)}")
    base_note_lines.append(f"AdjustmentId: {adjustment_id}")
    if note_user:
        base_note_lines.extend(['', f"Nota: {note_user}"])

    try:
        lot.qty_initial = float(new_qty_initial or 0.0)
        lot.unit_cost = float(new_unit_cost or 0.0)
        lot.qty_available = max(0.0, float(new_qty_initial or 0.0) - float(qty_sold_total or 0.0))

        qty_delta_adj = float(new_qty_initial or 0.0) - float(old_qty_initial or 0.0)
        db.session.add(InventoryMovement(
            company_id=cid,
            movement_date=date_adj,
            type='lot_adjust',
            sale_ticket=None,
            product_id=lot.product_id,
            lot_id=lot.id,
            qty_delta=qty_delta_adj,
            unit_cost=float(new_unit_cost or 0.0),
            total_cost=qty_delta_adj * float(new_unit_cost or 0.0),
        ))

        delta_unit_cost = float(new_unit_cost or 0.0) - float(old_unit_cost or 0.0)
        if abs(delta_unit_cost) > 1e-9 and sold_rows:
            for t, d, qty in (sold_rows or []):
                ticket = str(t or '').strip()
                if not ticket:
                    continue
                sold_qty = float(qty or 0.0)
                if abs(sold_qty) <= 1e-9:
                    continue
                db.session.add(InventoryMovement(
                    company_id=cid,
                    movement_date=d or date_adj,
                    type='sale_adjust',
                    sale_ticket=ticket,
                    product_id=lot.product_id,
                    lot_id=lot.id,
                    qty_delta=0.0,
                    unit_cost=delta_unit_cost,
                    total_cost=sold_qty * delta_unit_cost,
                ))

        created = {
            'adjustment_id': adjustment_id,
            'sales': [],
            'expenses': [],
            'difference_valor': round(neto, 4),
        }

        if amount_abs > 1e-9:
            mv_notes = '\n'.join(base_note_lines)
            mv_origin_ref = {
                'kind': 'lot_adjustment',
                'adjustment_id': adjustment_id,
                'product_id': lot.product_id,
                'lot_id': lot.id,
            }
            mv_meta = {
                'origin_ref': mv_origin_ref,
                'inventory_adjustment': {
                    'provider_attribution': {
                        'supplier_id': supplier_id,
                        'supplier_name': supplier_name,
                    },
                    'calculation': {
                        'old_qty': float(old_qty_initial or 0.0),
                        'new_qty': float(new_qty_initial or 0.0),
                        'old_unit_cost': float(old_unit_cost or 0.0),
                        'new_unit_cost': float(new_unit_cost or 0.0),
                        'delta_qty': float(delta_qty or 0.0),
                        'delta_cost': float(delta_cost or 0.0),
                        'stock_base': float(stock_base or 0.0),
                        'impact_cost': float(impact_cost or 0.0),
                        'impact_qty': float(impact_qty or 0.0),
                        'ingreso': float(ingreso or 0.0),
                        'gasto': float(gasto or 0.0),
                        'neto': float(neto or 0.0),
                    },
                },
            }

            mv_eid = uuid4().hex
            mv_amount = -amount_abs if neto > 0 else amount_abs
            db.session.add(Expense(
                id=mv_eid,
                company_id=cid,
                expense_date=date_adj,
                payment_method='Ajuste interno',
                amount=mv_amount,
                category='Inventario',
                supplier_id=supplier_id,
                supplier_name=supplier_name,
                note=mv_notes,
                expense_type='AjusteInventario',
                origin='inventory',
                meta_json=json.dumps(mv_meta, ensure_ascii=False),
            ))
            created['expenses'] = [mv_eid]

        db.session.commit()
        return jsonify({
            'ok': True,
            'item': _serialize_lot(lot),
            'created': created,
        })
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception('Failed to adjust inventory lot')
        payload = {'ok': False, 'error': 'db_error'}
        try:
            if bool(getattr(current_app, 'debug', False)):
                payload['details'] = str(e)
        except Exception:
            pass
        return jsonify(payload), 400


@bp.delete('/api/lots/<int:lot_id>')
@login_required
@module_required('inventory')
def delete_lot(lot_id: int):
    lot = db.session.get(InventoryLot, int(lot_id))
    if not lot:
        return jsonify({'ok': False, 'error': 'not_found'}), 404

    # No permitir borrar si el lote tiene consumos registrados
    used = db.session.query(InventoryMovement.id).filter(InventoryMovement.lot_id == lot.id).filter(InventoryMovement.type == 'sale').first()
    if used:
        return jsonify({'ok': False, 'error': 'lot_has_sales'}), 400

    # Borrado consistente:
    # - movimientos de inventario asociados (purchase/return/etc)
    # - egresos creados desde inventario que referencien este lote (meta_json.origin_ref.lot_id)
    try:
        cid = str(getattr(lot, 'company_id', '') or '').strip()

        # borrar egresos vinculados al lote
        # Nota: Expense.meta_json guarda origin_ref (dict) y puede tener lot_id como número o string.
        patterns = [
            f'%"lot_id":{lot.id}%',
            f'%"lot_id": {lot.id}%',
            f'%"lot_id":"{lot.id}"%',
            f'%"lot_id": "{lot.id}"%',
        ]
        try:
            conds = []
            for p in patterns:
                conds.append(Expense.meta_json.ilike(p))
            if cid and conds:
                (
                    db.session.query(Expense)
                    .filter(Expense.company_id == cid)
                    .filter(Expense.origin == 'inventory')
                    .filter(or_(*conds))
                    .delete(synchronize_session=False)
                )
        except Exception:
            current_app.logger.exception('Failed to delete inventory-origin expenses for lot')

        # borrar movimientos asociados (compras/ajustes/return) y el lote
        db.session.query(InventoryMovement).filter(InventoryMovement.lot_id == lot.id).delete(synchronize_session=False)
        db.session.delete(lot)
        db.session.commit()
    except Exception:
        db.session.rollback()
        current_app.logger.exception('Failed to delete inventory lot')
        return jsonify({'ok': False, 'error': 'db_error'}), 400
    return jsonify({'ok': True})


@bp.get('/api/movements')
@login_required
@module_required('inventory')
def list_inventory_movements():
    raw_from = (request.args.get('from') or '').strip()
    raw_to = (request.args.get('to') or '').strip()
    product_id = (request.args.get('product_id') or '').strip()
    limit = int(request.args.get('limit') or 2000)
    if limit <= 0 or limit > 5000:
        limit = 500

    try:
        ensure_request_context()
    except Exception:
        pass
    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    d_from = _parse_date_iso(raw_from, None)
    d_to = _parse_date_iso(raw_to, None)
    q = db.session.query(InventoryMovement).filter(InventoryMovement.company_id == cid)
    if d_from:
        q = q.filter(InventoryMovement.movement_date >= d_from)
    if d_to:
        q = q.filter(InventoryMovement.movement_date <= d_to)
    if product_id:
        try:
            q = q.filter(InventoryMovement.product_id == int(product_id))
        except Exception:
            return jsonify({'ok': True, 'items': []})
    q = q.order_by(InventoryMovement.movement_date.desc(), InventoryMovement.id.desc()).limit(limit)
    rows = q.all()
    return jsonify({'ok': True, 'items': [_serialize_movement(r) for r in rows]})
