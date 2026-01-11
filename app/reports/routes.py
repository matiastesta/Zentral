import json
import os
from datetime import date as dt_date
from datetime import datetime, timedelta
from io import BytesIO

from flask import current_app, g, jsonify, render_template, request, send_file
from flask_login import login_required
from sqlalchemy import and_

from app import db
from app.models import BusinessSettings, CashCount, Category, Employee, Expense, Installment, InstallmentPlan, InventoryLot, InventoryMovement, Product, Sale, SaleItem
from app.permissions import module_required
from app.reports import bp


def _parse_date_iso(raw, default=None):
    s = str(raw or '').strip()
    if not s:
        return default
    try:
        return dt_date.fromisoformat(s)
    except Exception:
        return default


def _num(v):
    try:
        return float(v)
    except Exception:
        return 0.0


def _company_id() -> str:
    try:
        return str(getattr(g, 'company_id', '') or '').strip()
    except Exception:
        return ''


def _load_crm_config(company_id: str) -> dict:
    try:
        from app.customers.routes import _load_crm_config as _load
        return _load(company_id)
    except Exception:
        return {
            'debt_overdue_days': 30,
            'debt_critical_days': 60,
        }


def _load_meta(row: Expense) -> dict:
    try:
        if row and row.meta_json:
            parsed = json.loads(row.meta_json) if isinstance(row.meta_json, str) else {}
            return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}
    return {}


def _is_payroll_expense(row: Expense) -> bool:
    try:
        ccat = str(row.category or '').lower()
        return ('nomina' in ccat) or ('nómina' in ccat)
    except Exception:
        return False


def _is_supplier_cc_payment(row: Expense) -> bool:
    meta = _load_meta(row)
    if meta.get('supplier_cc_payment') is True:
        return True
    try:
        ccat = str(row.category or '').lower()
        if 'pago cuenta corriente proveedor' in ccat:
            return True
    except Exception:
        return False
    return False


def _is_supplier_cc_pending(row: Expense) -> bool:
    meta = _load_meta(row)
    cc = meta.get('supplier_cc')
    if not isinstance(cc, dict):
        return False
    return bool(cc.get('enabled'))


def _format_currency_ars(n) -> str:
    try:
        v = float(n or 0.0)
    except Exception:
        v = 0.0
    try:
        s = f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return '$ ' + s
    except Exception:
        return '$ ' + str(v)


def _get_business_info():
    bs = None
    try:
        bs = BusinessSettings.get_for_company(getattr(g, 'company_id', None))
    except Exception:
        bs = None
    name = (getattr(bs, 'name', None) or '').strip() or 'Zentral'
    logo_path = None
    try:
        if bs and getattr(bs, 'logo_filename', None):
            p = os.path.join(current_app.static_folder, 'uploads', str(bs.logo_filename))
            if os.path.exists(p):
                logo_path = p
    except Exception:
        logo_path = None
    if not logo_path:
        try:
            p = os.path.join(current_app.static_folder, 'uploads', 'business_logo.png')
            if os.path.exists(p):
                logo_path = p
        except Exception:
            logo_path = None
    return name, logo_path


@bp.route("/")
@bp.route("/index")
@login_required
@module_required('reports')
def index():
    """Vista general de reportes (dummy)."""
    return render_template("reports/index.html", title="Reportes")


@bp.get('/api/eerr')
@login_required
@module_required('reports')
def eerr_api():
    raw_from = (request.args.get('from') or '').strip()
    raw_to = (request.args.get('to') or '').strip()
    raw_compare = (request.args.get('compare') or '').strip().lower()  # none | prev | yoy
    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'compare_mode': 'none', 'period': {'from': '', 'to': '', 'days': 0}, 'kpis': {}, 'kpis_prev': None, 'deltas': {}, 'rows': [], 'sub': {}, 'insights': []})

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'compare_mode': 'none', 'kpis': {}, 'kpis_prev': None, 'deltas': {}, 'table': [], 'series': [], 'insights': [], 'breakdowns': {}})

    d_from = _parse_date_iso(raw_from, None)
    d_to = _parse_date_iso(raw_to, None)
    if not d_from or not d_to:
        today = dt_date.today()
        d_from = d_from or dt_date(today.year, today.month, 1)
        d_to = d_to or today

    # 1) Ventas válidas
    sales_q = (
        db.session.query(Sale)
        .filter(Sale.company_id == cid)
        .filter(Sale.sale_date >= d_from)
        .filter(Sale.sale_date <= d_to)
        .filter(Sale.sale_type == 'Venta')
        .filter(db.func.lower(Sale.status).like('completad%'))
    )
    sales_rows = sales_q.all()

    gross_sales = sum(_num(r.total) for r in sales_rows)
    discounts = sum(_num(r.discount_general_amount) for r in sales_rows)

    # 2) Cambios: se registran como Sale(sale_type='Cambio', status='Cambio', total negativo)
    exchange_q = (
        db.session.query(Sale)
        .filter(Sale.company_id == cid)
        .filter(Sale.sale_date >= d_from)
        .filter(Sale.sale_date <= d_to)
        .filter(Sale.sale_type == 'Cambio')
    )
    exchange_rows = exchange_q.all()
    exchange_total = sum(_num(r.total) for r in exchange_rows)  # normalmente negativo

    net_sales = gross_sales + exchange_total

    # Desglose informativo de ventas (devengado)
    sales_contado = 0.0
    sales_cc = 0.0
    sales_cuotas = 0.0
    try:
        for r in (sales_rows or []):
            amt = _num(getattr(r, 'total', 0.0))
            if amt <= 0:
                continue
            is_inst = bool(getattr(r, 'is_installments', False))
            on_acc = bool(getattr(r, 'on_account', False))
            if is_inst:
                sales_cuotas += amt
            elif on_acc:
                sales_cc += amt
            else:
                sales_contado += amt
    except Exception:
        sales_contado = 0.0
        sales_cc = 0.0
        sales_cuotas = 0.0

    # 3) CMV: desde movimientos de inventario (FIFO congelado)
    valid_tickets = [str(r.ticket) for r in sales_rows if str(getattr(r, 'ticket', '') or '').strip()]
    cmv = 0.0
    if valid_tickets:
        cmv = (
            db.session.query(db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
            .filter(InventoryMovement.sale_ticket.in_(valid_tickets))
            .scalar()
        )
        cmv = _num(cmv)

    gross_margin = net_sales - cmv

    # 4) Gastos (incluye pendientes por CC proveedor; excluye pagos de CC para no duplicar)
    exp_q = (
        db.session.query(Expense)
        .filter(Expense.company_id == cid)
        .filter(Expense.expense_date >= d_from)
        .filter(Expense.expense_date <= d_to)
    )
    exp_rows = exp_q.all()

    payroll_expenses = 0.0
    opex_paid = 0.0
    opex_pending = 0.0
    for e in (exp_rows or []):
        if _is_supplier_cc_payment(e):
            continue
        amt = _num(e.amount)
        if amt <= 0:
            continue
        if _is_payroll_expense(e):
            payroll_expenses += amt
            continue
        if _is_supplier_cc_pending(e):
            opex_pending += amt
        else:
            opex_paid += amt

    operating_expenses = opex_paid + opex_pending
    ebitda = gross_margin - operating_expenses - payroll_expenses

    def _safe_pct(num, den):
        den = _num(den)
        if abs(den) <= 1e-9:
            return 0.0
        return float(num or 0.0) / den

    kpis = {
        'period': {'from': d_from.isoformat(), 'to': d_to.isoformat()},
        'sales_gross': round(gross_sales, 2),
        'sales_net': round(net_sales, 2),
        'discounts': round(discounts, 2),
        'exchanges_total': round(exchange_total, 2),
        'sales_contado': round(sales_contado, 2),
        'sales_cc': round(sales_cc, 2),
        'sales_cuotas': round(sales_cuotas, 2),
        'cmv': round(cmv, 2),
        'gross_margin': round(gross_margin, 2),
        'gross_margin_pct': round(_safe_pct(gross_margin, net_sales) * 100.0, 2),
        'operating_expenses': round(operating_expenses, 2),
        'operating_expenses_paid': round(opex_paid, 2),
        'operating_expenses_pending': round(opex_pending, 2),
        'payroll_expenses': round(payroll_expenses, 2),
        'net_result': round(ebitda, 2),
        'net_result_pct': round(_safe_pct(ebitda, net_sales) * 100.0, 2),
    }

    # 5) Serie temporal (para gráfico principal)
    # Ventas netas por día: Venta - Cambios
    sales_by_day = {}
    try:
        rows = (
            db.session.query(Sale.sale_date, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .group_by(Sale.sale_date)
            .all()
        )
        for day, amt in (rows or []):
            if day:
                sales_by_day[day.isoformat()] = _num(amt)
    except Exception:
        sales_by_day = {}

    exchanges_by_day = {}
    try:
        rows = (
            db.session.query(Sale.sale_date, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Cambio')
            .group_by(Sale.sale_date)
            .all()
        )
        for day, amt in (rows or []):
            if day:
                exchanges_by_day[day.isoformat()] = _num(amt)  # negativo
    except Exception:
        exchanges_by_day = {}

    cmv_by_day = {}
    try:
        # Join por ticket para traer CMV por fecha de venta
        rows = (
            db.session.query(Sale.sale_date, db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
            .join(InventoryMovement, and_(InventoryMovement.company_id == Sale.company_id, InventoryMovement.sale_ticket == Sale.ticket))
            .filter(Sale.company_id == cid)
            .filter(InventoryMovement.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
            .group_by(Sale.sale_date)
            .all()
        )
        for day, amt in (rows or []):
            if day:
                cmv_by_day[day.isoformat()] = _num(amt)
    except Exception:
        cmv_by_day = {}

    # Expenses by day (paid + pending), without payroll and without CC payment
    opex_by_day = {}
    payroll_by_day = {}
    try:
        for e in (exp_rows or []):
            if _is_supplier_cc_payment(e):
                continue
            if not e.expense_date:
                continue
            key = e.expense_date.isoformat()
            amt = _num(e.amount)
            if amt <= 0:
                continue
            if _is_payroll_expense(e):
                payroll_by_day[key] = payroll_by_day.get(key, 0.0) + amt
            else:
                opex_by_day[key] = opex_by_day.get(key, 0.0) + amt
    except Exception:
        opex_by_day = {}
        payroll_by_day = {}

    # Generate dense series (one point per day)
    series = []
    try:
        cur = d_from
        while cur <= d_to:
            k = cur.isoformat()
            s_day = _num(sales_by_day.get(k))
            ex_day = _num(exchanges_by_day.get(k))
            net_day = s_day + ex_day
            cmv_day = _num(cmv_by_day.get(k))
            opex_day = _num(opex_by_day.get(k))
            pay_day = _num(payroll_by_day.get(k))
            net_res_day = net_day - cmv_day - opex_day - pay_day
            series.append({
                'date': k,
                'sales_net': round(net_day, 2),
                'sales_gross': round(s_day, 2),
                'exchanges': round(ex_day, 2),
                'cmv': round(cmv_day, 2),
                'opex': round(opex_day, 2),
                'payroll': round(pay_day, 2),
                'net_result': round(net_res_day, 2),
            })
            cur = dt_date.fromordinal(cur.toordinal() + 1)
    except Exception:
        series = []

    # 6) Breakdowns
    sales_by_payment = {}
    try:
        rows = (
            db.session.query(Sale.payment_method, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .group_by(Sale.payment_method)
            .all()
        )
        for pm, amt in (rows or []):
            key = (str(pm or '').strip() or '—')
            sales_by_payment[key] = round(_num(amt), 2)
    except Exception:
        sales_by_payment = {}

    exchanges_by_payment = {}
    try:
        rows = (
            db.session.query(Sale.payment_method, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Cambio')
            .group_by(Sale.payment_method)
            .all()
        )
        for pm, amt in (rows or []):
            key = (str(pm or '').strip() or '—')
            exchanges_by_payment[key] = round(_num(amt), 2)  # negativo
    except Exception:
        exchanges_by_payment = {}

    emp_name_by_id = {}
    try:
        emp_ids = set()
        for e in (exp_rows or []):
            if _is_supplier_cc_payment(e):
                continue
            if not _is_payroll_expense(e):
                continue
            if e.employee_id:
                emp_ids.add(str(e.employee_id).strip())
        if emp_ids:
            rows = (
                db.session.query(Employee)
                .filter(Employee.company_id == cid)
                .filter(Employee.id.in_(list(emp_ids)))
                .all()
            )
            for emp in (rows or []):
                nm = (str(getattr(emp, 'name', '') or '').strip())
                if not nm:
                    first = (str(getattr(emp, 'first_name', '') or '').strip())
                    last = (str(getattr(emp, 'last_name', '') or '').strip())
                    nm = (first + ' ' + last).strip()
                if not nm:
                    nm = str(getattr(emp, 'id', '') or '').strip()
                if nm:
                    emp_name_by_id[str(getattr(emp, 'id', '') or '').strip()] = nm
    except Exception:
        emp_name_by_id = {}

    expenses_by_category = {}
    expenses_paid_by_category = {}
    expenses_pending_by_category = {}
    expenses_by_payment_method = {}
    payroll_by_employee = {}
    try:
        for e in (exp_rows or []):
            if _is_supplier_cc_payment(e):
                continue
            amt = _num(e.amount)
            if amt <= 0:
                continue
            cat = (str(e.category or '').strip() or 'Sin categoría')
            if _is_payroll_expense(e):
                emp = (str(e.employee_name or '').strip() or '')
                if (not emp) and e.employee_id:
                    emp = emp_name_by_id.get(str(e.employee_id).strip(), '')
                if not emp:
                    emp = (str(e.employee_id or '').strip() or 'Empleado')
                payroll_by_employee[emp] = payroll_by_employee.get(emp, 0.0) + amt
                continue
            pm = (str(e.payment_method or '').strip() or '—')
            expenses_by_payment_method[pm] = expenses_by_payment_method.get(pm, 0.0) + amt
            if _is_supplier_cc_pending(e):
                expenses_pending_by_category[cat] = expenses_pending_by_category.get(cat, 0.0) + amt
            else:
                expenses_paid_by_category[cat] = expenses_paid_by_category.get(cat, 0.0) + amt
            expenses_by_category[cat] = expenses_by_category.get(cat, 0.0) + amt
    except Exception:
        expenses_by_category = {}
        expenses_paid_by_category = {}
        expenses_pending_by_category = {}
        expenses_by_payment_method = {}
        payroll_by_employee = {}

    def _top_n(d: dict, n: int = 8):
        items = [{'key': k, 'amount': round(_num(v), 2)} for k, v in (d or {}).items()]
        items.sort(key=lambda x: abs(_num(x.get('amount'))), reverse=True)
        return items[:n]

    net_sales_by_payment = {}
    try:
        keys = set(list(sales_by_payment.keys()) + list(exchanges_by_payment.keys()))
        for k in keys:
            net_sales_by_payment[k] = _num(sales_by_payment.get(k)) + _num(exchanges_by_payment.get(k))
    except Exception:
        net_sales_by_payment = {}

    # Top productos por facturación (SaleItem) - solo ventas válidas
    top_products_by_revenue = []
    try:
        rows = (
            db.session.query(SaleItem.product_name, db.func.coalesce(db.func.sum(SaleItem.subtotal), 0.0))
            .join(Sale, Sale.id == SaleItem.sale_id)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(Sale.status == 'Completada')
            .group_by(SaleItem.product_name)
            .all()
        )
        items = []
        for name, amt in (rows or []):
            items.append({'key': (str(name or '').strip() or 'Producto'), 'amount': round(_num(amt), 2)})
        items.sort(key=lambda x: abs(_num(x.get('amount'))), reverse=True)
        top_products_by_revenue = items[:12]
    except Exception:
        top_products_by_revenue = []

    # Top productos por CMV (InventoryMovement) - ligado a ventas válidas
    top_products_by_cmv = []
    try:
        rows = (
            db.session.query(Product.name, db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0), db.func.coalesce(db.func.sum(-InventoryMovement.qty_delta), 0.0))
            .join(InventoryMovement, and_(InventoryMovement.company_id == Product.company_id, InventoryMovement.product_id == Product.id))
            .join(Sale, and_(Sale.company_id == InventoryMovement.company_id, Sale.ticket == InventoryMovement.sale_ticket))
            .filter(Product.company_id == cid)
            .filter(InventoryMovement.company_id == cid)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(Sale.status == 'Completada')
            .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
            .group_by(Product.name)
            .all()
        )
        items = []
        for name, cost, qty in (rows or []):
            qty = _num(qty)
            cost = _num(cost)
            items.append({
                'key': (str(name or '').strip() or 'Producto'),
                'amount': round(cost, 2),
                'qty': round(qty, 2),
                'unit_cost': round((cost / qty) if qty > 1e-9 else 0.0, 4),
            })
        items.sort(key=lambda x: abs(_num(x.get('amount'))), reverse=True)
        top_products_by_cmv = items[:12]
    except Exception:
        top_products_by_cmv = []

    # CMV por categoría de producto
    cmv_by_category = []
    try:
        rows = (
            db.session.query(Category.name, db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
            .join(Product, and_(Product.company_id == Category.company_id, Product.category_id == Category.id))
            .join(InventoryMovement, and_(InventoryMovement.company_id == Product.company_id, InventoryMovement.product_id == Product.id))
            .join(Sale, and_(Sale.company_id == InventoryMovement.company_id, Sale.ticket == InventoryMovement.sale_ticket))
            .filter(Category.company_id == cid)
            .filter(Product.company_id == cid)
            .filter(InventoryMovement.company_id == cid)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
            .group_by(Category.name)
            .all()
        )
        items = []
        for name, cost in (rows or []):
            items.append({'key': (str(name or '').strip() or 'Sin categoría'), 'amount': round(_num(cost), 2)})
        items.sort(key=lambda x: abs(_num(x.get('amount'))), reverse=True)
        cmv_by_category = items[:12]
    except Exception:
        cmv_by_category = []

    # 7) Insights (explicables, con contexto y acción)
    insights = []
    try:
        bs = None
        try:
            bs = BusinessSettings.get_for_company(getattr(g, 'company_id', None))
        except Exception:
            bs = None

        thresh_margin_pp = _num(getattr(bs, 'insight_margin_delta_pp', None)) if bs and getattr(bs, 'insight_margin_delta_pp', None) is not None else 5.0
        thresh_profit_pp = _num(getattr(bs, 'insight_profitability_delta_pp', None)) if bs and getattr(bs, 'insight_profitability_delta_pp', None) is not None else 4.0
        thresh_expenses_ratio = (_num(getattr(bs, 'insight_expenses_ratio_pct', None)) / 100.0) if bs and getattr(bs, 'insight_expenses_ratio_pct', None) is not None else 0.45

        def _pct(num, den):
            den = _num(den)
            if abs(den) <= 1e-9:
                return 0.0
            return (_num(num) / den) * 100.0

        def _severity_from_delta(delta_pp: float):
            d = abs(float(delta_pp or 0.0))
            if d >= 10.0:
                return 'alta'
            if d >= 5.0:
                return 'media'
            return 'baja'

        def _mk(kind, typ, key, title, description, key_data, action, severity, rule):
            it = {
                'kind': kind,
                'type': typ,
                'module': 'eerr',
                'key': key,
                'title': title,
                'description': description,
                'key_data': key_data,
                'suggested_action': action,
                'severity': severity,
                'rule': rule,
            }
            # compat UI actual
            it['detail'] = description
            return it

        # Período anterior comparable (mismo número de días, inmediatamente previo)
        prev = None
        try:
            def _month_end(d: dt_date) -> dt_date:
                if d.month == 12:
                    nxt = dt_date(d.year + 1, 1, 1)
                else:
                    nxt = dt_date(d.year, d.month + 1, 1)
                return nxt - timedelta(days=1)

            def _add_months(d: dt_date, months: int) -> dt_date:
                y = int(d.year)
                m = int(d.month) + int(months)
                while m > 12:
                    y += 1
                    m -= 12
                while m < 1:
                    y -= 1
                    m += 12
                last = _month_end(dt_date(y, m, 1)).day
                dd = min(int(d.day), int(last))
                return dt_date(y, m, dd)

            days = int((d_to - d_from).days) if d_from and d_to else 0
            if days < 0:
                days = 0

            prev_from = None
            prev_to = None
            if raw_compare == 'yoy':
                prev_from = _add_months(d_from, -12)
                prev_to = _add_months(d_to, -12)
            else:
                prev_to = dt_date.fromordinal(d_from.toordinal() - 1) if d_from else None
                prev_from = dt_date.fromordinal(prev_to.toordinal() - days) if prev_to else None

            if prev_from and prev_to:
                # Ventas válidas
                sales_rows_prev = (
                    db.session.query(Sale)
                    .filter(Sale.company_id == cid)
                    .filter(Sale.sale_date >= prev_from)
                    .filter(Sale.sale_date <= prev_to)
                    .filter(Sale.sale_type == 'Venta')
                    .filter(db.func.lower(Sale.status).like('completad%'))
                    .all()
                )
                gross_sales_prev = sum(_num(r.total) for r in (sales_rows_prev or []))

                exchange_rows_prev = (
                    db.session.query(Sale)
                    .filter(Sale.company_id == cid)
                    .filter(Sale.sale_date >= prev_from)
                    .filter(Sale.sale_date <= prev_to)
                    .filter(Sale.sale_type == 'Cambio')
                    .all()
                )
                exchange_total_prev = sum(_num(r.total) for r in (exchange_rows_prev or []))
                net_sales_prev = gross_sales_prev + exchange_total_prev

                # CMV del período previo
                valid_tickets_prev = [str(r.ticket) for r in (sales_rows_prev or []) if str(getattr(r, 'ticket', '') or '').strip()]
                cmv_prev = 0.0
                if valid_tickets_prev:
                    cmv_prev = (
                        db.session.query(db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
                        .filter(InventoryMovement.company_id == cid)
                        .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
                        .filter(InventoryMovement.sale_ticket.in_(valid_tickets_prev))
                        .scalar()
                    )
                    cmv_prev = _num(cmv_prev)

                # Gastos previos
                exp_rows_prev = (
                    db.session.query(Expense)
                    .filter(Expense.company_id == cid)
                    .filter(Expense.expense_date >= prev_from)
                    .filter(Expense.expense_date <= prev_to)
                    .all()
                )
                payroll_prev = 0.0
                opex_paid_prev = 0.0
                opex_pending_prev = 0.0
                for e in (exp_rows_prev or []):
                    if _is_supplier_cc_payment(e):
                        continue
                    amt = _num(e.amount)
                    if amt <= 0:
                        continue
                    if _is_payroll_expense(e):
                        payroll_prev += amt
                        continue
                    if _is_supplier_cc_pending(e):
                        opex_pending_prev += amt
                    else:
                        opex_paid_prev += amt

                opex_prev = opex_paid_prev + opex_pending_prev
                gross_margin_prev = net_sales_prev - cmv_prev
                net_result_prev = gross_margin_prev - opex_prev - payroll_prev

                prev = {
                    'period': {'from': prev_from.isoformat(), 'to': prev_to.isoformat()},
                    'sales_net': round(net_sales_prev, 2),
                    'cmv': round(cmv_prev, 2),
                    'gross_margin_pct': round(_pct(gross_margin_prev, net_sales_prev), 2),
                    'net_result_pct': round(_pct(net_result_prev, net_sales_prev), 2),
                }
        except Exception:
            prev = None

        # Insight base: sin ventas
        if net_sales <= 0:
            insights.append(_mk(
                'warning',
                'negativo',
                'eerr.no_sales',
                'Sin ventas netas en el período',
                'No hay ventas netas en el período seleccionado (o hubo solo cambios/devoluciones).',
                {'sales_net': round(_num(net_sales), 2)},
                'Revisar el período seleccionado y la clasificación de ventas/cambios. Comparar con períodos anteriores para detectar cortes operativos.',
                'alta',
                'Se genera si Ventas netas <= 0.',
            ))

        # Caída/mejora de margen bruto vs período anterior
        if prev and _num(net_sales) > 0 and _num(prev.get('sales_net')) > 0:
            cur_margin = _num(kpis.get('gross_margin_pct'))
            prev_margin = _num(prev.get('gross_margin_pct'))
            delta = cur_margin - prev_margin
            thresh = float(thresh_margin_pp or 5.0)
            if delta <= -thresh:
                insights.append(_mk(
                    'warning',
                    'negativo',
                    'eerr.margin_drop',
                    'Caída de margen bruto',
                    ('El margen bruto cayó ' + str(abs(round(delta, 2))).replace('.', ',') + ' pp respecto al período anterior.'),
                    {'gross_margin_pct': round(cur_margin, 2), 'prev_gross_margin_pct': round(prev_margin, 2), 'period_prev': prev.get('period')},
                    'Revisar CMV (costos), descuentos y precios. Mirar categorías/productos con mayor CMV y validar si hubo aumentos de costo o promos agresivas.',
                    _severity_from_delta(delta),
                    ('Se genera si Margen bruto % actual < Margen bruto % período anterior - ' + str(round(thresh, 2)).replace('.', ',') + ' pp.'),
                ))
            elif delta >= thresh:
                insights.append(_mk(
                    'ok',
                    'positivo',
                    'eerr.margin_improve',
                    'Mejora de margen bruto',
                    ('El margen bruto subió ' + str(abs(round(delta, 2))).replace('.', ',') + ' pp respecto al período anterior.'),
                    {'gross_margin_pct': round(cur_margin, 2), 'prev_gross_margin_pct': round(prev_margin, 2), 'period_prev': prev.get('period')},
                    'Identificar qué impulsó la mejora (precios, mix de ventas, menores descuentos, menor CMV). Repetir la estrategia en productos/canales clave.',
                    _severity_from_delta(delta),
                    ('Se genera si Margen bruto % actual > Margen bruto % período anterior + ' + str(round(thresh, 2)).replace('.', ',') + ' pp.'),
                ))

            # Mejora/caída de rentabilidad (resultado neto %)
            cur_net = _num(kpis.get('net_result_pct'))
            prev_net = _num(prev.get('net_result_pct'))
            delta_net = cur_net - prev_net
            thresh_net = float(thresh_profit_pp or 4.0)
            if delta_net >= thresh_net:
                insights.append(_mk(
                    'ok',
                    'positivo',
                    'eerr.profitability_improve',
                    'Mejora de rentabilidad',
                    ('La rentabilidad mejoró ' + str(abs(round(delta_net, 2))).replace('.', ',') + ' pp respecto al período anterior.'),
                    {'net_result_pct': round(cur_net, 2), 'prev_net_result_pct': round(prev_net, 2), 'period_prev': prev.get('period')},
                    'Identificar qué bajó (CMV o gastos) o qué subió (ventas netas). Capitalizarlo: reforzar lo que funcionó y evitar volver al nivel anterior.',
                    _severity_from_delta(delta_net),
                    ('Se genera si Resultado neto % mejora >= ' + str(round(thresh_net, 2)).replace('.', ',') + ' pp vs período anterior.'),
                ))
            elif delta_net <= -thresh_net:
                insights.append(_mk(
                    'warning',
                    'negativo',
                    'eerr.profitability_drop',
                    'Caída de rentabilidad',
                    ('La rentabilidad cayó ' + str(abs(round(delta_net, 2))).replace('.', ',') + ' pp respecto al período anterior.'),
                    {'net_result_pct': round(cur_net, 2), 'prev_net_result_pct': round(prev_net, 2), 'period_prev': prev.get('period')},
                    'Separar el problema: ¿bajó margen (CMV) o subieron gastos? Revisar gastos fijos/variables y validar precios/descuentos.',
                    _severity_from_delta(delta_net),
                    ('Se genera si Resultado neto % cae >= ' + str(round(thresh_net, 2)).replace('.', ',') + ' pp vs período anterior.'),
                ))

        # Gastos desalineados (gastos totales / ventas)
        if _num(net_sales) > 0:
            gastos_tot = abs(_num(operating_expenses)) + abs(_num(payroll_expenses))
            ratio = (gastos_tot / max(1.0, abs(_num(net_sales))))
            thresh_ratio = float(thresh_expenses_ratio or 0.45)
            if ratio > thresh_ratio:
                pct = round(ratio * 100.0, 2)
                times = round(ratio, 2)
                desc = ('Los gastos (operativos + nómina) representan el ' + str(pct).replace('.', ',') + '% de las ventas netas.')
                if pct >= 300.0:
                    desc = ('Los gastos (operativos + nómina) superan las ventas netas (' + str(times).replace('.', ',') + '×).')
                insights.append(_mk(
                    'warning',
                    'informativo',
                    'eerr.expenses_misaligned',
                    'Gastos desalineados vs ventas',
                    desc,
                    {'expenses_total': round(gastos_tot, 2), 'sales_net': round(_num(net_sales), 2), 'ratio': round(ratio, 4)},
                    'Revisar gastos fijos (contratos/servicios) y costos variables. Si el nivel de ventas bajó, ajustar estructura o impulsar ventas para recuperar escala.',
                    'media' if pct < 65 else 'alta',
                    ('Se genera si (Gastos operativos + Nómina) / Ventas netas > ' + str(round(thresh_ratio * 100.0, 2)).replace('.', ',') + '%.'),
                ))

        if not insights:
            insights.append(_mk(
                'ok',
                'positivo',
                'eerr.no_strong_signals',
                'Sin señales fuertes en el período',
                'No se detectaron variaciones relevantes para este período con las reglas actuales.',
                {'period': {'from': d_from.isoformat(), 'to': d_to.isoformat()}},
                'Bajar al detalle en las secciones expandibles y comparar con el período anterior para detectar cambios finos (mix, promociones, categorías).',
                'baja',
                'Se genera si ninguna regla de insights dispara.',
            ))
    except Exception:
        insights = [{'kind': 'ok', 'title': 'Sin insights', 'detail': 'No se pudo generar insights para el período.'}]

    # Tabla “contable” base (expandible en UI; por ahora sin sublíneas reales)
    table = [
        {
            'key': 'ingresos',
            'label': 'Ingresos (Ventas netas)',
            'amount': kpis['sales_net'],
            'pct_of_sales': 100.0,
            'children': [
                {'key': 'ventas_brutas', 'label': 'Ventas brutas', 'amount': kpis['sales_gross']},
                {'key': 'cambios', 'label': 'Cambios (devoluciones)', 'amount': kpis['exchanges_total']},
                {'key': 'descuentos', 'label': 'Descuentos', 'amount': -abs(kpis['discounts'])},
            ],
        },
        {
            'key': 'cmv',
            'label': 'Costo de mercadería vendida (CMV)',
            'amount': -abs(kpis['cmv']),
            'pct_of_sales': round(_safe_pct(-abs(kpis['cmv']), net_sales) * 100.0, 2),
            'children': [],
        },
        {
            'key': 'margen_bruto',
            'label': 'Margen bruto',
            'amount': kpis['gross_margin'],
            'pct_of_sales': kpis['gross_margin_pct'],
            'children': [],
        },
        {
            'key': 'gastos_operativos',
            'label': 'Gastos operativos',
            'amount': -abs(kpis['operating_expenses']),
            'pct_of_sales': round(_safe_pct(-abs(kpis['operating_expenses']), net_sales) * 100.0, 2),
            'children': [
                {'key': 'gastos_pagados', 'label': 'Pagados', 'amount': -abs(kpis['operating_expenses_paid'])},
                {'key': 'gastos_pendientes', 'label': 'Pendientes (CC Proveedores)', 'amount': -abs(kpis['operating_expenses_pending'])},
            ],
        },
        {
            'key': 'nomina',
            'label': 'Nómina',
            'amount': -abs(kpis['payroll_expenses']),
            'pct_of_sales': round(_safe_pct(-abs(kpis['payroll_expenses']), net_sales) * 100.0, 2),
            'children': [],
        },
        {
            'key': 'resultado_neto',
            'label': 'Resultado neto',
            'amount': kpis['net_result'],
            'pct_of_sales': kpis['net_result_pct'],
            'children': [],
        },
    ]

    kpis_prev = None
    deltas = {}
    try:
        if raw_compare in ('prev', 'yoy') and prev:
            kpis_prev = {
                'period': prev.get('period'),
                'sales_net': _num(prev.get('sales_net')),
                'gross_margin_pct': _num(prev.get('gross_margin_pct')),
                'net_result_pct': _num(prev.get('net_result_pct')),
            }

            def _delta(curv, prevv):
                c = _num(curv)
                p = _num(prevv)
                return {
                    'abs': round(c - p, 2),
                    'pct': round((_safe_pct(c - p, p) * 100.0) if abs(p) > 1e-9 else 0.0, 2),
                }

            deltas = {
                'sales_net': _delta(kpis.get('sales_net'), _num(prev.get('sales_net'))),
                'gross_margin_pct': {'abs': round(_num(kpis.get('gross_margin_pct')) - _num(prev.get('gross_margin_pct')), 2)},
                'net_result_pct': {'abs': round(_num(kpis.get('net_result_pct')) - _num(prev.get('net_result_pct')), 2)},
            }
    except Exception:
        kpis_prev = None
        deltas = {}

    return jsonify({
        'ok': True,
        'compare_mode': raw_compare if raw_compare in ('prev', 'yoy') else 'none',
        'kpis': kpis,
        'kpis_prev': kpis_prev,
        'deltas': deltas,
        'table': table,
        'series': series,
        'insights': insights,
        'breakdowns': {
            'sales_by_payment_method': _top_n(sales_by_payment, 10),
            'exchanges_by_payment_method': _top_n(exchanges_by_payment, 10),
            'net_sales_by_payment_method': _top_n(net_sales_by_payment, 10),
            'top_products_by_revenue': top_products_by_revenue,
            'top_products_by_cmv': top_products_by_cmv,
            'cmv_by_category': cmv_by_category,
            'expenses_by_category': _top_n(expenses_by_category, 12),
            'expenses_paid_by_category': _top_n(expenses_paid_by_category, 12),
            'expenses_pending_by_category': _top_n(expenses_pending_by_category, 12),
            'expenses_by_payment_method': _top_n(expenses_by_payment_method, 10),
            'payroll_by_employee': _top_n(payroll_by_employee, 12),
        },
    })


@bp.get('/api/sales_analysis')
@login_required
@module_required('reports')
def sales_analysis_api():
    raw_from = (request.args.get('from') or '').strip()
    raw_to = (request.args.get('to') or '').strip()
    raw_compare = (request.args.get('compare') or '').strip().lower()  # prev | yoy (future)
    group_by = (request.args.get('group_by') or 'product').strip().lower()  # product | category

    category_id = (request.args.get('category_id') or '').strip()
    product_id = (request.args.get('product_id') or '').strip()
    customer_id = (request.args.get('customer_id') or '').strip()
    payment_method = (request.args.get('payment_method') or '').strip()

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'group_by': group_by, 'kpis': {}, 'kpis_prev': None, 'deltas': {}, 'rows': [], 'scatter': [], 'sub': {}, 'insights': []})

    d_from = _parse_date_iso(raw_from, None)
    d_to = _parse_date_iso(raw_to, None)
    if not d_from or not d_to:
        today = dt_date.today()
        d_from = d_from or dt_date(today.year, today.month, 1)
        d_to = d_to or today

    def _safe_pct(num, den):
        den = _num(den)
        if abs(den) <= 1e-9:
            return 0.0
        return float(num or 0.0) / den

    def _period_prev(frm: dt_date, to: dt_date):
        try:
            days = int((to - frm).days) if frm and to else 0
            if days < 0:
                days = 0
            prev_to = dt_date.fromordinal(frm.toordinal() - 1) if frm else None
            prev_from = dt_date.fromordinal(prev_to.toordinal() - days) if prev_to else None
            return prev_from, prev_to
        except Exception:
            return None, None

    def _period_yoy(frm: dt_date, to: dt_date):
        try:
            # Same number of days but one year back (approx). Prefer exact same dates if possible.
            return dt_date(frm.year - 1, frm.month, frm.day), dt_date(to.year - 1, to.month, to.day)
        except Exception:
            return None, None

    def _compute(frm: dt_date, to: dt_date):
        # Robust item amount: some older payloads might not persist SaleItem.subtotal.
        # Use qty*unit_price*(1-discount_pct) as canonical amount.
        line_amount = (SaleItem.qty * SaleItem.unit_price) * (1.0 - (SaleItem.discount_pct / 100.0))

        sq_sales = (
            db.session.query(Sale)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= frm)
            .filter(Sale.sale_date <= to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
        )
        sq_returns = (
            db.session.query(Sale)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= frm)
            .filter(Sale.sale_date <= to)
            .filter(Sale.sale_type == 'Cambio')
        )
        if payment_method:
            sq_sales = sq_sales.filter(Sale.payment_method == payment_method)
            sq_returns = sq_returns.filter(Sale.payment_method == payment_method)
        if customer_id:
            sq_sales = sq_sales.filter(Sale.customer_id == customer_id)
            sq_returns = sq_returns.filter(Sale.customer_id == customer_id)

        sales = sq_sales.all()
        returns = sq_returns.all()

        sale_tickets = [str(r.ticket) for r in (sales or []) if str(getattr(r, 'ticket', '') or '').strip()]
        return_tickets = [str(r.ticket) for r in (returns or []) if str(getattr(r, 'ticket', '') or '').strip()]
        sale_count = len(sales or [])

        si_sales_q = (
            db.session.query(
                SaleItem.product_id,
                SaleItem.product_name,
                db.func.coalesce(db.func.sum(line_amount), 0.0),
                db.func.coalesce(db.func.sum(SaleItem.qty), 0.0),
            )
            .join(Sale, Sale.id == SaleItem.sale_id)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= frm)
            .filter(Sale.sale_date <= to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .filter(SaleItem.direction == 'out')
        )
        si_returns_q = (
            db.session.query(
                SaleItem.product_id,
                SaleItem.product_name,
                db.func.coalesce(db.func.sum(line_amount), 0.0),
                db.func.coalesce(db.func.sum(SaleItem.qty), 0.0),
            )
            .join(Sale, Sale.id == SaleItem.sale_id)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= frm)
            .filter(Sale.sale_date <= to)
            .filter(Sale.sale_type == 'Cambio')
            .filter(SaleItem.direction == 'in')
        )
        if payment_method:
            si_sales_q = si_sales_q.filter(Sale.payment_method == payment_method)
            si_returns_q = si_returns_q.filter(Sale.payment_method == payment_method)
        if customer_id:
            si_sales_q = si_sales_q.filter(Sale.customer_id == customer_id)
            si_returns_q = si_returns_q.filter(Sale.customer_id == customer_id)
        if product_id:
            si_sales_q = si_sales_q.filter(SaleItem.product_id == product_id)
            si_returns_q = si_returns_q.filter(SaleItem.product_id == product_id)

        si_sales_q = si_sales_q.group_by(SaleItem.product_id, SaleItem.product_name)
        si_returns_q = si_returns_q.group_by(SaleItem.product_id, SaleItem.product_name)
        si_sales_rows = si_sales_q.all()
        si_return_rows = si_returns_q.all()

        agg_items = {}
        def _agg_add(pid_raw, pname, s_amt, q_amt, sign: float):
            pid = str(pid_raw or '').strip()
            key = pid or (str(pname or '').strip() or '—')
            if key not in agg_items:
                agg_items[key] = {
                    'product_id': pid or None,
                    'product_name': str(pname or '').strip() or 'Producto',
                    'sales': 0.0,
                    'qty': 0.0,
                }
            agg_items[key]['sales'] += (sign * _num(s_amt))
            agg_items[key]['qty'] += (sign * _num(q_amt))

        for pid_raw, pname, s_amt, q_amt in (si_sales_rows or []):
            _agg_add(pid_raw, pname, s_amt, q_amt, 1.0)
        for pid_raw, pname, s_amt, q_amt in (si_return_rows or []):
            _agg_add(pid_raw, pname, s_amt, q_amt, -1.0)

        # Map product info
        prod_map = {}
        try:
            p_rows = db.session.query(Product).filter(Product.company_id == cid).all()
            for p in (p_rows or []):
                pid = str(getattr(p, 'id', '') or '').strip()
                if pid:
                    prod_map[pid] = p
        except Exception:
            prod_map = {}

        # Category map
        cat_map = {}
        try:
            c_rows = db.session.query(Category).filter(Category.company_id == cid).all()
            for c in (c_rows or []):
                cat_key = str(getattr(c, 'id', '') or '').strip()
                if cat_key:
                    cat_map[cat_key] = c
        except Exception:
            cat_map = {}

        cmv_sale_by_pid = {}
        if sale_tickets:
            try:
                rows = (
                    db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
                    .filter(InventoryMovement.company_id == cid)
                    .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
                    .filter(InventoryMovement.sale_ticket.in_(sale_tickets))
                    .group_by(InventoryMovement.product_id)
                    .all()
                )
                for pid_int, amt in (rows or []):
                    cmv_sale_by_pid[str(pid_int)] = _num(amt)
            except Exception:
                cmv_sale_by_pid = {}

        cmv_return_by_pid = {}
        if return_tickets:
            try:
                rows = (
                    db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
                    .filter(InventoryMovement.company_id == cid)
                    .filter(InventoryMovement.type == 'return')
                    .filter(InventoryMovement.sale_ticket.in_(return_tickets))
                    .group_by(InventoryMovement.product_id)
                    .all()
                )
                for pid_int, amt in (rows or []):
                    cmv_return_by_pid[str(pid_int)] = _num(amt)
            except Exception:
                cmv_return_by_pid = {}

        # Build rows per product
        rows_prod = []
        sales_total = 0.0
        cmv_total = 0.0
        qty_total = 0.0
        for k0, a in (agg_items or {}).items():
            pid = str(a.get('product_id') or '').strip()
            pname = str(a.get('product_name') or '').strip()
            p = prod_map.get(pid)
            if not p and pname:
                for pp in (prod_map.values() or []):
                    if str(getattr(pp, 'name', '') or '').strip() == str(pname).strip():
                        p = pp
                        pid = str(getattr(pp, 'id', '') or '').strip() or pid
                        break

            cat_id = ''
            cname = ''
            try:
                cat_id = str(getattr(p, 'category_id', '') or '').strip() if p else ''
                cname = str(getattr(cat_map.get(cat_id), 'name', '') or '').strip() if cat_id else ''
            except Exception:
                cat_id = ''
                cname = ''

            if category_id and cat_id and str(cat_id) != str(category_id):
                continue
            if category_id and not cat_id:
                continue

            sales_amt = _num(a.get('sales'))
            qty_amt = _num(a.get('qty'))
            cmv_amt = _num(cmv_sale_by_pid.get(pid)) - _num(cmv_return_by_pid.get(pid))
            margin_amt = sales_amt - cmv_amt
            margin_pct = (_safe_pct(margin_amt, sales_amt) * 100.0) if abs(sales_amt) > 1e-9 else 0.0

            sales_total += sales_amt
            cmv_total += cmv_amt
            qty_total += qty_amt

            rows_prod.append({
                'key': pid or (str(pname or '').strip() or '—'),
                'product_id': pid or None,
                'label': str(pname or '').strip() or (str(getattr(p, 'name', '') or '').strip() if p else 'Producto'),
                'category_id': cat_id or None,
                'category': cname or 'Sin categoría',
                'sales': round(sales_amt, 2),
                'cmv': round(cmv_amt, 2),
                'margin': round(margin_amt, 2),
                'margin_pct': round(margin_pct, 2),
                'qty': round(qty_amt, 2),
            })

        # If grouped by category, aggregate rows_prod
        rows = []
        if group_by == 'category':
            agg = {}
            for r in rows_prod:
                ck = str(r.get('category_id') or 'none')
                if ck not in agg:
                    agg[ck] = {
                        'key': ck,
                        'category_id': r.get('category_id'),
                        'label': r.get('category') or 'Sin categoría',
                        'sales': 0.0,
                        'cmv': 0.0,
                        'margin': 0.0,
                        'qty': 0.0,
                    }
                agg[ck]['sales'] += _num(r.get('sales'))
                agg[ck]['cmv'] += _num(r.get('cmv'))
                agg[ck]['margin'] += _num(r.get('margin'))
                agg[ck]['qty'] += _num(r.get('qty'))
            for ck, r in agg.items():
                s = _num(r.get('sales'))
                m = _num(r.get('margin'))
                r['margin_pct'] = round((_safe_pct(m, s) * 100.0) if abs(s) > 1e-9 else 0.0, 2)
                r['sales'] = round(s, 2)
                r['cmv'] = round(_num(r.get('cmv')), 2)
                r['margin'] = round(m, 2)
                r['qty'] = round(_num(r.get('qty')), 2)
                rows.append(r)
        else:
            rows = rows_prod

        gross_margin = sales_total - cmv_total
        margin_pct_total = round(_safe_pct(gross_margin, sales_total) * 100.0, 2)
        avg_margin_per_sale = (gross_margin / sale_count) if sale_count > 0 else 0.0

        kpis = {
            'period': {'from': frm.isoformat(), 'to': to.isoformat()},
            'sales_total': round(sales_total, 2),
            'cmv_total': round(cmv_total, 2),
            'cmv_pct': round(_safe_pct(cmv_total, sales_total) * 100.0, 2),
            'gross_margin_total': round(gross_margin, 2),
            'gross_margin_pct': margin_pct_total,
            'avg_margin_per_sale': round(avg_margin_per_sale, 2),
            'sales_count': int(sale_count),
            'qty_total': round(qty_total, 2),
        }

        # Enrich rows with % of sales
        for r in rows:
            r['sales_pct'] = round(_safe_pct(_num(r.get('sales')), sales_total) * 100.0, 2)

        # Subanalysis
        rows_sorted_sales = sorted(rows, key=lambda x: abs(_num(x.get('sales'))), reverse=True)
        rows_sorted_margin = sorted(rows, key=lambda x: abs(_num(x.get('margin'))), reverse=True)
        top_sales = rows_sorted_sales[:10]
        top_margin = rows_sorted_margin[:10]

        # Problematic: high sales share and low margin
        low_margin_thr = 15.0
        high_sales_share_thr = 0.15
        problematic = [
            r for r in rows_sorted_sales
            if (_num(r.get('sales_pct')) >= (high_sales_share_thr * 100.0)) and (_num(r.get('margin_pct')) <= low_margin_thr)
        ][:10]

        return {
            'kpis': kpis,
            'rows': rows,
            'scatter': [
                {
                    'key': r.get('key'),
                    'label': r.get('label'),
                    'sales': _num(r.get('sales')),
                    'margin_pct': _num(r.get('margin_pct')),
                    'margin': _num(r.get('margin')),
                    'category': r.get('category') if group_by != 'category' else r.get('label'),
                    'category_id': r.get('category_id'),
                    'product_id': r.get('product_id'),
                }
                for r in rows
            ],
            'sub': {
                'top_sales': top_sales,
                'top_margin': top_margin,
                'problematic': problematic,
            },
        }

    cur = _compute(d_from, d_to)

    prev = None
    if raw_compare == 'prev':
        p_from, p_to = _period_prev(d_from, d_to)
        if p_from and p_to:
            prev = _compute(p_from, p_to)
    elif raw_compare == 'yoy':
        p_from, p_to = _period_yoy(d_from, d_to)
        if p_from and p_to:
            prev = _compute(p_from, p_to)

    # Insights (ventas) - reglas claras
    insights = []
    try:
        def _mk(kind, typ, key, title, description, key_data, action, severity, rule):
            it = {
                'kind': kind,
                'type': typ,
                'module': 'sales',
                'key': key,
                'title': title,
                'description': description,
                'key_data': key_data,
                'suggested_action': action,
                'severity': severity,
                'rule': rule,
            }
            it['detail'] = description
            return it

        rows = cur.get('rows') if isinstance(cur.get('rows'), list) else []
        k = cur.get('kpis') if isinstance(cur.get('kpis'), dict) else {}

        # Insight: riesgo de rentabilidad por productos de bajo margen
        low_thr = 15.0
        low_sales = sum(_num(r.get('sales')) for r in rows if _num(r.get('margin_pct')) < low_thr)
        total_sales = _num(k.get('sales_total'))
        share = (_safe_pct(low_sales, total_sales) * 100.0) if total_sales > 0 else 0.0
        if total_sales > 0 and share >= 30.0:
            insights.append(_mk(
                'warning',
                'negativo',
                'sales.low_margin_risk',
                'Riesgo de rentabilidad',
                ('El ' + str(round(share, 2)).replace('.', ',') + '% de tus ventas proviene de items con margen menor al ' + str(int(low_thr)) + '%.'),
                {'low_margin_threshold_pct': low_thr, 'sales_low_margin': round(low_sales, 2), 'sales_total': round(total_sales, 2), 'share_pct': round(share, 2)},
                'Revisar precios, promociones y costos de compra en los items de mayor volumen con margen bajo.',
                'alta' if share >= 45 else 'media',
                'Se genera si % ventas con margen < 15% >= 30%.',
            ))

        # Insight: oportunidad comercial (alto margen, baja participación)
        high_thr = 40.0
        candidates = [r for r in rows if _num(r.get('margin_pct')) >= high_thr and _num(r.get('sales_pct')) <= 5.0]
        candidates = sorted(candidates, key=lambda x: _num(x.get('margin_pct')), reverse=True)[:3]
        if candidates:
            best = candidates[0]
            insights.append(_mk(
                'info',
                'informativo',
                'sales.high_margin_opportunity',
                'Oportunidad comercial',
                ('"' + str(best.get('label') or '').strip() + '" tiene margen promedio del ' + str(round(_num(best.get('margin_pct')), 2)).replace('.', ',') + '% con baja participación en ventas.'),
                {'label': best.get('label'), 'margin_pct': round(_num(best.get('margin_pct')), 2), 'sales_pct': round(_num(best.get('sales_pct')), 2)},
                'Impulsar comercialmente (stock, exhibición, bundles) y medir si sube participación sin perder margen.',
                'media',
                'Se genera si existe item con margen >= 40% y % ventas <= 5%.',
            ))

        # Insight: caída de margen promedio vs período anterior
        if prev and isinstance(prev.get('kpis'), dict):
            cur_m = _num(k.get('gross_margin_pct'))
            prev_m = _num(prev['kpis'].get('gross_margin_pct'))
            delta = cur_m - prev_m
            if delta <= -5.0:
                insights.append(_mk(
                    'warning',
                    'negativo',
                    'sales.margin_drop',
                    'Caída de margen promedio',
                    ('El margen promedio cayó ' + str(abs(round(delta, 2))).replace('.', ',') + ' pp respecto al período anterior.'),
                    {'gross_margin_pct': round(cur_m, 2), 'prev_gross_margin_pct': round(prev_m, 2), 'delta_pp': round(delta, 2)},
                    'Revisar si subió el CMV, si hubo promos agresivas o cambio en el mix de ventas. Enfocar en los items de mayor volumen.',
                    'alta' if abs(delta) >= 10 else 'media',
                    'Se genera si Margen % cae >= 5 pp vs período anterior.',
                ))

        if not insights:
            insights.append(_mk(
                'ok',
                'positivo',
                'sales.no_strong_signals',
                'Sin señales fuertes en el período',
                'No se detectaron variaciones relevantes con las reglas actuales. Podés usar el scatter para ver cuadrantes y la tabla para ordenar por margen.',
                {'period': {'from': d_from.isoformat(), 'to': d_to.isoformat()}},
                'Ordenar por Margen % y revisar items de alto volumen con bajo margen. Evaluar foco comercial en los de alto margen.',
                'baja',
                'Se genera si ninguna regla de insights dispara.',
            ))
    except Exception:
        insights = [{'kind': 'ok', 'title': 'Sin insights', 'detail': 'No se pudo generar insights para el período.'}]

    # Comparison deltas
    k_cur = cur.get('kpis') if isinstance(cur.get('kpis'), dict) else {}
    k_prev = prev.get('kpis') if prev and isinstance(prev.get('kpis'), dict) else {}

    def _delta(curv, prevv):
        c = _num(curv)
        p = _num(prevv)
        return {
            'abs': round(c - p, 2),
            'pct': round((_safe_pct(c - p, p) * 100.0) if abs(p) > 1e-9 else 0.0, 2),
        }

    deltas = {
        'sales_total': _delta(k_cur.get('sales_total'), k_prev.get('sales_total')),
        'cmv_total': _delta(k_cur.get('cmv_total'), k_prev.get('cmv_total')),
        'gross_margin_total': _delta(k_cur.get('gross_margin_total'), k_prev.get('gross_margin_total')),
        'gross_margin_pct': {'abs': round(_num(k_cur.get('gross_margin_pct')) - _num(k_prev.get('gross_margin_pct')), 2)},
        'avg_margin_per_sale': _delta(k_cur.get('avg_margin_per_sale'), k_prev.get('avg_margin_per_sale')),
    } if prev else {}

    return jsonify({
        'ok': True,
        'group_by': group_by,
        'kpis': k_cur,
        'kpis_prev': k_prev if prev else None,
        'deltas': deltas,
        'rows': cur.get('rows'),
        'scatter': cur.get('scatter'),
        'sub': cur.get('sub'),
        'insights': insights,
    })


@bp.get('/api/finance')
@login_required
@module_required('reports')
def finance_api():
    raw_from = (request.args.get('from') or '').strip()
    raw_to = (request.args.get('to') or '').strip()
    raw_compare = (request.args.get('compare') or '').strip().lower()  # none | prev | yoy

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'period': {'from': '', 'to': ''}, 'compare_mode': 'none', 'kpis': {}, 'kpis_prev': None, 'deltas': {}, 'series': [], 'trends': {}, 'months_summary': {}, 'breakdowns': {}, 'insights': []})

    installments_enabled = False
    try:
        bs = BusinessSettings.get_for_company(cid)
        installments_enabled = bool(bs and bool(getattr(bs, 'habilitar_sistema_cuotas', False)))
    except Exception:
        installments_enabled = False

    d_from = _parse_date_iso(raw_from, None)
    d_to = _parse_date_iso(raw_to, None)
    if not d_from or not d_to:
        today = dt_date.today()
        d_from = d_from or dt_date(today.year, today.month, 1)
        d_to = d_to or today

    if d_to < d_from:
        d_from, d_to = d_to, d_from

    def _safe_pct(num, den):
        den = _num(den)
        if abs(den) <= 1e-9:
            return 0.0
        return float(num or 0.0) / den

    def _month_start(d: dt_date) -> dt_date:
        return dt_date(d.year, d.month, 1)

    def _month_end(d: dt_date) -> dt_date:
        if d.month == 12:
            nxt = dt_date(d.year + 1, 1, 1)
        else:
            nxt = dt_date(d.year, d.month + 1, 1)
        return nxt - timedelta(days=1)

    def _add_months(d: dt_date, months: int) -> dt_date:
        y = int(d.year)
        m = int(d.month) + int(months)
        while m > 12:
            y += 1
            m -= 12
        while m < 1:
            y -= 1
            m += 12
        last = _month_end(dt_date(y, m, 1)).day
        dd = min(int(d.day), int(last))
        return dt_date(y, m, dd)

    def _compute_period(p_from: dt_date, p_to: dt_date):
        # Cash-flow only: collections are real cash movements.
        # - Venta: only the collected part (paid_amount)
        # - Cambio: affects cash (typically negative)
        # - CobroCC / CobroCuota: explicit cash collections
        sales_paid = 0.0
        sales_cash_only = 0.0
        sales_paid_cc_or_partial = 0.0
        sales_paid_installment_down = 0.0
        collections_cc = 0.0
        collections_installments = 0.0
        collections_other = 0.0
        collections_exchanges = 0.0

        try:
            sales_paid = (
                db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
                .filter(Sale.company_id == cid)
                .filter(Sale.sale_date >= p_from)
                .filter(Sale.sale_date <= p_to)
                .filter(Sale.sale_type == 'CobroVenta')
                .filter(db.func.lower(Sale.status).like('completad%'))
                .scalar()
            )
            sales_paid = _num(sales_paid)
            sales_cash_only = float(sales_paid)
            sales_paid_cc_or_partial = 0.0
            sales_paid_installment_down = 0.0
        except Exception:
            sales_paid = 0.0
            sales_cash_only = 0.0
            sales_paid_cc_or_partial = 0.0
            sales_paid_installment_down = 0.0

        try:
            collections_exchanges = (
                db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
                .filter(Sale.company_id == cid)
                .filter(Sale.sale_date >= p_from)
                .filter(Sale.sale_date <= p_to)
                .filter(Sale.sale_type == 'Cambio')
                .scalar()
            )
            collections_exchanges = _num(collections_exchanges)
        except Exception:
            collections_exchanges = 0.0

        try:
            collections_cc = (
                db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
                .filter(Sale.company_id == cid)
                .filter(Sale.sale_date >= p_from)
                .filter(Sale.sale_date <= p_to)
                .filter(Sale.sale_type == 'CobroCC')
                .filter(db.func.lower(Sale.status).like('completad%'))
                .scalar()
            )
            collections_cc = _num(collections_cc)
        except Exception:
            collections_cc = 0.0

        if installments_enabled:
            try:
                collections_installments = (
                    db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
                    .filter(Sale.company_id == cid)
                    .filter(Sale.sale_date >= p_from)
                    .filter(Sale.sale_date <= p_to)
                    .filter(Sale.sale_type == 'CobroCuota')
                    .filter(db.func.lower(Sale.status).like('completad%'))
                    .scalar()
                )
                collections_installments = _num(collections_installments)
            except Exception:
                collections_installments = 0.0

        try:
            collections_other = (
                db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
                .filter(Sale.company_id == cid)
                .filter(Sale.sale_date >= p_from)
                .filter(Sale.sale_date <= p_to)
                .filter(Sale.sale_type.in_(['IngresoAjusteInv']))
                .scalar()
            )
            collections_other = _num(collections_other)
        except Exception:
            collections_other = 0.0

        income_total = sales_paid + collections_cc + (collections_installments if installments_enabled else 0.0) + collections_other + collections_exchanges

        payroll_paid = 0.0
        opex_paid = 0.0
        opex_pending = 0.0

        exp_rows = []
        try:
            exp_rows = (
                db.session.query(Expense)
                .filter(Expense.company_id == cid)
                .filter(Expense.expense_date >= p_from)
                .filter(Expense.expense_date <= p_to)
                .all()
            )
        except Exception:
            exp_rows = []

        for e in (exp_rows or []):
            amt = _num(getattr(e, 'amount', 0.0))
            if amt <= 0:
                continue
            if _is_supplier_cc_pending(e):
                opex_pending += amt
                continue
            if _is_payroll_expense(e):
                payroll_paid += amt
            else:
                opex_paid += amt

        expense_total = opex_paid + payroll_paid
        result_total = income_total - expense_total
        ratio = _safe_pct(expense_total, income_total)

        return {
            'period': {'from': p_from.isoformat(), 'to': p_to.isoformat()},
            'income_total': round(income_total, 2),
            # Legacy-ish keys (still useful)
            'income_sales': round(sales_paid, 2),
            'income_exchanges': round(collections_exchanges, 2),
            'income_installments': round(collections_installments, 2),
            'expense_total': round(expense_total, 2),
            'operating_expenses': round(opex_paid, 2),
            'operating_expenses_paid': round(opex_paid, 2),
            'operating_expenses_pending': round(opex_pending, 2),
            'payroll_expenses': round(payroll_paid, 2),
            'result_total': round(result_total, 2),
            'result_pct': round(_safe_pct(result_total, income_total) * 100.0, 2),
            'expense_income_ratio': round(ratio, 4),
            # Cash breakdown for the new UI
            'collections_sales_paid': round(sales_paid, 2),
            'collections_sales_cash_only': round(sales_cash_only, 2),
            'collections_sales_paid_cc_or_partial': round(sales_paid_cc_or_partial, 2),
            'collections_sales_paid_installment_down': round(sales_paid_installment_down, 2),
            'collections_cc': round(collections_cc, 2),
            'collections_installments': round(collections_installments, 2),
            'collections_other': round(collections_other, 2),
        }

    cur = _compute_period(d_from, d_to)

    # Cash balance from CashCount (best-effort)
    cash_initial = None
    cash_final = None
    try:
        row = db.session.query(CashCount).filter(CashCount.company_id == cid, CashCount.count_date == d_from).first()
        if row:
            cash_initial = _num(getattr(row, 'opening_amount', 0.0))
        else:
            prev_row = (
                db.session.query(CashCount)
                .filter(CashCount.company_id == cid)
                .filter(CashCount.count_date < d_from)
                .order_by(CashCount.count_date.desc())
                .first()
            )
            if prev_row:
                cash_initial = _num(getattr(prev_row, 'closing_amount', 0.0))
    except Exception:
        cash_initial = None

    try:
        row = db.session.query(CashCount).filter(CashCount.company_id == cid, CashCount.count_date == d_to).first()
        if row:
            cash_final = _num(getattr(row, 'closing_amount', 0.0))
        else:
            prev_row = (
                db.session.query(CashCount)
                .filter(CashCount.company_id == cid)
                .filter(CashCount.count_date <= d_to)
                .order_by(CashCount.count_date.desc())
                .first()
            )
            if prev_row:
                cash_final = _num(getattr(prev_row, 'closing_amount', 0.0))
    except Exception:
        cash_final = None

    cash_balance = {
        'initial': round(_num(cash_initial), 2) if cash_initial is not None else None,
        'final': round(_num(cash_final), 2) if cash_final is not None else None,
        'initial_source': 'cash_count' if cash_initial is not None else None,
        'final_source': 'cash_count' if cash_final is not None else None,
    }

    # Projections as-of end date (best-effort)
    crm_cfg = _load_crm_config(cid)
    overdue_days = int(_num(crm_cfg.get('debt_overdue_days') or 30) or 30)
    critical_days = int(_num(crm_cfg.get('debt_critical_days') or 60) or 60)
    if overdue_days <= 0:
        overdue_days = 30
    if critical_days <= overdue_days:
        critical_days = overdue_days + 30

    cc_pending_total = 0.0
    cc_overdue_total = 0.0
    cc_critical_total = 0.0
    cc_due_7 = 0.0
    cc_due_15 = 0.0
    cc_due_30 = 0.0
    cc_next_due_date = None
    cc_next_due_amount = 0.0
    try:
        rows = (
            db.session.query(Sale)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .filter(Sale.due_amount > 0)
            .all()
        )
        for r in (rows or []):
            if bool(getattr(r, 'is_installments', False)):
                continue
            due_amt = _num(getattr(r, 'due_amount', 0.0))
            if due_amt <= 0:
                continue
            sd = getattr(r, 'sale_date', None)
            if not sd:
                continue
            cc_pending_total += due_amt
            age = int((d_to - sd).days)
            if age >= critical_days:
                cc_critical_total += due_amt
                cc_overdue_total += due_amt
            elif age >= overdue_days:
                cc_overdue_total += due_amt
            due_date = sd + timedelta(days=overdue_days)
            days_until = int((due_date - d_to).days)
            if days_until > 0:
                if cc_next_due_date is None or due_date < cc_next_due_date:
                    cc_next_due_date = due_date
                    cc_next_due_amount = due_amt
                if days_until <= 7:
                    cc_due_7 += due_amt
                if days_until <= 15:
                    cc_due_15 += due_amt
                if days_until <= 30:
                    cc_due_30 += due_amt
    except Exception:
        cc_pending_total = 0.0
        cc_overdue_total = 0.0
        cc_critical_total = 0.0
        cc_due_7 = 0.0
        cc_due_15 = 0.0
        cc_due_30 = 0.0
        cc_next_due_date = None
        cc_next_due_amount = 0.0

    inst_pending_total = 0.0
    inst_overdue_total = 0.0
    inst_due_7 = 0.0
    inst_due_15 = 0.0
    inst_due_30 = 0.0
    inst_next_due_date = None
    inst_next_due_amount = 0.0
    inst_due_in_period = 0.0
    if installments_enabled:
        try:
            rows = (
                db.session.query(Installment)
                .filter(Installment.company_id == cid)
                .all()
            )
            for it in (rows or []):
                amt = _num(getattr(it, 'amount', 0.0))
                if amt <= 0:
                    continue
                st = str(getattr(it, 'status', '') or '').strip().lower()
                dd = getattr(it, 'due_date', None)
                if dd and d_from <= dd <= d_to:
                    inst_due_in_period += amt
                if st == 'pagada':
                    continue
                inst_pending_total += amt
                if dd:
                    days_until = int((dd - d_to).days)
                    if dd < d_to:
                        inst_overdue_total += amt
                    if days_until > 0:
                        if inst_next_due_date is None or dd < inst_next_due_date:
                            inst_next_due_date = dd
                            inst_next_due_amount = amt
                        if days_until <= 7:
                            inst_due_7 += amt
                        if days_until <= 15:
                            inst_due_15 += amt
                        if days_until <= 30:
                            inst_due_30 += amt
        except Exception:
            inst_pending_total = 0.0
            inst_overdue_total = 0.0
            inst_due_7 = 0.0
            inst_due_15 = 0.0
            inst_due_30 = 0.0
            inst_next_due_date = None
            inst_next_due_amount = 0.0
            inst_due_in_period = 0.0

    projections = {
        'cc': {
            'pending_total': round(cc_pending_total, 2),
            'next_due_date': cc_next_due_date.isoformat() if cc_next_due_date else None,
            'next_due_amount': round(_num(cc_next_due_amount), 2),
            'due_7': round(cc_due_7, 2),
            'due_15': round(cc_due_15, 2),
            'due_30': round(cc_due_30, 2),
            'overdue_total': round(cc_overdue_total, 2),
            'critical_total': round(cc_critical_total, 2),
            'overdue_days': overdue_days,
            'critical_days': critical_days,
        },
        'installments': {
            'pending_total': round(inst_pending_total, 2),
            'next_due_date': inst_next_due_date.isoformat() if inst_next_due_date else None,
            'next_due_amount': round(_num(inst_next_due_amount), 2),
            'due_7': round(inst_due_7, 2),
            'due_15': round(inst_due_15, 2),
            'due_30': round(inst_due_30, 2),
            'overdue_total': round(inst_overdue_total, 2),
            'due_in_period_total': round(inst_due_in_period, 2),
        },
    }

    # Gap devengado vs caja (financiación a clientes)
    accrued_sales_net = 0.0
    try:
        gross = (
            db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .scalar()
        )
        exch = (
            db.session.query(db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Cambio')
            .scalar()
        )
        accrued_sales_net = _num(gross) + _num(exch)
    except Exception:
        accrued_sales_net = 0.0

    gap = {
        'accrued_sales_net': round(accrued_sales_net, 2),
        'cash_collections_total': round(_num(cur.get('income_total')), 2),
        'difference': round(accrued_sales_net - _num(cur.get('income_total')), 2),
    }

    prev = None
    if raw_compare in ('prev', 'yoy'):
        if raw_compare == 'prev':
            span_days = int((d_to - d_from).days) + 1
            prev_to = d_from - timedelta(days=1)
            prev_from = prev_to - timedelta(days=max(1, span_days) - 1)
        else:
            prev_from = _add_months(d_from, -12)
            prev_to = _add_months(d_to, -12)
        prev = _compute_period(prev_from, prev_to)

    def _delta(curv, prevv):
        c = _num(curv)
        p = _num(prevv)
        return {
            'abs': round(c - p, 2),
            'pct': round((_safe_pct(c - p, p) * 100.0) if abs(p) > 1e-9 else 0.0, 2),
        }

    deltas = {}
    if prev:
        deltas = {
            'income_total': _delta(cur.get('income_total'), prev.get('income_total')),
            'expense_total': _delta(cur.get('expense_total'), prev.get('expense_total')),
            'result_total': _delta(cur.get('result_total'), prev.get('result_total')),
            'expense_income_ratio': {'abs': round(_num(cur.get('expense_income_ratio')) - _num(prev.get('expense_income_ratio')), 4)},
        }

    series = []
    try:
        m = _month_start(d_from)
        end_m = _month_start(d_to)
        while m <= end_m:
            m_from = max(m, d_from)
            m_to = min(_month_end(m), d_to)
            vals = _compute_period(m_from, m_to)
            series.append({
                'month': m.strftime('%Y-%m'),
                'from': vals['period']['from'],
                'to': vals['period']['to'],
                'income_total': vals['income_total'],
                'expense_total': vals['expense_total'],
                'result_total': vals['result_total'],
                'expense_income_ratio': vals['expense_income_ratio'],
            })
            m = _add_months(m, 1)
    except Exception:
        series = []

    pos = 0
    neg = 0
    eq = 0
    try:
        for r in (series or []):
            inc = _num(r.get('income_total'))
            res = _num(r.get('result_total'))
            thr = abs(inc) * 0.01
            if thr < 1e-6:
                thr = 0.0
            if abs(res) <= thr:
                eq += 1
            elif res > 0:
                pos += 1
            else:
                neg += 1
    except Exception:
        pos = 0
        neg = 0
        eq = 0

    # Advanced analysis (replace simple trend/months messaging)
    # 1) Compute CMV to estimate contribution margin and breakeven.
    cmv_total = 0.0
    try:
        sales_rows = (
            db.session.query(Sale)
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .all()
        )
        sale_tickets = [str(r.ticket) for r in (sales_rows or []) if str(getattr(r, 'ticket', '') or '').strip()]
        if sale_tickets:
            cmv_total = (
                db.session.query(db.func.coalesce(db.func.sum(InventoryMovement.total_cost), 0.0))
                .filter(InventoryMovement.company_id == cid)
                .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
                .filter(InventoryMovement.sale_ticket.in_(sale_tickets))
                .scalar()
            )
            cmv_total = _num(cmv_total)
    except Exception:
        cmv_total = 0.0

    income_total = _num(cur.get('income_total'))
    expense_total = _num(cur.get('expense_total'))
    result_total = _num(cur.get('result_total'))

    gross_margin = income_total - cmv_total
    gross_margin_pct = round(_safe_pct(gross_margin, income_total) * 100.0, 2)
    net_margin_pct = round(_safe_pct(result_total, income_total) * 100.0, 2)

    # Contribution margin ratio for breakeven
    contrib_ratio = _safe_pct(gross_margin, income_total)
    break_even_income = None
    if contrib_ratio > 1e-9:
        break_even_income = expense_total / contrib_ratio

    # Result accumulated (from series)
    result_acc = 0.0
    try:
        result_acc = sum(_num(r.get('result_total')) for r in (series or []))
    except Exception:
        result_acc = 0.0

    # Expense structure heuristic: fixed vs variable
    fixed_cost = 0.0
    variable_cost = 0.0
    fixed_payroll = 0.0
    fixed_recurring = 0.0
    try:
        exp_rows = (
            db.session.query(Expense)
            .filter(Expense.company_id == cid)
            .filter(Expense.expense_date >= d_from)
            .filter(Expense.expense_date <= d_to)
            .all()
        )
        for e in (exp_rows or []):
            if _is_supplier_cc_payment(e):
                continue
            amt = _num(getattr(e, 'amount', 0.0))
            if amt <= 0:
                continue
            if _is_payroll_expense(e):
                fixed_cost += amt
                fixed_payroll += amt
                continue
            freq = str(getattr(e, 'frequency', '') or '').strip().lower()
            if freq:
                fixed_cost += amt
                fixed_recurring += amt
            else:
                variable_cost += amt
    except Exception:
        fixed_cost = 0.0
        variable_cost = 0.0
        fixed_payroll = 0.0
        fixed_recurring = 0.0

    # Dependence on high-sales days
    day_concentration = {'top1_share_pct': 0.0, 'top3_share_pct': 0.0}
    try:
        rows = (
            db.session.query(Sale.sale_date, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .group_by(Sale.sale_date)
            .all()
        )
        vals = sorted([_num(v) for _, v in (rows or [])], reverse=True)
        tot = sum(vals) if vals else 0.0
        if tot > 1e-9:
            top1 = (vals[0] if len(vals) >= 1 else 0.0)
            top3 = sum(vals[:3])
            day_concentration = {
                'top1_share_pct': round(_safe_pct(top1, tot) * 100.0, 2),
                'top3_share_pct': round(_safe_pct(top3, tot) * 100.0, 2),
            }
    except Exception:
        day_concentration = {'top1_share_pct': 0.0, 'top3_share_pct': 0.0}

    def _top_n(d: dict, n: int):
        items = []
        for k, v in (d or {}).items():
            items.append({'label': str(k), 'value': round(_num(v), 2)})
        items = sorted(items, key=lambda x: _num(x.get('value')), reverse=True)
        return items[:n]

    sales_by_payment = {}
    exchanges_by_payment = {}
    try:
        rows = (
            db.session.query(Sale.payment_method, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .group_by(Sale.payment_method)
            .all()
        )
        for pm, amt in (rows or []):
            key = (str(pm or '').strip() or '—')
            sales_by_payment[key] = sales_by_payment.get(key, 0.0) + _num(amt)
    except Exception:
        sales_by_payment = {}

    try:
        rows = (
            db.session.query(Sale.payment_method, db.func.coalesce(db.func.sum(Sale.total), 0.0))
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .filter(Sale.sale_type == 'Cambio')
            .group_by(Sale.payment_method)
            .all()
        )
        for pm, amt in (rows or []):
            key = (str(pm or '').strip() or '—')
            exchanges_by_payment[key] = exchanges_by_payment.get(key, 0.0) + _num(amt)
    except Exception:
        exchanges_by_payment = {}

    net_sales_by_payment = {}
    for k in set(list(sales_by_payment.keys()) + list(exchanges_by_payment.keys())):
        net_sales_by_payment[k] = _num(sales_by_payment.get(k)) + _num(exchanges_by_payment.get(k))

    expenses_by_category = {}
    payroll_expenses = 0.0
    total_expenses = 0.0
    try:
        exp_rows = (
            db.session.query(Expense)
            .filter(Expense.expense_date >= d_from)
            .filter(Expense.expense_date <= d_to)
            .all()
        )
        for e in (exp_rows or []):
            if _is_supplier_cc_payment(e):
                continue
            amt = _num(getattr(e, 'amount', 0.0))
            if amt <= 0:
                continue
            total_expenses += amt
            if _is_payroll_expense(e):
                payroll_expenses += amt
                continue
            cat = (str(getattr(e, 'category', '') or '').strip() or 'Sin categoría')
            expenses_by_category[cat] = expenses_by_category.get(cat, 0.0) + amt
    except Exception:
        expenses_by_category = {}

    exp_cat_items = _top_n(expenses_by_category, 12)
    for it in exp_cat_items:
        it['pct_of_income'] = round(_safe_pct(it.get('value'), _num(cur.get('income_total'))) * 100.0, 2)

    insights = []
    try:
        # Result quality signals
        if _num(cur.get('income_total')) <= 1e-9:
            insights.append({
                'kind': 'info',
                'severity': 'media',
                'key': 'finance.no_income',
                'title': 'Sin ingresos en el período',
                'detail': 'No hay ventas registradas en el período seleccionado. El resultado depende completamente de los gastos.',
            })

        if _num(cur.get('result_total')) < 0:
            insights.append({
                'kind': 'warning',
                'severity': 'alta',
                'key': 'finance.negative_result',
                'title': 'Resultado negativo',
                'detail': 'El período cerró con resultado negativo. Ajustá gastos y/o mejorá el margen de ventas para volver al equilibrio.',
            })

        # Expense ratio
        ratio = _num(cur.get('expense_income_ratio'))
        if _num(cur.get('income_total')) > 1e-9 and ratio >= 0.75:
            insights.append({
                'kind': 'warning',
                'severity': 'alta' if ratio >= 0.9 else 'media',
                'key': 'finance.high_expense_ratio',
                'title': 'Gastos muy altos sobre ingresos',
                'detail': 'El ratio de gastos/ingresos es alto. Revisá el desglose de gastos y recortá lo no esencial.',
            })

        # Break-even gap
        if break_even_income is not None and _num(cur.get('income_total')) > 1e-9:
            gap = _num(break_even_income) - _num(cur.get('income_total'))
            if gap > max(1.0, _num(cur.get('income_total')) * 0.05):
                insights.append({
                    'kind': 'warning',
                    'severity': 'alta' if gap > _num(cur.get('income_total')) * 0.25 else 'media',
                    'key': 'finance.break_even_gap',
                    'title': 'Falta volumen para cubrir costos',
                    'detail': 'Con el margen bruto actual, falta volumen de ventas para cubrir los gastos del período.',
                    'key_data': {
                        'break_even_income': round(_num(break_even_income), 2),
                        'income_total': round(_num(cur.get('income_total')), 2),
                        'gap': round(_num(gap), 2),
                    },
                })

        # Concentration risk
        if _num(day_concentration.get('top3_share_pct')) >= 45.0:
            insights.append({
                'kind': 'warning',
                'severity': 'media',
                'key': 'finance.concentration_risk',
                'title': 'Dependencia de pocos días',
                'detail': 'Una parte importante de los ingresos se concentró en pocos días. Si esos días fallan, el período puede volverse negativo.',
                'key_data': day_concentration,
            })

        if not insights:
            insights.append({
                'kind': 'ok',
                'severity': 'baja',
                'key': 'finance.no_strong_signals',
                'title': 'Sin alertas fuertes',
                'detail': 'No se detectaron alertas críticas con las reglas actuales. Usá el punto de equilibrio y la estructura de gastos para afinar decisiones.',
            })
    except Exception:
        insights = [{'kind': 'info', 'severity': 'baja', 'title': 'Sin insights', 'detail': 'No se pudo generar insights para el período.'}]

    return jsonify({
        'ok': True,
        'period': {'from': d_from.isoformat(), 'to': d_to.isoformat()},
        'compare_mode': raw_compare if raw_compare in ('prev', 'yoy') else 'none',
        'kpis': {
            **cur,
            'cmv_total': round(_num(cmv_total), 2),
            'gross_margin_total': round(_num(gross_margin), 2),
            'gross_margin_pct': gross_margin_pct,
            'net_margin_pct': net_margin_pct,
            'result_accumulated': round(_num(result_acc), 2),
            'fixed_cost_estimate': round(_num(fixed_cost), 2),
            'variable_cost_estimate': round(_num(variable_cost), 2),
            'fixed_cost_payroll': round(_num(fixed_payroll), 2),
            'fixed_cost_recurring': round(_num(fixed_recurring), 2),
            'break_even_income_estimate': round(_num(break_even_income), 2) if break_even_income is not None else None,
            'day_concentration': day_concentration,
        },
        'kpis_prev': prev if prev else None,
        'deltas': deltas,
        'series': series,
        'trends': {},
        'months_summary': {},
        'breakdowns': {
            'net_sales_by_payment_method': _top_n(net_sales_by_payment, 10),
            'expenses_by_category': exp_cat_items,
        },
        'cash_balance': cash_balance,
        'projections': projections,
        'gap': gap,
        'insights': insights,
    })


@bp.get('/api/inventory_rotation')
@login_required
@module_required('reports')
def inventory_rotation_api():
    raw_from = (request.args.get('from') or '').strip()
    raw_to = (request.args.get('to') or '').strip()
    raw_compare = (request.args.get('compare') or '').strip().lower()  # none | prev | yoy
    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'period': {'from': '', 'to': ''}, 'compare_mode': 'none', 'kpis': {}, 'kpis_prev': None, 'deltas': {}, 'rows': [], 'sub': {}, 'insights': []})
    d_from = _parse_date_iso(raw_from, None)
    d_to = _parse_date_iso(raw_to, None)
    if not d_from or not d_to:
        today = dt_date.today()
        d_from = d_from or dt_date(today.year, today.month, 1)
        d_to = d_to or today

    try:
        period_days = int((d_to - d_from).days) + 1
        if period_days <= 0:
            period_days = 1
    except Exception:
        period_days = 1

    def _safe_div(a, b):
        b = _num(b)
        if abs(b) <= 1e-9:
            return 0.0
        return _num(a) / b

    green_days = _num(request.args.get('green_days') or 30)
    yellow_days = _num(request.args.get('yellow_days') or 60)
    if green_days <= 0:
        green_days = 30
    if yellow_days <= green_days:
        yellow_days = green_days * 2

    # Products & category map
    products = db.session.query(Product).filter(Product.company_id == cid).all()
    cat_map = {}
    try:
        for c in (db.session.query(Category).filter(Category.company_id == cid).all() or []):
            cat_map[int(c.id)] = c
    except Exception:
        cat_map = {}

    # Stock today from lots (qty_available + valuation)
    stock_now = {}
    stock_value_now = {}
    try:
        rows = (
            db.session.query(
                InventoryLot.product_id,
                db.func.coalesce(db.func.sum(InventoryLot.qty_available), 0.0),
                db.func.coalesce(db.func.sum(InventoryLot.qty_available * InventoryLot.unit_cost), 0.0),
            )
            .filter(InventoryLot.company_id == cid)
            .group_by(InventoryLot.product_id)
            .all()
        )
        for pid, qty, val in (rows or []):
            stock_now[int(pid)] = _num(qty)
            stock_value_now[int(pid)] = _num(val)
    except Exception:
        stock_now = {}
        stock_value_now = {}

    # Movement deltas
    delta_after_to = {}
    delta_within = {}
    try:
        rows = (
            db.session.query(
                InventoryMovement.product_id,
                db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0),
            )
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.movement_date > d_to)
            .group_by(InventoryMovement.product_id)
            .all()
        )
        for pid, s in (rows or []):
            delta_after_to[int(pid)] = _num(s)
    except Exception:
        delta_after_to = {}

    try:
        rows = (
            db.session.query(
                InventoryMovement.product_id,
                db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0),
            )
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.movement_date >= d_from)
            .filter(InventoryMovement.movement_date <= d_to)
            .group_by(InventoryMovement.product_id)
            .all()
        )
        for pid, s in (rows or []):
            delta_within[int(pid)] = _num(s)
    except Exception:
        delta_within = {}

    # Units sold (confirmed sales) and units returned
    units_sold = {}
    units_returned = {}
    try:
        rows = (
            db.session.query(
                InventoryMovement.product_id,
                db.func.coalesce(db.func.sum(-InventoryMovement.qty_delta), 0.0),
            )
            .join(Sale, Sale.ticket == InventoryMovement.sale_ticket)
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_type == 'Venta')
            .filter(db.func.lower(Sale.status).like('completad%'))
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .group_by(InventoryMovement.product_id)
            .all()
        )
        for pid, q in (rows or []):
            units_sold[int(pid)] = _num(q)
    except Exception:
        units_sold = {}

    try:
        rows = (
            db.session.query(
                InventoryMovement.product_id,
                db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0),
            )
            .join(Sale, Sale.ticket == InventoryMovement.sale_ticket)
            .filter(InventoryMovement.company_id == cid)
            .filter(InventoryMovement.type == 'return')
            .filter(Sale.company_id == cid)
            .filter(Sale.sale_type == 'Cambio')
            .filter(Sale.sale_date >= d_from)
            .filter(Sale.sale_date <= d_to)
            .group_by(InventoryMovement.product_id)
            .all()
        )
        for pid, q in (rows or []):
            units_returned[int(pid)] = _num(q)
    except Exception:
        units_returned = {}

    rows_out = []
    totals_stock_value = 0.0
    totals_dead_value = 0.0
    avg_days_list = []
    rot_list = []
    products_with_stock = 0

    for p in (products or []):
        pid = int(getattr(p, 'id', 0) or 0)
        if pid <= 0:
            continue
        active = bool(getattr(p, 'active', True))
        qty_now = _num(stock_now.get(pid))

        # Follow the requested "activo" filter, but keep items that still have stock.
        if (not active) and qty_now <= 1e-9:
            continue

        val_now = _num(stock_value_now.get(pid))
        totals_stock_value += val_now

        stock_end = qty_now - _num(delta_after_to.get(pid))
        stock_start = stock_end - _num(delta_within.get(pid))
        if stock_end < 0:
            stock_end = 0.0
        if stock_start < 0:
            stock_start = 0.0
        stock_avg = (stock_start + stock_end) / 2.0

        sold = _num(units_sold.get(pid))
        returned = _num(units_returned.get(pid))
        net_sold = max(0.0, sold - returned)

        rotation = _safe_div(net_sold, stock_avg) if stock_avg > 1e-9 else 0.0
        days_stock = (_safe_div(period_days, rotation) if rotation > 1e-9 else None)

        if qty_now > 1e-9:
            products_with_stock += 1
        if qty_now > 1e-9 and net_sold <= 1e-9:
            totals_dead_value += val_now

        if days_stock is not None and qty_now > 1e-9:
            avg_days_list.append(float(days_stock))
        if rotation > 1e-9:
            rot_list.append(float(rotation))

        status = 'verde'
        if qty_now <= 1e-9:
            status = 'sin_stock'
        elif net_sold <= 1e-9:
            status = 'rojo'
        elif days_stock is None:
            status = 'rojo'
        elif days_stock <= green_days:
            status = 'verde'
        elif days_stock <= yellow_days:
            status = 'amarillo'
        else:
            status = 'rojo'

        cat_name = ''
        try:
            cat_id = getattr(p, 'category_id', None)
            if cat_id is not None and int(cat_id) in cat_map:
                cat_name = str(getattr(cat_map[int(cat_id)], 'name', '') or '').strip()
        except Exception:
            cat_name = ''

        rows_out.append({
            'product_id': pid,
            'label': str(getattr(p, 'name', '') or '').strip() or str(pid),
            'category': cat_name or 'Sin categoría',
            'active': active,
            'units_sold': round(net_sold, 2),
            'stock_start': round(stock_start, 2),
            'stock_end': round(stock_end, 2),
            'stock_avg': round(stock_avg, 2),
            'rotation': round(rotation, 4),
            'days_stock': round(float(days_stock), 2) if days_stock is not None else None,
            'stock_qty': round(qty_now, 2),
            'stock_value': round(val_now, 2),
            'status': status,
            'reorder_point': round(_num(getattr(p, 'reorder_point', 0.0)), 2),
        })

    # Aggregates / KPIs
    avg_days = (sum(avg_days_list) / len(avg_days_list)) if avg_days_list else 0.0
    avg_rot = (sum(rot_list) / len(rot_list)) if rot_list else 0.0

    # Sub-analyses
    no_movement = [r for r in rows_out if _num(r.get('units_sold')) <= 1e-9 and _num(r.get('stock_qty')) > 1e-9]
    no_movement_sorted = sorted(no_movement, key=lambda x: _num(x.get('stock_value')), reverse=True)[:20]

    high_rot_low_stock = [
        r for r in rows_out
        if (_num(r.get('units_sold')) > 1e-9)
        and (_num(r.get('rotation')) >= (avg_rot * 1.5 if avg_rot > 0 else 2.0))
        and (_num(r.get('stock_end')) <= max(_num(r.get('reorder_point')), 0.0))
        and (_num(r.get('stock_qty')) > 1e-9)
    ]
    high_rot_low_stock = sorted(high_rot_low_stock, key=lambda x: _num(x.get('rotation')), reverse=True)[:20]

    overstock_low_sales = [
        r for r in rows_out
        if (_num(r.get('stock_qty')) > 1e-9)
        and ((_num(r.get('days_stock')) if r.get('days_stock') is not None else 999999) >= max(90.0, yellow_days))
    ]
    overstock_low_sales = sorted(overstock_low_sales, key=lambda x: _num(x.get('stock_value')), reverse=True)[:20]

    # Insights
    insights = []
    if totals_dead_value > 0.01:
        insights.append({
            'kind': 'warning',
            'severity': 'alta',
            'key': 'inventory.dead_stock_value',
            'title': 'Stock inmovilizado',
            'detail': f'Tenés ${round(totals_dead_value, 2):,.2f} inmovilizados en productos sin rotación.'.replace(',', 'X').replace('.', ',').replace('X', '.'),
        })

    if avg_rot > 0:
        top_rot = sorted([r for r in rows_out if _num(r.get('rotation')) > 0], key=lambda x: _num(x.get('rotation')), reverse=True)[:1]
        if top_rot:
            r0 = top_rot[0]
            insights.append({
                'kind': 'ok',
                'severity': 'media',
                'key': 'inventory.top_rotation',
                'title': 'Producto destacado',
                'detail': f'"{str(r0.get("label") or "").strip()}" rota {round(_num(r0.get("rotation")), 2)} veces en el período (vs prom. {round(avg_rot, 2)}).',
            })

    if not insights:
        insights.append({
            'kind': 'info',
            'severity': 'baja',
            'key': 'inventory.no_strong_signals',
            'title': 'Sin señales fuertes',
            'detail': 'No se detectaron señales relevantes con las reglas actuales. Ordená por Días de stock y por Stock $ para priorizar decisiones.',
        })

    # Sort main table by stock value desc by default
    rows_out = sorted(rows_out, key=lambda x: _num(x.get('stock_value')), reverse=True)

    # Comparison (prev / yoy) - KPIs only
    kpis_prev = None
    deltas = {}
    try:
        if raw_compare in ('prev', 'yoy'):
            prev_from = None
            prev_to = None
            if raw_compare == 'prev':
                span_days = int((d_to - d_from).days) + 1
                prev_to = d_from - timedelta(days=1)
                prev_from = prev_to - timedelta(days=max(1, span_days) - 1)
            else:
                def _month_end(d: dt_date) -> dt_date:
                    if d.month == 12:
                        nxt = dt_date(d.year + 1, 1, 1)
                    else:
                        nxt = dt_date(d.year, d.month + 1, 1)
                    return nxt - timedelta(days=1)

                def _add_months(d: dt_date, months: int) -> dt_date:
                    y = int(d.year)
                    m = int(d.month) + int(months)
                    while m > 12:
                        y += 1
                        m -= 12
                    while m < 1:
                        y -= 1
                        m += 12
                    last = _month_end(dt_date(y, m, 1)).day
                    dd = min(int(d.day), int(last))
                    return dt_date(y, m, dd)

                prev_from = _add_months(d_from, -12)
                prev_to = _add_months(d_to, -12)

            if prev_from and prev_to:
                try:
                    prev_days = int((prev_to - prev_from).days) + 1
                    if prev_days <= 0:
                        prev_days = 1
                except Exception:
                    prev_days = 1

                # Movement deltas
                delta_after_prev = {}
                delta_within_prev = {}
                try:
                    rows = (
                        db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0))
                        .filter(InventoryMovement.company_id == cid)
                        .filter(InventoryMovement.movement_date > prev_to)
                        .group_by(InventoryMovement.product_id)
                        .all()
                    )
                    for pid, s in (rows or []):
                        delta_after_prev[int(pid)] = _num(s)
                except Exception:
                    delta_after_prev = {}

                try:
                    rows = (
                        db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0))
                        .filter(InventoryMovement.company_id == cid)
                        .filter(InventoryMovement.movement_date >= prev_from)
                        .filter(InventoryMovement.movement_date <= prev_to)
                        .group_by(InventoryMovement.product_id)
                        .all()
                    )
                    for pid, s in (rows or []):
                        delta_within_prev[int(pid)] = _num(s)
                except Exception:
                    delta_within_prev = {}

                # Units sold/returned
                sold_prev = {}
                returned_prev = {}
                try:
                    rows = (
                        db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(-InventoryMovement.qty_delta), 0.0))
                        .join(Sale, Sale.ticket == InventoryMovement.sale_ticket)
                        .filter(InventoryMovement.company_id == cid)
                        .filter(InventoryMovement.type.in_(['sale', 'sale_adjust']))
                        .filter(Sale.company_id == cid)
                        .filter(Sale.sale_type == 'Venta')
                        .filter(db.func.lower(Sale.status).like('completad%'))
                        .filter(Sale.sale_date >= prev_from)
                        .filter(Sale.sale_date <= prev_to)
                        .group_by(InventoryMovement.product_id)
                        .all()
                    )
                    for pid, q in (rows or []):
                        sold_prev[int(pid)] = _num(q)
                except Exception:
                    sold_prev = {}

                try:
                    rows = (
                        db.session.query(InventoryMovement.product_id, db.func.coalesce(db.func.sum(InventoryMovement.qty_delta), 0.0))
                        .join(Sale, Sale.ticket == InventoryMovement.sale_ticket)
                        .filter(InventoryMovement.company_id == cid)
                        .filter(InventoryMovement.type == 'return')
                        .filter(Sale.company_id == cid)
                        .filter(Sale.sale_type == 'Cambio')
                        .filter(Sale.sale_date >= prev_from)
                        .filter(Sale.sale_date <= prev_to)
                        .group_by(InventoryMovement.product_id)
                        .all()
                    )
                    for pid, q in (rows or []):
                        returned_prev[int(pid)] = _num(q)
                except Exception:
                    returned_prev = {}

                prev_dead_value = 0.0
                prev_days_list = []
                prev_rot_list = []
                for p in (products or []):
                    pid = int(getattr(p, 'id', 0) or 0)
                    if pid <= 0:
                        continue
                    active = bool(getattr(p, 'active', True))
                    qty_now = _num(stock_now.get(pid))
                    if (not active) and qty_now <= 1e-9:
                        continue

                    stock_end = qty_now - _num(delta_after_prev.get(pid))
                    stock_start = stock_end - _num(delta_within_prev.get(pid))
                    if stock_end < 0:
                        stock_end = 0.0
                    if stock_start < 0:
                        stock_start = 0.0
                    stock_avg = (stock_start + stock_end) / 2.0

                    net_sold = max(0.0, _num(sold_prev.get(pid)) - _num(returned_prev.get(pid)))
                    rotation = _safe_div(net_sold, stock_avg) if stock_avg > 1e-9 else 0.0
                    days_stock = (_safe_div(prev_days, rotation) if rotation > 1e-9 else None)

                    if qty_now > 1e-9 and net_sold <= 1e-9:
                        prev_dead_value += _num(stock_value_now.get(pid))
                    if days_stock is not None and qty_now > 1e-9:
                        prev_days_list.append(float(days_stock))
                    if rotation > 1e-9:
                        prev_rot_list.append(float(rotation))

                prev_avg_days = (sum(prev_days_list) / len(prev_days_list)) if prev_days_list else 0.0
                prev_avg_rot = (sum(prev_rot_list) / len(prev_rot_list)) if prev_rot_list else 0.0

                kpis_prev = {
                    'period': {'from': prev_from.isoformat(), 'to': prev_to.isoformat(), 'days': int(prev_days)},
                    'avg_days_stock': round(prev_avg_days, 2),
                    'avg_rotation': round(prev_avg_rot, 4),
                    'dead_stock_value': round(prev_dead_value, 2),
                }

                def _delta(curv, prevv):
                    c = _num(curv)
                    p = _num(prevv)
                    return {
                        'abs': round(c - p, 4),
                        'pct': round((_safe_div((c - p), p) * 100.0) if abs(p) > 1e-9 else 0.0, 2),
                    }

                deltas = {
                    'avg_rotation': _delta(_num(kpis.get('avg_rotation')), _num(kpis_prev.get('avg_rotation'))),
                    'avg_days_stock': _delta(_num(kpis.get('avg_days_stock')), _num(kpis_prev.get('avg_days_stock'))),
                    'dead_stock_value': _delta(_num(kpis.get('dead_stock_value')), _num(kpis_prev.get('dead_stock_value'))),
                }
    except Exception:
        kpis_prev = None
        deltas = {}

    return jsonify({
        'ok': True,
        'compare_mode': raw_compare if raw_compare in ('prev', 'yoy') else 'none',
        'period': {'from': d_from.isoformat(), 'to': d_to.isoformat(), 'days': int(period_days)},
        'kpis': {
            'products_count': int(len(rows_out)),
            'products_with_stock': int(products_with_stock),
            'stock_value_total': round(totals_stock_value, 2),
            'dead_stock_value': round(totals_dead_value, 2),
            'avg_days_stock': round(avg_days, 2),
            'avg_rotation': round(avg_rot, 4),
            'thresholds': {'green_days': round(green_days, 2), 'yellow_days': round(yellow_days, 2)},
        },
        'kpis_prev': kpis_prev,
        'deltas': deltas,
        'rows': rows_out,
        'sub': {
            'no_movement': no_movement_sorted,
            'high_rotation_low_stock': high_rot_low_stock,
            'overstock_low_sales': overstock_low_sales,
        },
        'insights': insights,
    })


@bp.get('/api/lookups/categories')
@login_required
@module_required('reports')
def lookup_categories_api():
    q = (request.args.get('q') or '').strip().lower()
    limit = int(request.args.get('limit') or 100)
    if limit <= 0 or limit > 300:
        limit = 100

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    query = db.session.query(Category).filter(Category.company_id == cid).filter(Category.active.is_(True))
    if q:
        like = f"%{q}%"
        query = query.filter(Category.name.ilike(like))
    rows = query.order_by(Category.name.asc()).limit(limit).all()
    return jsonify({
        'ok': True,
        'items': [{'id': str(r.id), 'label': str(r.name or '').strip() or str(r.id)} for r in (rows or [])],
    })


@bp.get('/api/lookups/products')
@login_required
@module_required('reports')
def lookup_products_api():
    q = (request.args.get('q') or '').strip().lower()
    category_id = (request.args.get('category_id') or '').strip()
    limit = int(request.args.get('limit') or 100)
    if limit <= 0 or limit > 300:
        limit = 100

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    query = db.session.query(Product).filter(Product.company_id == cid).filter(Product.active.is_(True))
    if category_id:
        try:
            query = query.filter(Product.category_id == int(category_id))
        except Exception:
            query = query.filter(Product.category_id == category_id)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Product.name.ilike(like))
            | (Product.internal_code.ilike(like))
            | (Product.barcode.ilike(like))
        )
    rows = query.order_by(Product.name.asc()).limit(limit).all()
    return jsonify({
        'ok': True,
        'items': [
            {
                'id': str(r.id),
                'label': str(r.name or '').strip() or str(r.id),
                'category_id': str(getattr(r, 'category_id', '') or '') if getattr(r, 'category_id', None) is not None else None,
                'category': str(getattr(getattr(r, 'category', None), 'name', '') or '').strip() or None,
            }
            for r in (rows or [])
        ],
    })


@bp.get('/api/lookups/customers')
@login_required
@module_required('reports')
def lookup_customers_api():
    q = (request.args.get('q') or '').strip().lower()
    limit = int(request.args.get('limit') or 100)
    if limit <= 0 or limit > 300:
        limit = 100

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    query = db.session.query(Customer).filter(Customer.company_id == cid)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Customer.name.ilike(like))
            | (Customer.first_name.ilike(like))
            | (Customer.last_name.ilike(like))
            | (Customer.email.ilike(like))
            | (Customer.phone.ilike(like))
        )
    rows = query.order_by(Customer.updated_at.desc(), Customer.created_at.desc()).limit(limit).all()

    def _full_name(c: Customer) -> str:
        full = (str(getattr(c, 'first_name', '') or '').strip() + ' ' + str(getattr(c, 'last_name', '') or '').strip()).strip()
        return (str(getattr(c, 'name', '') or '').strip() or full or str(getattr(c, 'id', '') or '').strip())

    return jsonify({
        'ok': True,
        'items': [{'id': str(r.id), 'label': _full_name(r)} for r in (rows or [])],
    })


@bp.get('/api/lookups/payment_methods')
@login_required
@module_required('reports')
def lookup_payment_methods_api():
    q = (request.args.get('q') or '').strip().lower()
    limit = int(request.args.get('limit') or 50)
    if limit <= 0 or limit > 200:
        limit = 50

    cid = _company_id()
    if not cid:
        return jsonify({'ok': True, 'items': []})

    query = db.session.query(Sale.payment_method).filter(Sale.company_id == cid).distinct().order_by(Sale.payment_method.asc())
    rows = query.limit(400).all()
    items = []
    for (pm,) in (rows or []):
        s = str(pm or '').strip()
        if not s:
            continue
        if q and q not in s.lower():
            continue
        items.append({'id': s, 'label': s})
        if len(items) >= limit:
            break

    return jsonify({'ok': True, 'items': items})


@bp.post('/api/finance/export/pdf')
@login_required
@module_required('reports')
def finance_export_pdf_api():
    payload = request.get_json(silent=True) or {}
    finance = payload.get('finance') if isinstance(payload.get('finance'), dict) else {}
    k = finance.get('kpis') if isinstance(finance.get('kpis'), dict) else {}
    series = finance.get('series') if isinstance(finance.get('series'), list) else []
    insights = finance.get('insights') if isinstance(finance.get('insights'), list) else []
    breakdowns = finance.get('breakdowns') if isinstance(finance.get('breakdowns'), dict) else {}

    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
    except Exception:
        return jsonify({'ok': False, 'error': 'reportlab_missing'}), 400

    business_name, logo_path = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 16 * mm
    y = height - margin

    def _wrap_text(text: str, font_name: str, font_size: int, max_width: float):
        words = str(text or '').replace('\n', ' ').split()
        if not words:
            return ['']
        lines = []
        cur = ''
        for w in words:
            cand = (cur + ' ' + w).strip() if cur else w
            try:
                cand_w = pdfmetrics.stringWidth(cand, font_name, font_size)
            except Exception:
                cand_w = len(cand) * (font_size * 0.5)
            if cand_w <= max_width:
                cur = cand
                continue
            if cur:
                lines.append(cur)
            cur = w
        if cur:
            lines.append(cur)
        return lines or ['']

    def _draw_wrapped(text: str, x: float, y0: float, max_width: float, line_gap_mm: float = 5.0, font_name: str = 'Helvetica', font_size: int = 9):
        nonlocal y
        c.setFont(font_name, font_size)
        line_h = line_gap_mm * mm
        for ln in _wrap_text(text, font_name, font_size, max_width):
            if y0 < 20 * mm:
                c.showPage()
                y0 = height - margin
                y = y0
                c.setFont(font_name, font_size)
            c.drawString(x, y0, ln)
            y0 -= line_h
            y = y0
        return y0

    logo_reserved_w = 0.0
    if logo_path:
        try:
            logo_size = 14 * mm
            logo_reserved_w = logo_size + (4 * mm)
            c.drawImage(
                logo_path,
                width - margin - logo_size,
                y - 12 * mm,
                width=logo_size,
                height=logo_size,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass

    c.setFillColor(colors.HexColor('#0d1067'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y, 'Informe Financiero Complementario')
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 9)
    c.drawString(margin, y - 5 * mm, business_name)
    c.drawRightString(width - margin - logo_reserved_w, y - 5 * mm, 'Generado: ' + datetime.now().strftime('%Y-%m-%d %H:%M'))
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y - 9 * mm, width - margin, y - 9 * mm)
    y -= 15 * mm

    c.setFont('Helvetica', 10)
    c.drawString(margin, y, f"Período: {p_from} a {p_to}")
    y -= 10 * mm

    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'KPIs')
    y -= 6 * mm

    c.setFont('Helvetica', 10)
    summary_rows = [
        ('Ingresos totales', _num(k.get('income_total'))),
        ('Gastos totales', _num(k.get('expense_total'))),
        ('Resultado', _num(k.get('result_total'))),
        ('Resultado %', _num(k.get('result_pct'))),
        ('Ratio Gastos/Ingresos', _num(k.get('expense_income_ratio')) * 100.0),
    ]
    for lbl, val in summary_rows:
        if y < 25 * mm:
            c.showPage()
            y = height - margin
            c.setFont('Helvetica', 10)
        c.drawString(margin, y, lbl)
        if lbl.endswith('%'):
            c.drawRightString(width - margin, y, f"{val:.2f}%")
        elif 'Ratio' in lbl:
            c.drawRightString(width - margin, y, f"{val:.2f}%")
        else:
            c.drawRightString(width - margin, y, _format_currency_ars(val))
        y -= 6 * mm
    y -= 2 * mm

    if series:
        if y < 40 * mm:
            c.showPage()
            y = height - margin
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, 'Serie mensual')
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica-Bold', 9)
        c.setFillColor(colors.HexColor('#374151'))
        c.drawString(margin, y, 'Mes')
        c.drawRightString(width - margin, y, 'Resultado')
        c.drawRightString(width - margin - 40 * mm, y, 'Gastos')
        c.drawRightString(width - margin - 80 * mm, y, 'Ingresos')
        c.setFillColor(colors.black)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        for r in (series or [])[:18]:
            if y < 20 * mm:
                c.showPage()
                y = height - margin
                c.setFont('Helvetica', 9)
            m = str(r.get('month') or '')
            inc = _num(r.get('income_total'))
            exp = _num(r.get('expense_total'))
            resv = _num(r.get('result_total'))
            c.drawString(margin, y, m)
            c.drawRightString(width - margin - 80 * mm, y, _format_currency_ars(inc))
            c.drawRightString(width - margin - 40 * mm, y, _format_currency_ars(exp))
            c.drawRightString(width - margin, y, _format_currency_ars(resv))
            y -= 5.5 * mm
        y -= 2 * mm

    exp_cat = breakdowns.get('expenses_by_category') if isinstance(breakdowns.get('expenses_by_category'), list) else []
    if exp_cat:
        if y < 50 * mm:
            c.showPage()
            y = height - margin
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, 'Gastos por categoría (top)')
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        for it in exp_cat[:12]:
            if y < 20 * mm:
                c.showPage()
                y = height - margin
                c.setFont('Helvetica', 9)
            label = str(it.get('label') or '')
            val = _num(it.get('value'))
            pct = _num(it.get('pct_of_income'))
            s = f"{label}: {_format_currency_ars(val)} ({pct:.2f}%)"
            y = _draw_wrapped(s, margin, y, (width - margin) - margin, line_gap_mm=5.0, font_name='Helvetica', font_size=9)
            y -= 1 * mm

    if insights:
        if y < 45 * mm:
            c.showPage()
            y = height - margin
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, 'Insights')
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        max_w = (width - margin) - margin
        for it in (insights or [])[:10]:
            title = str(it.get('title') or '').strip()
            detail = str(it.get('detail') or it.get('description') or '').strip()
            s = (title + ': ' + detail).strip() if title else detail
            y = _draw_wrapped(s, margin, y, max_w, line_gap_mm=5.0, font_name='Helvetica', font_size=9)
            y -= 1 * mm

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"Finanzas_{p_from}_a_{p_to}.pdf" if p_from and p_to else "Finanzas.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@bp.post('/api/finance/export/excel')
@login_required
@module_required('reports')
def finance_export_excel_api():
    payload = request.get_json(silent=True) or {}
    finance = payload.get('finance') if isinstance(payload.get('finance'), dict) else {}
    k = finance.get('kpis') if isinstance(finance.get('kpis'), dict) else {}
    series = finance.get('series') if isinstance(finance.get('series'), list) else []
    breakdowns = finance.get('breakdowns') if isinstance(finance.get('breakdowns'), dict) else {}
    insights = finance.get('insights') if isinstance(finance.get('insights'), list) else []

    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({
            'ok': False,
            'error': 'openpyxl_missing',
            'message': 'No se pudo exportar a Excel porque falta la dependencia openpyxl. Instalá las dependencias del proyecto (requirements.txt) o ejecutá: pip install openpyxl==3.1.2',
        }), 400

    business_name, _ = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Finanzas'
    header_fill = PatternFill('solid', fgColor='0D1067')
    sub_fill = PatternFill('solid', fgColor='F3F4F6')

    ws['A1'] = business_name
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:D1')

    ws['A2'] = 'Informe Financiero Complementario'
    ws['A2'].font = Font(bold=True)
    ws.merge_cells('A2:D2')
    ws['A3'] = f"Período: {p_from} a {p_to}"
    ws.merge_cells('A3:D3')

    ws['A5'] = 'KPIs'
    ws['A5'].font = Font(bold=True)
    ws['A5'].fill = sub_fill
    ws.merge_cells('A5:D5')

    kpi_rows = [
        ('Ingresos totales', _num(k.get('income_total'))),
        ('Gastos totales', _num(k.get('expense_total'))),
        ('Resultado', _num(k.get('result_total'))),
        ('Resultado %', _num(k.get('result_pct')) / 100.0),
        ('Ratio Gastos/Ingresos', _num(k.get('expense_income_ratio'))),
    ]
    r0 = 6
    for i, (lbl, val) in enumerate(kpi_rows):
        rr = r0 + i
        ws[f'A{rr}'] = lbl
        ws[f'B{rr}'] = val
        ws[f'A{rr}'].font = Font(bold=True)
        ws[f'B{rr}'].number_format = '0.00' if 'Ratio' in lbl else '#,##0.00'
        if lbl.endswith('%'):
            ws[f'B{rr}'].number_format = '0.00%'

    start = r0 + len(kpi_rows) + 2
    ws[f'A{start}'] = 'Serie mensual'
    ws[f'A{start}'].font = Font(bold=True)
    ws[f'A{start}'].fill = sub_fill
    ws.merge_cells(f'A{start}:D{start}')

    head_row = start + 1
    ws[f'A{head_row}'] = 'Mes'
    ws[f'B{head_row}'] = 'Ingresos'
    ws[f'C{head_row}'] = 'Gastos'
    ws[f'D{head_row}'] = 'Resultado'
    for c in ('A', 'B', 'C', 'D'):
        ws[f'{c}{head_row}'].font = Font(bold=True)

    cur_row = head_row + 1
    for r in (series or []):
        ws[f'A{cur_row}'] = str(r.get('month') or '')
        ws[f'B{cur_row}'] = _num(r.get('income_total'))
        ws[f'C{cur_row}'] = _num(r.get('expense_total'))
        ws[f'D{cur_row}'] = _num(r.get('result_total'))
        ws[f'B{cur_row}'].number_format = '#,##0.00'
        ws[f'C{cur_row}'].number_format = '#,##0.00'
        ws[f'D{cur_row}'].number_format = '#,##0.00'
        cur_row += 1

    cur_row += 1
    exp_cat = breakdowns.get('expenses_by_category') if isinstance(breakdowns.get('expenses_by_category'), list) else []
    ws[f'A{cur_row}'] = 'Gastos por categoría (top)'
    ws[f'A{cur_row}'].font = Font(bold=True)
    ws[f'A{cur_row}'].fill = sub_fill
    ws.merge_cells(f'A{cur_row}:D{cur_row}')
    cur_row += 1
    ws[f'A{cur_row}'] = 'Categoría'
    ws[f'B{cur_row}'] = 'Monto'
    ws[f'C{cur_row}'] = '% sobre ingresos'
    for c in ('A', 'B', 'C'):
        ws[f'{c}{cur_row}'].font = Font(bold=True)
    cur_row += 1
    for it in (exp_cat or [])[:20]:
        ws[f'A{cur_row}'] = str(it.get('label') or '')
        ws[f'B{cur_row}'] = _num(it.get('value'))
        ws[f'C{cur_row}'] = _num(it.get('pct_of_income')) / 100.0
        ws[f'B{cur_row}'].number_format = '#,##0.00'
        ws[f'C{cur_row}'].number_format = '0.00%'
        cur_row += 1

    if insights:
        cur_row += 1
        ws[f'A{cur_row}'] = 'Insights'
        ws[f'A{cur_row}'].font = Font(bold=True)
        ws[f'A{cur_row}'].fill = sub_fill
        ws.merge_cells(f'A{cur_row}:D{cur_row}')
        cur_row += 1
        for it in insights[:30]:
            title = str(it.get('title') or '').strip()
            detail = str(it.get('detail') or it.get('description') or '').strip()
            ws[f'A{cur_row}'] = (title + ': ' + detail).strip() if title else detail
            ws.merge_cells(f'A{cur_row}:D{cur_row}')
            cur_row += 1

    for i, w in enumerate((22, 16, 16, 16), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"Finanzas_{p_from}_a_{p_to}.xlsx" if p_from and p_to else "Finanzas.xlsx"
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.post('/api/sales_analysis/export/pdf')
@login_required
@module_required('reports')
def sales_analysis_export_pdf_api():
    payload = request.get_json(silent=True) or {}
    sales = payload.get('sales') if isinstance(payload.get('sales'), dict) else {}

    k = sales.get('kpis') if isinstance(sales.get('kpis'), dict) else {}
    rows = sales.get('rows') if isinstance(sales.get('rows'), list) else []
    insights = sales.get('insights') if isinstance(sales.get('insights'), list) else []
    group_by = str(sales.get('group_by') or '').strip().lower()

    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
    except Exception:
        return jsonify({'ok': False, 'error': 'reportlab_missing'}), 400

    business_name, logo_path = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')
    g_label = 'Por producto' if group_by != 'category' else 'Por categoría'

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 16 * mm
    y = height - margin

    def _wrap_text(text: str, font_name: str, font_size: int, max_width: float):
        words = str(text or '').replace('\n', ' ').split()
        if not words:
            return ['']
        lines = []
        cur = ''
        for w in words:
            cand = (cur + ' ' + w).strip() if cur else w
            try:
                cand_w = pdfmetrics.stringWidth(cand, font_name, font_size)
            except Exception:
                cand_w = len(cand) * (font_size * 0.5)
            if cand_w <= max_width:
                cur = cand
                continue
            if cur:
                lines.append(cur)
            cur = w
        if cur:
            lines.append(cur)
        return lines or ['']

    def _draw_wrapped(text: str, x: float, y0: float, max_width: float, line_gap_mm: float = 5.0, font_name: str = 'Helvetica', font_size: int = 9):
        nonlocal y
        c.setFont(font_name, font_size)
        line_h = line_gap_mm * mm
        for ln in _wrap_text(text, font_name, font_size, max_width):
            if y0 < 20 * mm:
                c.showPage()
                y0 = height - margin
                y = y0
                c.setFont(font_name, font_size)
            c.drawString(x, y0, ln)
            y0 -= line_h
            y = y0
        return y0

    # Header
    logo_reserved_w = 0.0
    if logo_path:
        try:
            logo_size = 14 * mm
            logo_reserved_w = logo_size + (4 * mm)
            c.drawImage(
                logo_path,
                width - margin - logo_size,
                y - 12 * mm,
                width=logo_size,
                height=logo_size,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass

    c.setFillColor(colors.HexColor('#0d1067'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y, 'Análisis de Ventas (Ventas vs Margen)')
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 9)
    c.drawString(margin, y - 5 * mm, business_name)
    c.drawRightString(width - margin - logo_reserved_w, y - 5 * mm, 'Generado: ' + datetime.now().strftime('%Y-%m-%d %H:%M'))
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y - 9 * mm, width - margin, y - 9 * mm)
    y -= 15 * mm

    c.setFont('Helvetica', 10)
    c.drawString(margin, y, f"Período: {p_from} a {p_to} · {g_label}")
    y -= 10 * mm

    # KPIs
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Resumen ejecutivo')
    y -= 6 * mm
    c.setFont('Helvetica', 10)
    summary_rows = [
        ('Ventas totales', _num(k.get('sales_total'))),
        ('CMV total', _num(k.get('cmv_total'))),
        ('Margen bruto', _num(k.get('gross_margin_total'))),
        ('Margen %', _num(k.get('gross_margin_pct'))),
        ('Margen promedio por venta', _num(k.get('avg_margin_per_sale'))),
        ('Cantidad de ventas', int(k.get('sales_count') or 0)),
    ]
    for lbl, val in summary_rows:
        if y < 25 * mm:
            c.showPage()
            y = height - margin
            c.setFont('Helvetica', 10)
        c.drawString(margin, y, lbl)
        if isinstance(val, int) and lbl == 'Cantidad de ventas':
            c.drawRightString(width - margin, y, str(val))
        elif lbl.endswith('%'):
            c.drawRightString(width - margin, y, f"{val:.2f}%")
        else:
            c.drawRightString(width - margin, y, _format_currency_ars(val))
        y -= 6 * mm
    y -= 3 * mm

    # Tabla
    if y < 40 * mm:
        c.showPage()
        y = height - margin
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Detalle')
    y -= 6 * mm
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y, width - margin, y)
    y -= 6 * mm

    # Column layout
    col_label_x = margin
    col_sales_x = width - margin
    col_margen_x = width - margin - 40 * mm
    col_margen_pct_x = width - margin - 68 * mm
    col_cmv_x = width - margin - 96 * mm

    c.setFont('Helvetica-Bold', 9)
    c.setFillColor(colors.HexColor('#374151'))
    c.drawString(col_label_x, y, 'Producto/Categoría')
    c.drawRightString(col_cmv_x, y, 'CMV')
    c.drawRightString(col_margen_pct_x, y, 'Margen %')
    c.drawRightString(col_margen_x, y, 'Margen')
    c.drawRightString(col_sales_x, y, 'Ventas')
    c.setFillColor(colors.black)
    y -= 6 * mm

    c.setFont('Helvetica', 9)
    for r in (rows or [])[:55]:
        if y < 20 * mm:
            c.showPage()
            y = height - margin
            c.setFont('Helvetica', 9)
        label = str(r.get('label') or '—')
        cmv = -abs(_num(r.get('cmv')))
        sales_amt = _num(r.get('sales'))
        margin_amt = _num(r.get('margin'))
        margin_pct = _num(r.get('margin_pct'))

        max_w = (col_cmv_x - 4 * mm) - col_label_x
        label_lines = _wrap_text(label, 'Helvetica', 9, max_w)
        c.drawString(col_label_x, y, label_lines[0][:70])
        c.drawRightString(col_cmv_x, y, _format_currency_ars(cmv))
        c.drawRightString(col_margen_pct_x, y, f"{margin_pct:.2f}%")
        c.drawRightString(col_margen_x, y, _format_currency_ars(margin_amt))
        c.drawRightString(col_sales_x, y, _format_currency_ars(sales_amt))
        y -= 5.5 * mm
    y -= 2 * mm

    # Insights
    if insights:
        if y < 45 * mm:
            c.showPage()
            y = height - margin
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, 'Insights')
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        max_w = (width - margin) - margin
        for it in (insights or [])[:10]:
            title = str(it.get('title') or '').strip()
            detail = str(it.get('detail') or it.get('description') or '').strip()
            s = (title + ': ' + detail).strip() if title else detail
            y = _draw_wrapped(s, margin, y, max_w, line_gap_mm=5.0, font_name='Helvetica', font_size=9)
            y -= 1 * mm

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"Ventas_{p_from}_a_{p_to}.pdf" if p_from and p_to else "Ventas.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@bp.post('/api/sales_analysis/export/excel')
@login_required
@module_required('reports')
def sales_analysis_export_excel_api():
    payload = request.get_json(silent=True) or {}
    sales = payload.get('sales') if isinstance(payload.get('sales'), dict) else {}

    k = sales.get('kpis') if isinstance(sales.get('kpis'), dict) else {}
    rows = sales.get('rows') if isinstance(sales.get('rows'), list) else []
    insights = sales.get('insights') if isinstance(sales.get('insights'), list) else []
    group_by = str(sales.get('group_by') or '').strip().lower()

    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({
            'ok': False,
            'error': 'openpyxl_missing',
            'message': 'No se pudo exportar a Excel porque falta la dependencia openpyxl. Instalá las dependencias del proyecto (requirements.txt) o ejecutá: pip install openpyxl==3.1.2',
        }), 400

    business_name, _ = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')
    g_label = 'producto' if group_by != 'category' else 'categoría'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ventas vs Margen'
    header_fill = PatternFill('solid', fgColor='0D1067')
    sub_fill = PatternFill('solid', fgColor='F3F4F6')

    ws['A1'] = business_name
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Análisis de Ventas (Ventas vs Margen)'
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:H2')

    ws['A3'] = 'Período'
    ws['B3'] = f"{p_from} a {p_to}"
    ws['A4'] = 'Agrupación'
    ws['B4'] = g_label
    ws['A5'] = 'Generado'
    ws['B5'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    ws.append([])
    ws.append(['Resumen', 'Valor'])
    ws['A7'].font = Font(bold=True)
    ws['B7'].font = Font(bold=True)
    ws['A7'].fill = sub_fill
    ws['B7'].fill = sub_fill
    ws.append(['Ventas totales', _num(k.get('sales_total'))])
    ws.append(['CMV total', _num(k.get('cmv_total'))])
    ws.append(['Margen bruto', _num(k.get('gross_margin_total'))])
    ws.append(['Margen %', _num(k.get('gross_margin_pct'))])
    ws.append(['Margen promedio por venta', _num(k.get('avg_margin_per_sale'))])
    ws.append(['Cantidad de ventas', int(k.get('sales_count') or 0)])

    ws.append([])
    ws.append(['Detalle', '', '', '', '', '', '', ''])
    ws.merge_cells(f"A{ws.max_row}:H{ws.max_row}")
    ws[f"A{ws.max_row}"].font = Font(bold=True)
    ws[f"A{ws.max_row}"].fill = sub_fill

    ws.append(['Label', 'Categoría', 'Ventas', '% Ventas', 'CMV', 'Margen', 'Margen %', 'Cantidad'])
    header_row = ws.max_row
    for col in range(1, 9):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')

    for r in (rows or []):
        ws.append([
            str(r.get('label') or ''),
            str(r.get('category') or ''),
            _num(r.get('sales')),
            _num(r.get('sales_pct')),
            -abs(_num(r.get('cmv'))),
            _num(r.get('margin')),
            _num(r.get('margin_pct')),
            _num(r.get('qty')),
        ])

    # Formatting
    for col in [3, 5, 6]:
        for rr in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=rr, column=col).number_format = '$ #,##0.00'
    for col in [4, 7]:
        for rr in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=rr, column=col).number_format = '0.00"%"'
    for rr in range(header_row + 1, ws.max_row + 1):
        for col in range(1, 9):
            ws.cell(row=rr, column=col).alignment = Alignment(vertical='top')

    # Auto width
    for col in range(1, 9):
        max_len = 10
        for rr in range(1, ws.max_row + 1):
            v = ws.cell(row=rr, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)) if len(str(v)) < 60 else 60)
        ws.column_dimensions[get_column_letter(col)].width = min(48, max_len + 2)

    if insights:
        ws2 = wb.create_sheet('Insights')
        ws2.append(['Título', 'Detalle', 'Severidad', 'Acción sugerida', 'Regla'])
        for col in range(1, 6):
            cell = ws2.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = sub_fill
        for it in (insights or [])[:30]:
            ws2.append([
                str(it.get('title') or ''),
                str(it.get('detail') or it.get('description') or ''),
                str(it.get('severity') or ''),
                str(it.get('suggested_action') or ''),
                str(it.get('rule') or ''),
            ])
        ws2.column_dimensions['A'].width = 32
        ws2.column_dimensions['B'].width = 80
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 46
        ws2.column_dimensions['E'].width = 46
        for rr in range(2, ws2.max_row + 1):
            ws2.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"Ventas_{p_from}_a_{p_to}.xlsx" if p_from and p_to else "Ventas.xlsx"
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.post('/api/eerr/export/pdf')
@login_required
@module_required('reports')
def eerr_export_pdf_api():
    payload = request.get_json(silent=True) or {}
    eerr = payload.get('eerr') if isinstance(payload.get('eerr'), dict) else {}
    expanded = payload.get('expanded') if isinstance(payload.get('expanded'), dict) else {}

    k = eerr.get('kpis') if isinstance(eerr.get('kpis'), dict) else {}
    b = eerr.get('breakdowns') if isinstance(eerr.get('breakdowns'), dict) else {}
    insights = eerr.get('insights') if isinstance(eerr.get('insights'), list) else []
    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
    except Exception:
        return jsonify({'ok': False, 'error': 'reportlab_missing'}), 400

    business_name, logo_path = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 16 * mm
    y = height - margin

    def _wrap_text(text: str, font_name: str, font_size: int, max_width: float):
        words = str(text or '').replace('\n', ' ').split()
        if not words:
            return ['']
        lines = []
        cur = ''
        for w in words:
            cand = (cur + ' ' + w).strip() if cur else w
            try:
                cand_w = pdfmetrics.stringWidth(cand, font_name, font_size)
            except Exception:
                cand_w = len(cand) * (font_size * 0.5)
            if cand_w <= max_width:
                cur = cand
                continue
            if cur:
                lines.append(cur)
            cur = w
        if cur:
            lines.append(cur)
        return lines or ['']

    def _draw_wrapped(text: str, x: float, y0: float, max_width: float, line_gap_mm: float = 5.0, font_name: str = 'Helvetica', font_size: int = 9):
        nonlocal y
        c.setFont(font_name, font_size)
        line_h = line_gap_mm * mm
        for ln in _wrap_text(text, font_name, font_size, max_width):
            if y0 < 20 * mm:
                c.showPage()
                y0 = height - margin
                y = y0
                c.setFont(font_name, font_size)
            c.drawString(x, y0, ln)
            y0 -= line_h
            y = y0
        return y0

    # Header
    logo_reserved_w = 0.0
    if logo_path:
        try:
            logo_size = 14 * mm
            logo_reserved_w = logo_size + (4 * mm)
            c.drawImage(
                logo_path,
                width - margin - logo_size,
                y - 12 * mm,
                width=logo_size,
                height=logo_size,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass
    c.setFillColor(colors.HexColor('#0d1067'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y, 'Estado de Resultados')
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 9)
    c.drawString(margin, y - 5 * mm, business_name)
    c.drawRightString(width - margin - logo_reserved_w, y - 5 * mm, 'Generado: ' + datetime.now().strftime('%Y-%m-%d %H:%M'))
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y - 9 * mm, width - margin, y - 9 * mm)
    y -= 15 * mm

    c.setFont('Helvetica', 10)
    c.drawString(margin, y, f"Período: {p_from} a {p_to}")
    y -= 10 * mm

    # Resumen ejecutivo
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Resumen ejecutivo')
    y -= 6 * mm
    c.setFont('Helvetica', 10)
    summary_rows = [
        ('Ventas netas', _num(k.get('sales_net'))),
        ('Margen bruto', _num(k.get('gross_margin'))),
        ('Gastos operativos', -abs(_num(k.get('operating_expenses')))),
        ('Resultado neto', _num(k.get('net_result'))),
    ]
    for lbl, val in summary_rows:
        c.drawString(margin, y, lbl)
        c.drawRightString(width - margin, y, _format_currency_ars(val))
        y -= 6 * mm
    y -= 4 * mm

    # Tabla EERR
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Estado de Resultados')
    y -= 6 * mm
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y, width - margin, y)
    y -= 6 * mm
    c.setFont('Helvetica', 10)
    table_rows = [
        ('Ventas netas', _num(k.get('sales_net'))),
        ('CMV', -abs(_num(k.get('cmv')))),
        ('Margen bruto', _num(k.get('gross_margin'))),
        ('Gastos operativos', -abs(_num(k.get('operating_expenses')))),
        ('Nómina', -abs(_num(k.get('payroll_expenses')))),
        ('Resultado neto', _num(k.get('net_result'))),
    ]
    for lbl, val in table_rows:
        c.drawString(margin, y, lbl)
        c.drawRightString(width - margin, y, _format_currency_ars(val))
        y -= 6 * mm
    y -= 4 * mm

    def draw_detail(title, items):
        nonlocal y
        if y < 40 * mm:
            c.showPage()
            y = height - margin
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, title)
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        for it in (items or [])[:14]:
            if y < 25 * mm:
                break
            c.drawString(margin, y, str(it.get('key') or '—')[:55])
            c.drawRightString(width - margin, y, _format_currency_ars(it.get('amount')))
            y -= 5 * mm
        y -= 3 * mm

    # Detalles (solo lo expandido)
    if expanded.get('income') is True:
        draw_detail('Detalle Ingresos (ventas netas por medio de pago)', b.get('net_sales_by_payment_method') if isinstance(b.get('net_sales_by_payment_method'), list) else [])
    if expanded.get('cmv') is True:
        draw_detail('Detalle CMV (top productos por costo)', b.get('top_products_by_cmv') if isinstance(b.get('top_products_by_cmv'), list) else [])
    if expanded.get('opex') is True:
        draw_detail('Detalle Gastos (por categoría)', b.get('expenses_by_category') if isinstance(b.get('expenses_by_category'), list) else [])

    # Insights
    if y < 45 * mm:
        c.showPage()
        y = height - margin
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Insights')
    y -= 6 * mm
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y, width - margin, y)
    y -= 6 * mm
    c.setFont('Helvetica', 9)
    for it in (insights or [])[:10]:
        title = str(it.get('title') or '').strip()
        detail = str(it.get('detail') or '').strip()
        s = (title + ': ' + detail).strip() if title else detail
        max_w = (width - margin) - margin
        y = _draw_wrapped(s, margin, y, max_w, line_gap_mm=5.0, font_name='Helvetica', font_size=9)
        y -= 1 * mm

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"EERR_{p_from}_a_{p_to}.pdf" if p_from and p_to else "EERR.pdf"
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@bp.post('/api/eerr/export/excel')
@login_required
@module_required('reports')
def eerr_export_excel_api():
    payload = request.get_json(silent=True) or {}
    eerr = payload.get('eerr') if isinstance(payload.get('eerr'), dict) else {}
    expanded = payload.get('expanded') if isinstance(payload.get('expanded'), dict) else {}

    k = eerr.get('kpis') if isinstance(eerr.get('kpis'), dict) else {}
    b = eerr.get('breakdowns') if isinstance(eerr.get('breakdowns'), dict) else {}
    insights = eerr.get('insights') if isinstance(eerr.get('insights'), list) else []
    if not k:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({
            'ok': False,
            'error': 'openpyxl_missing',
            'message': 'No se pudo exportar a Excel porque falta la dependencia openpyxl. Instalá las dependencias del proyecto (requirements.txt) o ejecutá: pip install openpyxl==3.1.2',
        }), 400

    business_name, _ = _get_business_info()
    period = k.get('period') if isinstance(k.get('period'), dict) else {}
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen EERR'
    header_fill = PatternFill('solid', fgColor='0D1067')
    sub_fill = PatternFill('solid', fgColor='F3F4F6')

    ws['A1'] = business_name
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A2'] = 'Estado de Resultados'
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')
    ws['A3'] = 'Período'
    ws['B3'] = f"{p_from} a {p_to}"
    ws['A4'] = 'Generado'
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    ws.append([])
    ws.append(['Concepto', 'Monto'])
    ws['A6'].font = Font(bold=True)
    ws['B6'].font = Font(bold=True)
    ws['A6'].fill = sub_fill
    ws['B6'].fill = sub_fill

    base_rows = [
        ('Ventas netas', _num(k.get('sales_net'))),
        ('CMV', -abs(_num(k.get('cmv')))),
        ('Margen bruto', _num(k.get('gross_margin'))),
        ('Gastos operativos', -abs(_num(k.get('operating_expenses')))),
        ('Nómina', -abs(_num(k.get('payroll_expenses')))),
        ('Resultado neto', _num(k.get('net_result'))),
    ]
    for lbl, val in base_rows:
        ws.append([lbl, float(val)])
    for r in range(7, 13):
        ws[f'B{r}'].number_format = '"$" #,##0.00'

    ws.append([])
    ws.append(['Insights'])
    ws['A14'].font = Font(bold=True)
    ws.merge_cells('A14:F14')
    for it in (insights or [])[:20]:
        s = (str(it.get('title') or '').strip() + ' - ' + str(it.get('detail') or '').strip()).strip(' -')
        if s:
            ws.append([s])
            ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 22

    def add_sheet(title, header, rows_data, money_cols=None, pct_cols=None):
        sh = wb.create_sheet(title)
        sh.append(header)
        for cell in sh[1]:
            cell.font = Font(bold=True)
            cell.fill = sub_fill
            cell.alignment = Alignment(horizontal='center')
        for r in rows_data:
            sh.append(r)
        if money_cols:
            for col in money_cols:
                for cell in sh[get_column_letter(col)][1:]:
                    cell.number_format = '"$" #,##0.00'
        if pct_cols:
            for col in pct_cols:
                for cell in sh[get_column_letter(col)][1:]:
                    cell.number_format = '0.00%'
        for i in range(1, len(header) + 1):
            sh.column_dimensions[get_column_letter(i)].width = 20
        return sh

    # Ingresos (solo si expandido)
    if expanded.get('income') is True:
        pms = b.get('net_sales_by_payment_method') if isinstance(b.get('net_sales_by_payment_method'), list) else []
        rows_data = [[str(x.get('key') or ''), float(_num(x.get('amount')))] for x in pms]
        add_sheet('Ingresos', ['Medio de pago', 'Ventas netas $'], rows_data, money_cols=[2])

    # CMV (solo si expandido)
    if expanded.get('cmv') is True:
        items = b.get('top_products_by_cmv') if isinstance(b.get('top_products_by_cmv'), list) else []
        rows_data = [[str(x.get('key') or ''), float(_num(x.get('qty'))), float(_num(x.get('unit_cost'))), float(_num(x.get('amount')))] for x in items]
        sh = add_sheet('CMV', ['Producto', 'Cantidad', 'Costo unit.', 'CMV total'], rows_data, money_cols=[3, 4])
        for cell in sh['C'][1:]:
            cell.number_format = '"$" #,##0.0000'

    # Gastos (solo si expandido)
    if expanded.get('opex') is True:
        items = b.get('expenses_by_category') if isinstance(b.get('expenses_by_category'), list) else []
        rows_data = [[str(x.get('key') or ''), float(_num(x.get('amount')))] for x in items]
        add_sheet('Gastos', ['Categoría', 'Monto $'], rows_data, money_cols=[2])

        pay = b.get('payroll_by_employee') if isinstance(b.get('payroll_by_employee'), list) else []
        rows_data = [[str(x.get('key') or ''), float(_num(x.get('amount')))] for x in pay]
        add_sheet('Nómina', ['Empleado', 'Monto $'], rows_data, money_cols=[2])

    # Margen (derivado del mismo objeto, sin DB)
    rev = b.get('top_products_by_revenue') if isinstance(b.get('top_products_by_revenue'), list) else []
    cmv_top = b.get('top_products_by_cmv') if isinstance(b.get('top_products_by_cmv'), list) else []
    rev_map = {str(x.get('key') or '').strip(): _num(x.get('amount')) for x in rev}
    cmv_map = {str(x.get('key') or '').strip(): _num(x.get('amount')) for x in cmv_top}
    keys = set(list(rev_map.keys()) + list(cmv_map.keys()))
    rows_data = []
    for name in sorted(keys):
        v = _num(rev_map.get(name))
        cst = _num(cmv_map.get(name))
        m = v - cst
        mp = (m / v) if abs(v) > 1e-9 else 0.0
        rows_data.append([name, float(v), float(cst), float(m), float(mp)])
    add_sheet('Margen', ['Producto', 'Ventas $', 'CMV $', 'Margen $', 'Margen %'], rows_data, money_cols=[2, 3, 4], pct_cols=[5])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"EERR_{p_from}_a_{p_to}.xlsx" if p_from and p_to else "EERR.xlsx"
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@bp.post('/api/inventory_rotation/export/pdf')
@login_required
@module_required('reports')
def inventory_rotation_export_pdf_api():
    payload = request.get_json(silent=True) or {}
    inv = payload.get('inventory') if isinstance(payload.get('inventory'), dict) else {}

    period = inv.get('period') if isinstance(inv.get('period'), dict) else {}
    k = inv.get('kpis') if isinstance(inv.get('kpis'), dict) else {}
    rows = inv.get('rows') if isinstance(inv.get('rows'), list) else []
    insights = inv.get('insights') if isinstance(inv.get('insights'), list) else []
    sub = inv.get('sub') if isinstance(inv.get('sub'), dict) else {}

    if not k or not rows:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.pdfbase import pdfmetrics
    except Exception:
        return jsonify({'ok': False, 'error': 'reportlab_missing'}), 400

    business_name, logo_path = _get_business_info()
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4
    margin = 16 * mm
    y = height - margin

    def _wrap_text(text: str, font_name: str, font_size: int, max_width: float):
        words = str(text or '').replace('\n', ' ').split()
        if not words:
            return ['']
        lines = []
        cur = ''
        for w in words:
            cand = (cur + ' ' + w).strip() if cur else w
            try:
                cand_w = pdfmetrics.stringWidth(cand, font_name, font_size)
            except Exception:
                cand_w = len(cand) * (font_size * 0.5)
            if cand_w <= max_width:
                cur = cand
                continue
            if cur:
                lines.append(cur)
            cur = w
        if cur:
            lines.append(cur)
        return lines or ['']

    def _new_page():
        nonlocal y
        c.showPage()
        y = height - margin

    # Header
    logo_reserved_w = 0.0
    if logo_path:
        try:
            logo_size = 14 * mm
            logo_reserved_w = logo_size + (4 * mm)
            c.drawImage(
                logo_path,
                width - margin - logo_size,
                y - 12 * mm,
                width=logo_size,
                height=logo_size,
                preserveAspectRatio=True,
                mask='auto',
            )
        except Exception:
            pass
    c.setFillColor(colors.HexColor('#0d1067'))
    c.setFont('Helvetica-Bold', 16)
    c.drawString(margin, y, 'Inventario - Rotación de stock')
    c.setFillColor(colors.black)
    c.setFont('Helvetica', 9)
    c.drawString(margin, y - 5 * mm, business_name)
    c.drawRightString(width - margin - logo_reserved_w, y - 5 * mm, 'Generado: ' + datetime.now().strftime('%Y-%m-%d %H:%M'))
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y - 9 * mm, width - margin, y - 9 * mm)
    y -= 15 * mm

    c.setFont('Helvetica', 10)
    c.drawString(margin, y, f"Período: {p_from} a {p_to}")
    y -= 9 * mm

    # KPIs
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Resumen')
    y -= 6 * mm
    c.setFont('Helvetica', 9)
    kpi_rows = [
        ('Stock total (valor)', _format_currency_ars(_num(k.get('stock_value_total')))),
        ('Stock inmovilizado', _format_currency_ars(_num(k.get('dead_stock_value')))),
        ('Rotación promedio', str(round(_num(k.get('avg_rotation')), 2)).replace('.', ',')),
        ('Días stock prom.', str(int(round(_num(k.get('avg_days_stock')), 0)))),
        ('Productos con stock', str(int(k.get('products_with_stock') or 0))),
    ]
    for lbl, val in kpi_rows:
        c.setFillColor(colors.black)
        c.drawString(margin, y, lbl)
        c.drawRightString(width - margin, y, str(val))
        y -= 5.5 * mm
    y -= 2 * mm

    # Main table
    if y < 55 * mm:
        _new_page()
    c.setFont('Helvetica-Bold', 11)
    c.drawString(margin, y, 'Detalle (ordenado por Stock $)')
    y -= 6 * mm
    c.setStrokeColor(colors.HexColor('#e5e7eb'))
    c.line(margin, y, width - margin, y)
    y -= 6 * mm

    headers = ['Producto', 'Categoría', 'Unid.', 'Rot.', 'Días', 'Stock $', 'Estado']
    col_x = [margin, margin + 62 * mm, margin + 104 * mm, margin + 118 * mm, margin + 132 * mm, margin + 150 * mm, margin + 175 * mm]
    c.setFont('Helvetica-Bold', 8)
    for i, htxt in enumerate(headers):
        c.drawString(col_x[i], y, htxt)
    y -= 4.5 * mm
    c.setFont('Helvetica', 8)
    c.setStrokeColor(colors.HexColor('#f3f4f6'))
    c.line(margin, y, width - margin, y)
    y -= 4.0 * mm

    def _status_label(s):
        ss = str(s or '').lower()
        if ss == 'verde':
            return 'Verde'
        if ss == 'amarillo':
            return 'Amarillo'
        if ss == 'rojo':
            return 'Rojo'
        if ss == 'sin_stock':
            return 'Sin stock'
        return ss[:18]

    rows_sorted = sorted(rows, key=lambda r: _num((r or {}).get('stock_value')), reverse=True)
    for r in rows_sorted[:35]:
        if y < 18 * mm:
            _new_page()
            c.setFont('Helvetica-Bold', 8)
            for i, htxt in enumerate(headers):
                c.drawString(col_x[i], y, htxt)
            y -= 4.5 * mm
            c.setFont('Helvetica', 8)
            c.setStrokeColor(colors.HexColor('#f3f4f6'))
            c.line(margin, y, width - margin, y)
            y -= 4.0 * mm

        name = str(r.get('label') or '—')
        cat = str(r.get('category') or '—')
        units = str(round(_num(r.get('units_sold')), 2)).replace('.', ',')
        rot = str(round(_num(r.get('rotation')), 2)).replace('.', ',')
        days = '—'
        if r.get('days_stock') is not None:
            try:
                days = str(int(round(_num(r.get('days_stock')), 0)))
            except Exception:
                days = str(r.get('days_stock'))
        stock_val = _format_currency_ars(_num(r.get('stock_value')))
        st = _status_label(r.get('status'))

        max_name_w = col_x[1] - col_x[0] - 2
        name_lines = _wrap_text(name, 'Helvetica', 8, max_name_w)
        c.drawString(col_x[0], y, name_lines[0][:40])
        c.drawString(col_x[1], y, cat[:26])
        c.drawRightString(col_x[2] + 10, y, units)
        c.drawRightString(col_x[3] + 10, y, rot)
        c.drawRightString(col_x[4] + 10, y, days)
        c.drawRightString(col_x[5] + 20, y, stock_val)
        c.drawString(col_x[6], y, st)
        y -= 5.0 * mm
        for extra in name_lines[1:2]:
            if y < 18 * mm:
                break
            c.drawString(col_x[0], y, extra[:45])
            y -= 5.0 * mm

    # Insights
    if insights:
        if y < 45 * mm:
            _new_page()
        c.setFont('Helvetica-Bold', 11)
        c.drawString(margin, y, 'Insights')
        y -= 6 * mm
        c.setStrokeColor(colors.HexColor('#e5e7eb'))
        c.line(margin, y, width - margin, y)
        y -= 6 * mm
        c.setFont('Helvetica', 9)
        for it in (insights or [])[:10]:
            title = str(it.get('title') or '').strip()
            detail = str(it.get('detail') or it.get('description') or '').strip()
            s = (title + ': ' + detail).strip() if title else detail
            if not s:
                continue
            for ln in _wrap_text(s, 'Helvetica', 9, (width - margin) - margin):
                if y < 18 * mm:
                    _new_page()
                    c.setFont('Helvetica', 9)
                c.drawString(margin, y, ln)
                y -= 5.0 * mm
            y -= 1.0 * mm

    # Sub-analyses summary (top)
    try:
        nm = sub.get('no_movement') if isinstance(sub.get('no_movement'), list) else []
        br = sub.get('high_rotation_low_stock') if isinstance(sub.get('high_rotation_low_stock'), list) else []
        ov = sub.get('overstock_low_sales') if isinstance(sub.get('overstock_low_sales'), list) else []
        if nm or br or ov:
            if y < 45 * mm:
                _new_page()
            c.setFont('Helvetica-Bold', 11)
            c.drawString(margin, y, 'Alertas (Top)')
            y -= 6 * mm
            c.setStrokeColor(colors.HexColor('#e5e7eb'))
            c.line(margin, y, width - margin, y)
            y -= 6 * mm
            c.setFont('Helvetica', 9)

            def _list(title, items):
                nonlocal y
                if not items:
                    return
                if y < 30 * mm:
                    _new_page()
                    c.setFont('Helvetica', 9)
                c.setFont('Helvetica-Bold', 9)
                c.drawString(margin, y, title)
                y -= 5.0 * mm
                c.setFont('Helvetica', 9)
                for it in items[:8]:
                    if y < 18 * mm:
                        break
                    s = str(it.get('label') or '—')
                    c.drawString(margin, y, '• ' + s[:78])
                    y -= 4.6 * mm
                y -= 1.0 * mm

            _list('Sin movimiento', nm)
            _list('Alta rotación + stock bajo', br)
            _list('Acumulación', ov)
    except Exception:
        pass

    c.showPage()
    c.save()
    buf.seek(0)
    filename = f"Inventario_{p_from}_a_{p_to}.pdf" if p_from and p_to else 'Inventario.pdf'
    return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)


@bp.post('/api/inventory_rotation/export/excel')
@login_required
@module_required('reports')
def inventory_rotation_export_excel_api():
    payload = request.get_json(silent=True) or {}
    inv = payload.get('inventory') if isinstance(payload.get('inventory'), dict) else {}

    period = inv.get('period') if isinstance(inv.get('period'), dict) else {}
    k = inv.get('kpis') if isinstance(inv.get('kpis'), dict) else {}
    rows = inv.get('rows') if isinstance(inv.get('rows'), list) else []
    insights = inv.get('insights') if isinstance(inv.get('insights'), list) else []
    sub = inv.get('sub') if isinstance(inv.get('sub'), dict) else {}

    if not k or not rows:
        return jsonify({'ok': False, 'error': 'invalid_payload'}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({
            'ok': False,
            'error': 'openpyxl_missing',
            'message': 'No se pudo exportar a Excel porque falta la dependencia openpyxl. Instalá las dependencias del proyecto (requirements.txt) o ejecutá: pip install openpyxl==3.1.2',
        }), 400

    business_name, _ = _get_business_info()
    p_from = str(period.get('from') or '')
    p_to = str(period.get('to') or '')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumen'
    header_fill = PatternFill('solid', fgColor='0D1067')
    sub_fill = PatternFill('solid', fgColor='F3F4F6')

    ws['A1'] = business_name
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=14)
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:F1')
    ws['A2'] = 'Inventario - Rotación de stock'
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:F2')
    ws['A3'] = 'Período'
    ws['B3'] = f"{p_from} a {p_to}"
    ws['A4'] = 'Generado'
    ws['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M')

    ws.append([])
    ws.append(['KPI', 'Valor'])
    ws['A6'].font = Font(bold=True)
    ws['B6'].font = Font(bold=True)
    ws['A6'].fill = sub_fill
    ws['B6'].fill = sub_fill

    ws.append(['Stock total (valor)', float(_num(k.get('stock_value_total')))])
    ws.append(['Stock inmovilizado', float(_num(k.get('dead_stock_value')))])
    ws.append(['Rotación promedio', float(_num(k.get('avg_rotation')))])
    ws.append(['Días stock prom.', float(_num(k.get('avg_days_stock')))])
    ws.append(['Productos (tabla)', int(len(rows))])
    ws.append(['Productos con stock', int(k.get('products_with_stock') or 0)])

    for rr in range(7, 9):
        ws[f'B{rr}'].number_format = '"$" #,##0.00'
    ws['B9'].number_format = '0.0000'
    ws['B10'].number_format = '0'

    ws_detail = wb.create_sheet('Detalle')
    headers = ['Producto', 'Categoría', 'Unidades vendidas', 'Stock inicio', 'Stock fin', 'Stock prom.', 'Rotación', 'Días stock', 'Stock qty', 'Stock $', 'Estado']
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
        cell.fill = sub_fill
        cell.alignment = Alignment(horizontal='center')

    for r in rows[:5000]:
        ws_detail.append([
            str(r.get('label') or ''),
            str(r.get('category') or ''),
            float(_num(r.get('units_sold'))),
            float(_num(r.get('stock_start'))),
            float(_num(r.get('stock_end'))),
            float(_num(r.get('stock_avg'))),
            float(_num(r.get('rotation'))),
            (float(_num(r.get('days_stock'))) if r.get('days_stock') is not None else None),
            float(_num(r.get('stock_qty'))),
            float(_num(r.get('stock_value'))),
            str(r.get('status') or ''),
        ])

    for rr in range(2, ws_detail.max_row + 1):
        ws_detail.cell(row=rr, column=3).number_format = '0.00'
        ws_detail.cell(row=rr, column=4).number_format = '0.00'
        ws_detail.cell(row=rr, column=5).number_format = '0.00'
        ws_detail.cell(row=rr, column=6).number_format = '0.00'
        ws_detail.cell(row=rr, column=7).number_format = '0.0000'
        ws_detail.cell(row=rr, column=8).number_format = '0'
        ws_detail.cell(row=rr, column=9).number_format = '0.00'
        ws_detail.cell(row=rr, column=10).number_format = '"$" #,##0.00'
        for cc in range(1, 12):
            ws_detail.cell(row=rr, column=cc).alignment = Alignment(vertical='top')

    # Auto width
    for col in range(1, len(headers) + 1):
        max_len = 10
        for rr in range(1, min(ws_detail.max_row, 350) + 1):
            v = ws_detail.cell(row=rr, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)) if len(str(v)) < 60 else 60)
        ws_detail.column_dimensions[get_column_letter(col)].width = min(44, max_len + 2)

    if insights:
        ws_i = wb.create_sheet('Insights')
        ws_i.append(['Título', 'Detalle', 'Severidad'])
        for cell in ws_i[1]:
            cell.font = Font(bold=True)
            cell.fill = sub_fill
        for it in (insights or [])[:30]:
            ws_i.append([
                str(it.get('title') or ''),
                str(it.get('detail') or it.get('description') or ''),
                str(it.get('severity') or ''),
            ])
        ws_i.column_dimensions['A'].width = 32
        ws_i.column_dimensions['B'].width = 90
        ws_i.column_dimensions['C'].width = 12
        for rr in range(2, ws_i.max_row + 1):
            ws_i.cell(row=rr, column=2).alignment = Alignment(wrap_text=True, vertical='top')

    try:
        nm = sub.get('no_movement') if isinstance(sub.get('no_movement'), list) else []
        br = sub.get('high_rotation_low_stock') if isinstance(sub.get('high_rotation_low_stock'), list) else []
        ov = sub.get('overstock_low_sales') if isinstance(sub.get('overstock_low_sales'), list) else []
        if nm or br or ov:
            ws_s = wb.create_sheet('Subanálisis')
            ws_s.append(['Tipo', 'Producto', 'Categoría', 'Stock $', 'Días stock', 'Rotación', 'Unidades vendidas'])
            for cell in ws_s[1]:
                cell.font = Font(bold=True)
                cell.fill = sub_fill
            def _add_items(kind, items):
                for it in (items or [])[:150]:
                    ws_s.append([
                        kind,
                        str(it.get('label') or ''),
                        str(it.get('category') or ''),
                        float(_num(it.get('stock_value'))),
                        (float(_num(it.get('days_stock'))) if it.get('days_stock') is not None else None),
                        float(_num(it.get('rotation'))),
                        float(_num(it.get('units_sold'))),
                    ])
            _add_items('Sin movimiento', nm)
            _add_items('Alta rotación + stock bajo', br)
            _add_items('Acumulación', ov)
            for rr in range(2, ws_s.max_row + 1):
                ws_s.cell(row=rr, column=4).number_format = '"$" #,##0.00'
                ws_s.cell(row=rr, column=5).number_format = '0'
                ws_s.cell(row=rr, column=6).number_format = '0.0000'
                ws_s.cell(row=rr, column=7).number_format = '0.00'
            ws_s.column_dimensions['A'].width = 28
            ws_s.column_dimensions['B'].width = 44
            ws_s.column_dimensions['C'].width = 22
            ws_s.column_dimensions['D'].width = 16
            ws_s.column_dimensions['E'].width = 10
            ws_s.column_dimensions['F'].width = 12
            ws_s.column_dimensions['G'].width = 16
    except Exception:
        pass

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    filename = f"Inventario_{p_from}_a_{p_to}.xlsx" if p_from and p_to else 'Inventario.xlsx'
    return send_file(out, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
